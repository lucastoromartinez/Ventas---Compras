import re
import unicodedata
from io import BytesIO

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
# CARGA
# ─────────────────────────────────────────────

def load_excel_file(file) -> pd.DataFrame:
    return pd.read_excel(file, dtype=str)


# ─────────────────────────────────────────────
# DEPURACIÓN SISTEMA
# ─────────────────────────────────────────────

def depurar_sistema(df: pd.DataFrame) -> pd.DataFrame:
    """
    Depura el DataFrame del sistema (robusto a variaciones de nombres de columnas:
    mayúsculas/minúsculas, tildes, puntos, espacios, guiones).

    - Normaliza CUIT: elimina guiones/espacios
    - Divide Nro. en "Pto. Venta" y "N°Comprobante" (split por guion, sin ceros a izquierda)
      Maneja: sin guion, más de un guion, em/en dash, espacios alrededor del guion,
      y Pto. Venta vacío por error de carga
    - Filtra filas donde Tipo Doc. == 'Factura Gastos' (robusto)
    - Convierte columnas de importes a float
    - Dropea columnas originales CUIT y Nro. (las reales detectadas)
    """

    df = df.copy()
    df.columns = df.columns.astype(str).str.strip()

    def _norm_col(s: str) -> str:
        s = str(s)
        s = s.replace("\u00a0", " ")
        s = s.strip().lower()
        s = unicodedata.normalize("NFKD", s)
        s = "".join(ch for ch in s if not unicodedata.combining(ch))
        s = re.sub(r"[.\-_/]", " ", s)
        s = re.sub(r"\s+", " ", s).strip()
        return s

    norm_map: dict[str, str] = {}
    for c in df.columns:
        k = _norm_col(c)
        norm_map.setdefault(k, c)

    def _resolve(candidates: list[str], required: bool = True) -> str | None:
        for cand in candidates:
            k = _norm_col(cand)
            if k in norm_map:
                return norm_map[k]
        if required:
            raise KeyError(
                f"depurar_sistema: no pude resolver columna. "
                f"Candidatos={candidates}. Disponibles={list(df.columns)}"
            )
        return None

    c_cuit     = _resolve(["CUIT", "Cuit", "CUIT ", "C.U.I.T", "C U I T", "CUIT/CUIL"])
    c_nro      = _resolve(["Nro.", "Nro", "Numero", "Número", "N°", "Nro comprobante", "Nro Comprobante"])
    c_tipo_doc = _resolve(["Tipo Doc.", "Tipo Doc", "Tipo Documento", "Tipo de Documento"], required=False)

    df["CUIT_norm"] = (
        df[c_cuit]
        .astype(str)
        .str.replace("-", "", regex=False)
        .str.replace(" ", "", regex=False)
        .str.strip()
        .astype(str)
    )

    def _parse_nro(s: str) -> tuple[str, str]:
        s = str(s).strip()
        s = s.replace('\u2212', '-').replace('\u2013', '-')  # em dash / en dash
        s = re.sub(r'\s*-\s*', '-', s)
        count = s.count('-')
        if count == 0:
            left, right = s[:4], s[4:]
        elif count == 1:
            left, right = s.split('-', 1)
        else:
            idx   = s.index('-')
            left  = s[:idx]
            right = s[idx + 1:].replace('-', '')
        pto  = left.lstrip('0')
        comp = right.lstrip('0')
        if pto == '' and len(right) >= 3 and right[0] != '0' and right[1:3] == '00':
            pto  = right[0]
            comp = right[1:].lstrip('0')
        return pto, comp

    parsed = df[c_nro].astype(str).str.strip().apply(_parse_nro)

    df["Pto. Venta"]    = parsed.apply(lambda x: x[0]).astype(str)
    df["N°Comprobante"] = parsed.apply(lambda x: x[1]).astype(str)

    importes_aliases: dict[str, list[str]] = {
        "Imp. Neto Gravado":    ["Imp. Neto Gravado", "Imp Neto Gravado", "Neto Gravado", "Neto Grav"],
        "Imp. Neto No Gravado": ["Imp. Neto No Gravado", "Imp Neto No Gravado", "Neto No Gravado", "No Gravado"],
        "IVA 10,5%":            ["IVA 10,5%", "IVA 10.5%", "IVA 10,5", "IVA 10.5", "IVA 10"],
        "IVA 21%":              ["IVA 21%", "IVA 21", "IVA21"],
        "IVA 27%":              ["IVA 27%", "IVA 27", "IVA27"],
        "Imp. Int.":            ["Imp. Int.", "Imp Int", "Impuestos Internos", "Imp Internos"],
        "Perc. Gcias.":         ["Perc. Gcias.", "Perc Gcias", "Percepcion Ganancias", "Perc. Ganancias"],
        "Perc. IVA":            ["Perc. IVA", "Perc IVA", "Percepcion IVA"],
        "Perc. IIBB CABA":      ["Perc. IIBB CABA", "Perc IIBB CABA", "Percep IIBB CABA", "IIBB CABA"],
        "Perc. IIBB BS AS":     ["Perc. IIBB BS AS", "Perc IIBB BS AS", "Perc. IIBB Bs As", "IIBB BS AS", "IIBB Buenos Aires"],
        "Perc. SUSS":           ["Perc. SUSS", "Perc SUSS", "Percepcion SUSS", "SUSS"],
        "SIRCREB":              ["SIRCREB", "Sircreb"],
        "Total":                ["Total", "Importe Total", "Total Comprobante", "Total Factura"],
    }

    col_importes_reales: dict[str, str] = {}
    for canon, aliases in importes_aliases.items():
        col = _resolve(aliases, required=False)
        if col is not None:
            col_importes_reales[canon] = col

    for canon, col in col_importes_reales.items():
        df[col] = pd.to_numeric(df[col], errors="coerce").astype(float)

    if c_tipo_doc is not None:
        tipo_norm = df[c_tipo_doc].astype(str).str.strip().str.lower()
        df = df[~tipo_norm.eq("factura gastos")].copy()

    df = df.drop(columns=[c_cuit, c_nro], errors="ignore")

    return df


# ─────────────────────────────────────────────
# DEPURACIÓN ARCA
# ─────────────────────────────────────────────

def depurar_arca(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    df.columns = df.columns.str.strip()

    df = df.rename(columns={
        "Punto de Venta": "Pto. Venta",
        "NÃºmero Desde":  "N°Comprobante",
    })

    df = df.drop(columns=["NÃºmero Hasta"], errors="ignore")

    df["Nro. Doc. Emisor"] = df["Nro. Doc. Emisor"].astype(str).str.strip()

    columnas_importe = [
        "Imp. Neto Gravado IVA 0%",
        "IVA 2,5%",
        "Imp. Neto Gravado IVA 2,5%",
        "IVA 5%",
        "Imp. Neto Gravado IVA 5%",
        "IVA 10,5%",
        "Imp. Neto Gravado IVA 10,5%",
        "IVA 21%",
        "Imp. Neto Gravado IVA 21%",
        "IVA 27%",
        "Imp. Neto Gravado IVA 27%",
        "Imp. Neto Gravado Total",
        "Imp. Neto No Gravado",
        "Imp. Op. Exentas",
        "Otros Tributos",
        "Total IVA",
        "Imp. Total",
    ]
    columnas_importe = [c for c in columnas_importe if c in df.columns]

    df[columnas_importe] = (
        df[columnas_importe]
        .apply(pd.to_numeric, errors="coerce")
        .astype(float)
    )

    if (
        "Tipo de Comprobante" in df.columns
        and "Imp. Neto No Gravado" in df.columns
        and "Imp. Total" in df.columns
    ):
        tipo_tmp = pd.to_numeric(df["Tipo de Comprobante"].astype(str).str.strip(), errors="coerce")
        mask_111213 = tipo_tmp.isin([11, 12, 13])

        mask_fill = mask_111213 & (
            df["Imp. Neto No Gravado"].isna() | (df["Imp. Neto No Gravado"] == 0)
        )

        df.loc[mask_fill, "Imp. Neto No Gravado"] = df.loc[mask_fill, "Imp. Total"]

    if "Tipo Cambio" in df.columns:
        df["Tipo Cambio"] = pd.to_numeric(df["Tipo Cambio"], errors="coerce")

        mask_tc = df["Tipo Cambio"].notna() & (df["Tipo Cambio"] != 1)

        for col in columnas_importe:
            df.loc[mask_tc, col] = df.loc[mask_tc, col] * df.loc[mask_tc, "Tipo Cambio"]

    df["Tipo de Comprobante"] = pd.to_numeric(
        df["Tipo de Comprobante"].astype(str).str.strip(),
        errors="coerce"
    )

    mask_nc = df["Tipo de Comprobante"].isin([3, 8, 13, 53, 58, 63])
    df.loc[mask_nc, columnas_importe] *= -1

    if "Imp. Neto Gravado IVA 0%" in df.columns:
        df["Imp. Neto Gravado IVA 0%"] = pd.to_numeric(df["Imp. Neto Gravado IVA 0%"], errors="coerce")

    if "Imp. Neto No Gravado" in df.columns:
        df["Imp. Neto No Gravado"] = pd.to_numeric(df["Imp. Neto No Gravado"], errors="coerce")

    return df


# ─────────────────────────────────────────────
# CRUCE 1: por Pto. Venta + N°Comprobante + CUIT
# ─────────────────────────────────────────────

def cruce1(df_arca_dep: pd.DataFrame, df_sistema_dep: pd.DataFrame, tolerancia_total: float = 1.0):
    sis  = df_sistema_dep.copy().reset_index(drop=True)
    arca = df_arca_dep.copy().reset_index(drop=True)

    sis.columns  = sis.columns.str.strip()
    arca.columns = arca.columns.str.strip()

    sis["_idx_sis"]   = sis.index
    arca["_idx_arca"] = arca.index

    sis["_total_sis"]   = pd.to_numeric(sis["Total"],       errors="coerce")
    arca["_total_arca"] = pd.to_numeric(arca["Imp. Total"], errors="coerce")

    # Join many-to-many por Pto. Venta + N°Comprobante + CUIT
    cand = pd.merge(
        sis,
        arca,
        left_on  =["Pto. Venta", "N°Comprobante", "CUIT_norm"],
        right_on =["Pto. Venta", "N°Comprobante", "Nro. Doc. Emisor"],
        how      ="inner",
        suffixes =("_sis", "_arca"),
    )

    # Filtrar por tolerancia de importe total
    cand["_diff_total"] = (cand["_total_sis"] - cand["_total_arca"]).abs()
    cand_ok = cand[cand["_diff_total"] <= tolerancia_total].sort_values("_diff_total")

    # Resolver 1-a-1: entre duplicados de la misma clave, tomar el par con menor diferencia de importe
    usados_sis, usados_arca = set(), set()
    matched_sis_idx, matched_arca_idx = [], []

    for _, row in cand_ok.iterrows():
        i_sis  = int(row["_idx_sis"])
        i_arca = int(row["_idx_arca"])
        if i_sis not in usados_sis and i_arca not in usados_arca:
            usados_sis.add(i_sis)
            usados_arca.add(i_arca)
            matched_sis_idx.append(i_sis)
            matched_arca_idx.append(i_arca)

    sis_rename = {
        "Fecha":               "Fecha_Sistema",
        "Pto. Venta":          "Pto. Venta_sistema",
        "N°Comprobante":       "N°Comprobante_sistema",
        "CUIT_norm":           "cuit_sistema",
        "Imp. Neto Gravado":   "Gravado_sistema",
        "Imp. Neto No Gravado":"No Gravado_sistema",
        "Total":               "Imp. Total_sistema",
    }

    arca_cols_rename = {
        "Fecha de EmisiÃ³n":      "Fecha_arca",
        "Pto. Venta":             "Pto. Venta_arca",
        "N°Comprobante":          "N°Comprobante_arca",
        "Nro. Doc. Emisor":       "cuit_arca",
        "Imp. Neto No Gravado":   "No gravado_arca",
        "Imp. Neto Gravado Total":"Gravado_arca",
        "Imp. Op. Exentas":       "Exentas_arca",
        "Otros Tributos":         "Otros Tributos_arca",
        "Imp. Total":             "Imp. Total_arca",
    }

    arca_cols_available = [c for c in arca_cols_rename if c in arca.columns]

    temp_cols_sis  = ["_idx_sis",  "_total_sis"]
    temp_cols_arca = ["_idx_arca", "_total_arca"]

    # ---------------------------------------------------------------
    # Duplicados exactos: misma clave (Pto. Venta + N°Comprobante + CUIT)
    # Y mismo importe repetidos más de una vez de un mismo lado (la misma
    # factura cargada dos veces). La copia "sobrante" (la que no ganó el
    # match 1 a 1 de arriba) nunca va a encontrar contraparte en la otra
    # fuente -esta ya la "gastó" la primera copia-, así que terminaría
    # en falta_arca/falta_sistema como si fuera un comprobante realmente
    # faltante. Se manda a "revisar" con comentario "Duplicado" en vez de
    # a "falta".
    #
    # El caso de una Nota de Crédito cargada reutilizando el N° de la
    # factura (mismo Pto. Venta+N°Comprobante+CUIT, importe con signo
    # inverso) NO se trata acá: si el respaldo real está en ARCA, lo
    # termina resolviendo cruce3 por CUIT+Fecha+importe; si no está,
    # queda en falta_arca legítimamente (es una diferencia real, no un
    # artefacto de carga duplicada).
    # ---------------------------------------------------------------
    def _detectar_duplicados(df_full, key_cols, idx_col, total_col, usados):
        totales = df_full.set_index(idx_col)[total_col]
        filas, usados_extra = [], set()
        for idxs in df_full.groupby(key_cols)[idx_col].apply(list):
            if len(idxs) < 2:
                continue
            for i in idxs:
                if i in usados or i in usados_extra:
                    continue
                otros = [j for j in idxs if j != i]
                if any(abs(totales[i] - totales[j]) <= tolerancia_total for j in otros):
                    fila = df_full.loc[[i]].copy()
                    fila["comentario"] = "Duplicado"
                    filas.append(fila)
                    usados_extra.add(i)
        if filas:
            return pd.concat(filas, ignore_index=True), usados_extra
        vacio = df_full.iloc[0:0].copy()
        vacio["comentario"] = pd.Series(dtype="object")
        return vacio, usados_extra

    dup_sis, extra_usados_sis = _detectar_duplicados(
        sis, ["Pto. Venta", "N°Comprobante", "CUIT_norm"], "_idx_sis", "_total_sis", usados_sis
    )
    usados_sis |= extra_usados_sis

    dup_arca, extra_usados_arca = _detectar_duplicados(
        arca, ["Pto. Venta", "N°Comprobante", "Nro. Doc. Emisor"], "_idx_arca", "_total_arca", usados_arca
    )
    usados_arca |= extra_usados_arca

    revisar_duplicados = pd.concat([
        dup_sis.drop(columns=temp_cols_sis, errors="ignore").rename(columns=sis_rename),
        dup_arca[[c for c in arca_cols_available if c in dup_arca.columns] + ["comentario"]]
            .rename(columns=arca_cols_rename),
    ], ignore_index=True)

    match_sis = (
        sis.loc[matched_sis_idx]
        .drop(columns=temp_cols_sis, errors="ignore")
        .rename(columns=sis_rename)
        .reset_index(drop=True)
    )
    match_arca = (
        arca.loc[matched_arca_idx, arca_cols_available]
        .drop(columns=temp_cols_arca, errors="ignore")
        .rename(columns=arca_cols_rename)
        .reset_index(drop=True)
    )
    match = pd.concat([match_sis, match_arca], axis=1)

    falta_arca = (
        sis[~sis["_idx_sis"].isin(usados_sis)]
        .drop(columns=temp_cols_sis, errors="ignore")
        .rename(columns=sis_rename)
        .reset_index(drop=True)
    )

    falta_sistema = (
        arca[~arca["_idx_arca"].isin(usados_arca)]
        .drop(columns=temp_cols_arca, errors="ignore")
        .rename(columns=arca_cols_rename)
        .reset_index(drop=True)
    )

    # "revisar" se arma acá mismo (inconsistencias del match, con los
    # casos de composición Gravado/No Gravado ya separados a
    # "no_gravado" dentro de revisar_inconsistencias_en_match, + los
    # duplicados detectados en este cruce); cruce2 y cruce3 le van
    # agregando filas a "revisar".
    revisar_base, no_gravado = revisar_inconsistencias_en_match(match, tol_pesos=tolerancia_total)
    revisar = pd.concat([revisar_base, revisar_duplicados], ignore_index=True)

    return match, falta_sistema, falta_arca, revisar, no_gravado


# ─────────────────────────────────────────────
# REVISAR INCONSISTENCIAS EN MATCH
# ─────────────────────────────────────────────

# Columnas del sistema que pueden estar "escondiendo" un importe dentro
# de No Gravado en vez de discriminarlo aparte (percepciones/alícuotas
# que ARCA agrupa en Otros Tributos). Orden = prioridad de chequeo.
PERCEP_COLS_NO_GRAVADO = [
    "Perc. IIBB CABA", "Perc. IIBB BS AS", "Perc. SUSS", "Perc. Gcias.", "Perc. IVA",
    "SIRCREB", "Imp. Int.", "IVA 10,5%", "IVA 21%", "IVA 27%",
]


def revisar_inconsistencias_en_match(
    match: pd.DataFrame,
    tol_pesos: float = 1.0,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Arma `revisar` desde `match` cuando NO coinciden:
      - Fecha_Sistema vs Fecha_arca                → comparación exacta
      - Gravado_sistema vs Gravado_arca            → tolerancia +/- tol_pesos
      - No Gravado_sistema vs No gravado_arca      → tolerancia +/- tol_pesos.
        Si no coincide directo, "No Gravado" no se marca como
        inconsistencia si se explica por composición (no es una
        diferencia real, solo cambia cómo se reparte entre columnas):
          a) coincide con Otros Tributos_arca
          b) coincide con Exentas_arca
          c) IVA 0%: el sistema no tiene columna de IVA 0% y lo carga
             junto con No Gravado -> (Gravado_arca - Gravado_sistema)
             == (No Gravado_sistema - No gravado_arca). Esto también
             explica el Gravado (mismo movimiento).
          d) alguna columna de PERCEP_COLS_NO_GRAVADO: el sistema carga
             esa percepción junto con No Gravado en vez de
             discriminarla -> No Gravado_sistema + percepción ==
             Otros Tributos_arca.
        Las filas explicadas por (a)/(b)/(c)/(d) no van a "revisar":
        van a "no_gravado", con el motivo puntual en el comentario
        más cualquier otro motivo que la fila tenga sin explicar (ej.
        "Iva 0% registrado como no gravado, Fecha").
      - Imp. Total_sistema vs Imp. Total_arca      → tolerancia +/- tol_pesos

    Regla extra:
      - Si el motivo es SOLO "No Gravado" y Tipo de Comprobante == 11, NO incluir (Factura C)
      - Si Tipo de Comprobante no está en match, la exclusión se omite sin romper.

    Agrega `comentario` con: "Fecha" / "Gravado" / "No Gravado" / "Total".
    Pueden combinarse: "Fecha, Gravado, No Gravado, Total".

    Devuelve (revisar, no_gravado).
    """

    df = match.copy()
    df.columns = df.columns.astype(str).str.strip()

    def _pick(df_: pd.DataFrame, candidates: list[str]) -> str:
        cols = set(df_.columns)
        for c in candidates:
            if c in cols:
                return c
        raise KeyError(
            f"revisar_inconsistencias_en_match: no encontré columna. "
            f"Candidatos={candidates}. Disponibles={list(df_.columns)}"
        )

    def _pick_optional(df_: pd.DataFrame, candidates: list[str]) -> str | None:
        cols = set(df_.columns)
        for c in candidates:
            if c in cols:
                return c
        return None

    def _to_dt_norm(series: pd.Series) -> pd.Series:
        return pd.to_datetime(series, errors="coerce").dt.normalize()

    def _to_num(series: pd.Series) -> pd.Series:
        return pd.to_numeric(series, errors="coerce")

    c_fecha_sis   = _pick(df, ["Fecha_Sistema"])
    c_fecha_arca  = _pick(df, ["Fecha_arca"])

    c_grav_sis    = _pick(df, ["Gravado_sistema"])
    c_grav_arca   = _pick(df, ["Gravado_arca"])

    c_nograv_sis    = _pick(df, ["No Gravado_sistema"])
    c_nograv_arca   = _pick(df, ["No gravado_arca"])
    c_otros_trib    = _pick_optional(df, ["Otros Tributos_arca"])
    c_exentas_arca  = _pick_optional(df, ["Exentas_arca"])

    c_total_sis   = _pick(df, ["Imp. Total_sistema"])
    c_total_arca  = _pick(df, ["Imp. Total_arca"])

    c_tipo_comp   = _pick_optional(df, ["Tipo de Comprobante"])

    tol = float(tol_pesos)

    mask_fecha = ~(
        _to_dt_norm(df[c_fecha_sis]) == _to_dt_norm(df[c_fecha_arca])
    )

    grav_sis  = _to_num(df[c_grav_sis ]).fillna(0.0).round(2)
    grav_arca = _to_num(df[c_grav_arca]).fillna(0.0).round(2)
    mask_grav_raw = (grav_sis - grav_arca).abs() > tol

    nograv_sis  = _to_num(df[c_nograv_sis ]).fillna(0.0).round(2)
    nograv_arca = _to_num(df[c_nograv_arca]).fillna(0.0).round(2)

    total_sis  = _to_num(df[c_total_sis ]).fillna(0.0).round(2)
    total_arca = _to_num(df[c_total_arca]).fillna(0.0).round(2)
    mask_total = (total_sis - total_arca).abs() > tol

    if c_tipo_comp is not None:
        tipo_comp    = pd.to_numeric(df[c_tipo_comp], errors="coerce")
        mask_excluir = tipo_comp == 11
    else:
        mask_excluir = pd.Series(False, index=df.index)

    match_directo = (nograv_sis - nograv_arca).abs() <= tol

    # Motivo por el que "No Gravado" no coincide directo pero no es una
    # diferencia real (composición) -> a "no_gravado", no a "revisar".
    etiqueta_no_gravado = pd.Series(pd.NA, index=df.index, dtype="object")
    explica_gravado      = pd.Series(False, index=df.index)

    pendiente = (~match_directo) & (~mask_excluir)

    if c_otros_trib is not None:
        otros_trib = _to_num(df[c_otros_trib]).fillna(0.0).round(2)
        ok = pendiente & ((nograv_sis - otros_trib).abs() <= tol)
        etiqueta_no_gravado.loc[ok] = "No Gravado coincide con Otros Tributos"
        pendiente &= ~ok

    if c_exentas_arca is not None:
        exentas_arca = _to_num(df[c_exentas_arca]).fillna(0.0).round(2)
        ok = pendiente & ((nograv_sis - exentas_arca).abs() <= tol)
        etiqueta_no_gravado.loc[ok] = "No Gravado coincide con Exentas"
        pendiente &= ~ok

    diff_gravado   = grav_arca - grav_sis
    diff_nogravado = nograv_sis - nograv_arca
    ok = pendiente & ((diff_gravado - diff_nogravado).abs() <= tol)
    etiqueta_no_gravado.loc[ok] = "Iva 0% registrado como no gravado"
    explica_gravado.loc[ok] = True
    pendiente &= ~ok

    if c_otros_trib is not None:
        for col in PERCEP_COLS_NO_GRAVADO:
            if col not in df.columns or not pendiente.any():
                continue
            percep = _to_num(df[col]).fillna(0.0).round(2)
            ok = pendiente & ((nograv_sis + percep - otros_trib).abs() <= tol)
            etiqueta_no_gravado.loc[ok] = f"{col} registrada junto con No Gravado (coincide con Otros Tributos)"
            pendiente &= ~ok

    mask_nograv_explicado = etiqueta_no_gravado.notna()
    mask_nograv = pendiente
    mask_grav   = mask_grav_raw & ~explica_gravado

    def _comentario_row(i, incluir_no_gravado_flag: bool) -> str:
        motivos = []
        if bool(mask_fecha.loc[i]):  motivos.append("Fecha")
        if bool(mask_grav.loc[i]):   motivos.append("Gravado")
        if incluir_no_gravado_flag and bool(mask_nograv.loc[i]): motivos.append("No Gravado")
        if bool(mask_total.loc[i]):  motivos.append("Total")
        return ", ".join(motivos)

    mask_any = (mask_fecha | mask_grav | mask_nograv | mask_total) & ~mask_nograv_explicado
    revisar  = df.loc[mask_any].copy()
    revisar["comentario"] = (
        [_comentario_row(i, incluir_no_gravado_flag=True) for i in revisar.index]
        if not revisar.empty
        else pd.Series(dtype="object")
    )

    no_gravado = df.loc[mask_nograv_explicado].copy()
    no_gravado["comentario"] = (
        [
            ", ".join(
                [etiqueta_no_gravado.loc[i]]
                + [m for m in _comentario_row(i, incluir_no_gravado_flag=False).split(", ") if m]
            )
            for i in no_gravado.index
        ]
        if not no_gravado.empty
        else pd.Series(dtype="object")
    )

    return revisar.reset_index(drop=True), no_gravado.reset_index(drop=True)


# ─────────────────────────────────────────────
# CRUCE 2: faltantes por N°Comprobante + CUIT (sin Pto. Venta)
# ─────────────────────────────────────────────

def cruce2(revisar1: pd.DataFrame, falta_arca: pd.DataFrame, falta_sistema: pd.DataFrame):
    """
    Segundo cruce: cruza falta_arca y falta_sistema por N°Comprobante + CUIT (sin Pto. Venta).

    - Los que matchean se sacan de los faltantes y se agregan a revisar con comentario "Pto. Venta".
    - Los que no matchean permanecen en sus respectivos faltantes.
    """

    fa = falta_arca.copy().reset_index(drop=True)
    fs = falta_sistema.copy().reset_index(drop=True)

    fa.columns = fa.columns.str.strip()
    fs.columns = fs.columns.str.strip()

    fa["_idx_fa"] = fa.index
    fs["_idx_fs"] = fs.index

    fa["_conteo"] = fa.groupby(["N°Comprobante_sistema", "cuit_sistema"]).cumcount() + 1
    fs["_conteo"] = fs.groupby(["N°Comprobante_arca",    "cuit_arca"   ]).cumcount() + 1

    merge_df = pd.merge(
        fa,
        fs,
        left_on  =["N°Comprobante_sistema", "cuit_sistema", "_conteo"],
        right_on =["N°Comprobante_arca",    "cuit_arca",    "_conteo"],
        how      ="outer",
        indicator=True,
        suffixes =("_fa", "_fs"),
    )

    match_raw = merge_df[merge_df["_merge"] == "both"].copy()

    arca_cols_match = [
        "Fecha_arca", "Pto. Venta_arca", "N°Comprobante_arca",
        "cuit_arca", "No gravado_arca", "Gravado_arca",
        "Exentas_arca", "Otros Tributos_arca", "Imp. Total_arca",
    ]

    if not match_raw.empty:
        fa_idx = match_raw["_idx_fa"].astype(int).values
        fs_idx = match_raw["_idx_fs"].astype(int).values

        fa_match = (
            fa.loc[fa_idx]
            .drop(columns=["_conteo", "_idx_fa"], errors="ignore")
            .reset_index(drop=True)
        )

        arca_cols_available = [c for c in arca_cols_match if c in fs.columns]
        fs_match = (
            fs.loc[fs_idx, arca_cols_available]
            .reset_index(drop=True)
        )

        new_match = pd.concat([fa_match, fs_match], axis=1)
        new_match["comentario"] = "Pto. Venta"

        revisar = pd.concat([revisar1, new_match], ignore_index=True)
    else:
        revisar = revisar1.copy()

    matched_fa_idx = set(match_raw["_idx_fa"].dropna().astype(int))
    matched_fs_idx = set(match_raw["_idx_fs"].dropna().astype(int))

    falta_arca_new = (
        fa.loc[~fa["_idx_fa"].isin(matched_fa_idx)]
        .drop(columns=["_conteo", "_idx_fa"], errors="ignore")
        .reset_index(drop=True)
    )

    falta_sistema_new = (
        fs.loc[~fs["_idx_fs"].isin(matched_fs_idx)]
        .drop(columns=["_conteo", "_idx_fs"], errors="ignore")
        .reset_index(drop=True)
    )

    return revisar, falta_arca_new, falta_sistema_new


# ─────────────────────────────────────────────
# CRUCE 3: tolerante a errores en N°Comprobante o CUIT
# ─────────────────────────────────────────────

def cruce3(
    revisar: pd.DataFrame,
    falta_arca: pd.DataFrame,
    falta_sistema: pd.DataFrame,
    tol_pesos: float = 1.0,
    preferir_match_minimo: bool = True,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Tercer cruce (tolerante a errores en N°Comprobante o CUIT) entre faltantes.

    A) Primero busca por N°Comprobante + Fecha + importes (tol)
       => falló el CUIT en cruces anteriores → comentario "CUIT"
    B) Luego, sobre el remanente, busca por CUIT + Fecha + importes (tol)
       => falló el N°Comprobante → comentario "N°Comprobante"
    """

    def _to_date(series: pd.Series) -> pd.Series:
        return pd.to_datetime(series, errors="coerce").dt.date

    def _to_amount(series: pd.Series) -> pd.Series:
        return pd.to_numeric(series, errors="coerce").round(2)

    def _align_columns(base: pd.DataFrame, add: pd.DataFrame) -> pd.DataFrame:
        add2 = add.copy()
        for c in base.columns:
            if c not in add2.columns:
                add2[c] = np.nan
        extras = [c for c in add2.columns if c not in base.columns]
        return add2[base.columns.tolist() + extras]

    def _resolver_1a1(candidatos_ok: pd.DataFrame) -> tuple[pd.DataFrame, set[int], set[int]]:
        usados_fa: set[int] = set()
        usados_fs: set[int] = set()
        seleccionados = []
        for _, row in candidatos_ok.iterrows():
            id_fa = int(row["_id_fa"])
            id_fs = int(row["_id_fs"])
            if (id_fa not in usados_fa) and (id_fs not in usados_fs):
                usados_fa.add(id_fa)
                usados_fs.add(id_fs)
                seleccionados.append(row)
        return pd.DataFrame(seleccionados).copy(), usados_fa, usados_fs

    fa = falta_arca.copy().reset_index(drop=True)
    fs = falta_sistema.copy().reset_index(drop=True)
    fa.columns = fa.columns.str.strip()
    fs.columns = fs.columns.str.strip()

    fa["_id_fa"] = np.arange(len(fa), dtype=int)
    fs["_id_fs"] = np.arange(len(fs), dtype=int)

    fa["_nro_key"]  = fa["N°Comprobante_sistema"].astype(str).str.strip()
    fs["_nro_key"]  = fs["N°Comprobante_arca"].astype(str).str.strip()

    fa["_cuit_key"] = (
        fa["cuit_sistema"].astype(str).str.strip()
        .str.replace("-", "", regex=False).str.replace(" ", "", regex=False)
    )
    fs["_cuit_key"] = (
        fs["cuit_arca"].astype(str).str.strip()
        .str.replace("-", "", regex=False).str.replace(" ", "", regex=False)
    )

    fa["_fecha_key"] = _to_date(fa["Fecha_Sistema"])
    fs["_fecha_key"] = _to_date(fs["Fecha_arca"])

    fa["_ng_key"]  = _to_amount(fa["Gravado_sistema"  ]).fillna(0.0)
    fa["_nng_key"] = _to_amount(fa["No Gravado_sistema"]).fillna(0.0)
    fs["_ng_key"]  = _to_amount(fs["Gravado_arca"     ]).fillna(0.0)
    fs["_nng_key"] = _to_amount(fs["No gravado_arca"  ]).fillna(0.0)

    arca_cols_match     = ["Fecha_arca", "Pto. Venta_arca", "N°Comprobante_arca",
                           "cuit_arca", "No gravado_arca", "Gravado_arca",
                           "Exentas_arca", "Otros Tributos_arca", "Imp. Total_arca"]
    arca_cols_available = [c for c in arca_cols_match if c in fs.columns]

    def _build_match_rows(resolved: pd.DataFrame, comentario_val: str) -> pd.DataFrame:
        if resolved.empty:
            return pd.DataFrame()
        fa_idx = resolved["_id_fa"].astype(int).values
        fs_idx = resolved["_id_fs"].astype(int).values
        temp_cols = [c for c in fa.columns if c.startswith("_")]
        fa_match = (
            fa.loc[fa_idx]
            .drop(columns=temp_cols, errors="ignore")
            .reset_index(drop=True)
        )
        fs_match = (
            fs.loc[fs_idx, arca_cols_available]
            .reset_index(drop=True)
        )
        result = pd.concat([fa_match, fs_match], axis=1)
        result["comentario"] = comentario_val
        return result

    # A) N°Comprobante + Fecha + importes → "CUIT"
    cand_nro = pd.merge(
        fa, fs,
        on=["_nro_key", "_fecha_key"],
        how="inner",
        suffixes=("_fa", "_fs"),
        validate="many_to_many",
    )

    match_nro_rows = pd.DataFrame()
    usados_fa_nro: set[int] = set()
    usados_fs_nro: set[int] = set()

    if not cand_nro.empty:
        cand_nro["_diff_ng"]    = (cand_nro["_ng_key_fa"]  - cand_nro["_ng_key_fs"]).abs()
        cand_nro["_diff_nng"]   = (cand_nro["_nng_key_fa"] - cand_nro["_nng_key_fs"]).abs()
        cand_nro["_diff_total"] = cand_nro["_diff_ng"] + cand_nro["_diff_nng"]

        cand_nro_ok = cand_nro[
            (cand_nro["_diff_ng"]  <= tol_pesos) &
            (cand_nro["_diff_nng"] <= tol_pesos)
        ].copy()

        if not cand_nro_ok.empty and preferir_match_minimo:
            cand_nro_ok = cand_nro_ok.sort_values(
                by=["_nro_key", "_fecha_key", "_diff_total", "_diff_ng", "_diff_nng", "_id_fa", "_id_fs"],
                ascending=True,
            )

        if not cand_nro_ok.empty:
            resolved, usados_fa_nro, usados_fs_nro = _resolver_1a1(cand_nro_ok)
            match_nro_rows = _build_match_rows(resolved, "CUIT")

    fa_rem = fa[~fa["_id_fa"].isin(usados_fa_nro)].copy()
    fs_rem = fs[~fs["_id_fs"].isin(usados_fs_nro)].copy()

    # B) CUIT + Fecha + importes → "N°Comprobante"
    cand_cuit = pd.merge(
        fa_rem, fs_rem,
        on=["_cuit_key", "_fecha_key"],
        how="inner",
        suffixes=("_fa", "_fs"),
        validate="many_to_many",
    )

    match_cuit_rows = pd.DataFrame()
    usados_fa_cuit: set[int] = set()
    usados_fs_cuit: set[int] = set()

    if not cand_cuit.empty:
        cand_cuit["_diff_ng"]    = (cand_cuit["_ng_key_fa"]  - cand_cuit["_ng_key_fs"]).abs()
        cand_cuit["_diff_nng"]   = (cand_cuit["_nng_key_fa"] - cand_cuit["_nng_key_fs"]).abs()
        cand_cuit["_diff_total"] = cand_cuit["_diff_ng"] + cand_cuit["_diff_nng"]

        cand_cuit_ok = cand_cuit[
            (cand_cuit["_diff_ng"]  <= tol_pesos) &
            (cand_cuit["_diff_nng"] <= tol_pesos)
        ].copy()

        if not cand_cuit_ok.empty and preferir_match_minimo:
            cand_cuit_ok = cand_cuit_ok.sort_values(
                by=["_cuit_key", "_fecha_key", "_diff_total", "_diff_ng", "_diff_nng", "_id_fa", "_id_fs"],
                ascending=True,
            )

        if not cand_cuit_ok.empty:
            resolved, usados_fa_cuit, usados_fs_cuit = _resolver_1a1(cand_cuit_ok)
            match_cuit_rows = _build_match_rows(resolved, "N°Comprobante")

    falta_arca_new = (
        fa_rem[~fa_rem["_id_fa"].isin(usados_fa_cuit)]
        .drop(columns=[c for c in fa_rem.columns if c.startswith("_")], errors="ignore")
        .reset_index(drop=True)
    )
    falta_sistema_new = (
        fs_rem[~fs_rem["_id_fs"].isin(usados_fs_cuit)]
        .drop(columns=[c for c in fs_rem.columns if c.startswith("_")], errors="ignore")
        .reset_index(drop=True)
    )

    nuevos = [df for df in [match_nro_rows, match_cuit_rows] if not df.empty]

    if not nuevos:
        revisar_new = revisar.copy()
    else:
        revisar_2         = pd.concat(nuevos, ignore_index=True)
        revisar_2_aligned = _align_columns(revisar, revisar_2)
        revisar_new       = pd.concat([revisar, revisar_2_aligned], ignore_index=True)

    return revisar_new, falta_arca_new, falta_sistema_new


# ─────────────────────────────────────────────
# NETEAR FALTA_SISTEMA: comprobantes del mismo proveedor
# que se cancelan entre sí (mismo criterio que la fórmula de Excel)
# ─────────────────────────────────────────────

def netear_falta_sistema(
    falta_sistema: pd.DataFrame,
    tol_pesos: float = 1.0,
) -> pd.DataFrame:
    """
    Dentro de falta_sistema, agrupa por proveedor (cuit_arca) y busca pares
    de comprobantes cuyos Imp. Total_arca se cancelan entre sí (suma ~0,
    tolerancia +/- tol_pesos). Esos pares se sacan del faltante: si se
    netean, no representan una diferencia real y no hace falta buscarlos.

    Resolución 1 a 1 por proveedor: cada fila se usa como máximo en un par.
    """

    df = falta_sistema.copy().reset_index(drop=True)
    df.columns = df.columns.astype(str).str.strip()

    c_cuit  = "cuit_arca"
    c_total = "Imp. Total_arca"

    df["_id"]    = df.index
    df["_total"] = pd.to_numeric(df[c_total], errors="coerce").fillna(0.0).round(2)

    usados: set[int] = set()

    for _, grupo in df.groupby(c_cuit):
        idxs    = grupo["_id"].tolist()
        totales = grupo.set_index("_id")["_total"]

        for a in range(len(idxs)):
            id_a = idxs[a]
            if id_a in usados:
                continue
            for b in range(a + 1, len(idxs)):
                id_b = idxs[b]
                if id_b in usados:
                    continue
                if abs(totales[id_a] + totales[id_b]) <= tol_pesos:
                    usados.add(id_a)
                    usados.add(id_b)
                    break

    falta_sistema_new = (
        df[~df["_id"].isin(usados)]
        .drop(columns=["_id", "_total"], errors="ignore")
        .reset_index(drop=True)
    )

    return falta_sistema_new


# ─────────────────────────────────────────────
# EXPORTAR A BUFFER EN MEMORIA (descargable único)
# ─────────────────────────────────────────────

COLUMNAS_IMPORTE = [
    "Imp. Neto Gravado IVA 0%", "IVA 2,5%", "Imp. Neto Gravado IVA 2,5%",
    "IVA 5%", "Imp. Neto Gravado IVA 5%", "IVA 10,5%", "Imp. Neto Gravado IVA 10,5%",
    "IVA 21%", "Imp. Neto Gravado IVA 21%", "IVA 27%", "Imp. Neto Gravado IVA 27%",
    "Imp. Neto Gravado Total", "Imp. Neto No Gravado", "Imp. Op. Exentas",
    "Otros Tributos", "Total IVA", "Imp. Total",
    "Gravado_sistema", "Gravado_arca", "No Gravado_sistema", "No gravado_arca",
    "Otros Tributos_arca", "Exentas_arca", "Imp. Total_sistema", "Imp. Total_arca",
]


def _forzar_importes_float(df: pd.DataFrame, columnas: list[str] = COLUMNAS_IMPORTE) -> pd.DataFrame:
    df = df.copy()
    for col in columnas:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").astype(float)
    return df


def generar_excel_en_memoria(
    revisar: pd.DataFrame,
    falta_sistema: pd.DataFrame,
) -> bytes:
    DATE_FORMAT   = "DD/MM/YYYY"
    AMOUNT_FORMAT = "#,##0.00"

    ordered_rev = [
        "Fecha_Sistema", "Fecha_arca",
        "Razón Social",
        "Pto. Venta_sistema",    "Pto. Venta_arca",
        "N°Comprobante_sistema", "N°Comprobante_arca",
        "Gravado_sistema",       "Gravado_arca",
        "No Gravado_sistema",    "No gravado_arca",
        "Otros Tributos_arca",   "Exentas_arca",
        "Imp. Total_sistema",    "Imp. Total_arca",
        "comentario",
    ]
    rev_out = _forzar_importes_float(
        revisar[[c for c in ordered_rev if c in revisar.columns]]
    )
    rev_date_cols = ["Fecha_Sistema", "Fecha_arca"]

    fs_out       = _forzar_importes_float(falta_sistema)
    fs_date_cols = ["Fecha_arca"]

    def _prep_df(df: pd.DataFrame, date_col_names: list[str]) -> tuple[pd.DataFrame, list[int], list[int]]:
        df = df.copy()
        date_col_indices   = []
        amount_col_indices = []
        for i, col in enumerate(df.columns, start=1):
            if col in date_col_names:
                df[col] = pd.to_datetime(df[col], errors="coerce")
                date_col_indices.append(i)
            elif col in COLUMNAS_IMPORTE:
                amount_col_indices.append(i)
        return df, date_col_indices, amount_col_indices

    rev_prep, rev_date_idx, rev_amount_idx = _prep_df(rev_out, rev_date_cols)
    fs_prep,  fs_date_idx,  fs_amount_idx  = _prep_df(fs_out,  fs_date_cols)

    def _style_sheet(ws, date_col_indices: list[int], amount_col_indices: list[int]) -> None:
        thin          = Side(style="thin")
        header_fill   = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        header_font   = Font(bold=True)
        header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.border    = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx in date_col_indices:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = DATE_FORMAT

        for col_idx in amount_col_indices:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = AMOUNT_FORMAT

        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0 for cell in col),
                default=10,
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        rev_prep.to_excel(writer, sheet_name="revisar",       index=False)
        fs_prep.to_excel( writer, sheet_name="falta_sistema", index=False)

        wb = writer.book
        _style_sheet(wb["revisar"],       rev_date_idx, rev_amount_idx)
        _style_sheet(wb["falta_sistema"], fs_date_idx, fs_amount_idx)

    return buf.getvalue()


# ─────────────────────────────────────────────
# PIPELINE COMPLETO
# ─────────────────────────────────────────────

def correr_cruce(archivo_arca, archivo_sistema, tol_pesos: float = 1.0):
    df_arca    = load_excel_file(archivo_arca)
    df_sistema = load_excel_file(archivo_sistema)

    df_arca_dep    = depurar_arca(df_arca)
    df_sistema_dep = depurar_sistema(df_sistema)

    match, falta_sistema1, falta_arca1, revisar1, no_gravado = cruce1(
        df_arca_dep, df_sistema_dep, tolerancia_total=tol_pesos
    )

    revisar2, falta_arca2, falta_sistema2 = cruce2(revisar1, falta_arca1, falta_sistema1)

    revisar3, falta_arca3, falta_sistema3 = cruce3(
        revisar2, falta_arca2, falta_sistema2, tol_pesos=tol_pesos
    )

    falta_sistema_final = netear_falta_sistema(falta_sistema3, tol_pesos=tol_pesos)

    stats = {
        "match":             len(match),
        "revisar":           len(revisar3),
        "duplicados":        int((revisar1["comentario"] == "Duplicado").sum()),
        "no_gravado":        len(no_gravado),
        "faltante_arca":     len(falta_arca3),
        "faltante_sistema":  len(falta_sistema_final),
    }

    buf_reporte = generar_excel_en_memoria(revisar3, falta_sistema_final)

    return buf_reporte, stats
