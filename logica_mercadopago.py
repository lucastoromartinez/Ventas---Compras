"""
logica_mercadopago.py — Conciliación Mercado Pago

Adaptado del notebook de conciliación de Mercado Pago. Dos cruces:

  1) cruce_mayor_detalle: Mayor de Mercado Pago (mayor_mp) vs Detalle de
     Mercado Pago (detalle_mp). Soporta opcionalmente el Detalle del mes
     ANTERIOR (para acreditaciones de fin de mes que se registraron recién
     en el mayor del mes en curso) y el switch '9dD' (separa las
     transacciones de "meitre"/"irondriver" del Detalle para buscarlas
     contra las Acreditaciones del Mayor de Recaudación en vez del Mayor
     de Mercado Pago).
  2) cruce_comisiones_impuestos: compara Comisiones e Impuestos (totales)
     entre ambos Mayores y el Detalle.

  generar_df_impuestos: abre el detalle de Impuestos por categoría
  (Sistema vs Mercado) para inspección manual.

  correr_conciliacion_mercadopago(): pipeline de entrada para la app
  (Streamlit) — recibe los archivos subidos, corre los dos cruces y
  devuelve un Excel en memoria + estadísticas para mostrar en pantalla.
"""

import json
import unicodedata
from io import BytesIO

import pandas as pd
from openpyxl.styles import Font


# ─────────────────────────────────────────────
# CARGA
# ─────────────────────────────────────────────

def load_excel_file(archivo) -> pd.DataFrame:
    return pd.read_excel(archivo)


# ─────────────────────────────────────────────
# NORMALIZACIÓN DE TEXTO
# ─────────────────────────────────────────────

def _normalizar(texto):
    """Mayúsculas, sin tildes ni espacios, para comparar texto sin errores de formato."""
    if pd.isna(texto):
        return ''
    texto = str(texto).upper().strip()
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII')
    texto = texto.replace(' ', '')
    return texto


def _normalizar_puntos(texto):
    """Mayúsculas, sin tildes ni puntos (para Comentario, que puede traer 'MP 06.2026')."""
    if pd.isna(texto):
        return ''
    texto = str(texto).upper().strip()
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII')
    texto = texto.replace('.', '')
    return texto


TIPOS_EXCLUIDOS_ACREDITACION = ['PAYOUTS', 'SETTLEMENT', 'REFUND']

# Traducción de las claves de 'financial_entity' que trae Mercado Pago en la
# columna IMPUESTOS DESAGREGADOS a un nombre legible. Si en otro mes aparece
# una clave nueva (otra provincia, por ejemplo), se muestra tal cual viene
# (sin traducir) para no perder el dato.
ETIQUETAS_CATEGORIA = {
    'caba': 'IIBB CABA',
    'debitos_creditos': 'Impuesto Débitos/Créditos',
}


# ─────────────────────────────────────────────
# DEPURACIÓN
# ─────────────────────────────────────────────

def depurar_mayores(df: pd.DataFrame) -> pd.DataFrame:
    """
    Depura un DataFrame de Mayor contable:
    - Elimina la fila con Comentario == 'Saldo anterior'.
    - Redondea Debe y Haber a 2 decimales (tratando NaN como 0, ya que en
      el mayor cada movimiento suele tener solo uno de los dos cargado).
    - Calcula Importe = Debe - Haber (redondeado a 2 decimales).
    """
    df = df.copy()
    df.columns = [c.strip() for c in df.columns]

    df = df[df['Comentario'] != 'Saldo anterior'].reset_index(drop=True)

    df['Debe'] = df['Debe'].fillna(0).round(2)
    df['Haber'] = df['Haber'].fillna(0).round(2)
    df['Importe'] = (df['Debe'] - df['Haber']).round(2)

    return df


def depurar_mercado(df: pd.DataFrame) -> pd.DataFrame:
    """
    Depura el detalle de Mercado Pago:
    - Dropea filas completamente vacías (basura al final del Excel).
    - Convierte a datetime64[ns] toda columna cuyo nombre contenga 'FECHA'.
    - Redondea a 2 decimales toda columna numérica.
    - Respeta columnas booleanas (ej. 'LIQUIDADO') sin convertirlas.
    """
    df = df.copy()
    df.columns = [c.strip() for c in df.columns]

    df = df.dropna(how='all').reset_index(drop=True)

    date_cols = [c for c in df.columns if 'FECHA' in c.upper()]
    for c in date_cols:
        df[c] = (
            pd.to_datetime(df[c], errors='coerce')
              .dt.tz_localize(None)
              .astype('datetime64[ns]')
        )

    for c in df.columns:
        if c in date_cols:
            continue
        if pd.api.types.is_float_dtype(df[c]) or pd.api.types.is_integer_dtype(df[c]):
            df[c] = df[c].round(2)
        elif df[c].dtype == object:
            non_null = df[c].dropna()
            if non_null.empty:
                continue
            if non_null.apply(lambda x: isinstance(x, bool)).any():
                continue
            converted = pd.to_numeric(df[c], errors='coerce')
            if converted.notna().sum() / len(non_null) > 0.9:
                df[c] = converted.round(2)

    return df


def _clave_local(fuente_detalle_sin_tipos_excluidos):
    """
    A partir de un detalle ya filtrado (sin PAYOUTS/SETTLEMENT/REFUND), arma
    la columna _clave_local: NOMBRE DE LOCAL, y para las filas sin local usa
    DETALLE DE LA VENTA normalizado como clave alternativa (con el caso
    especial de Meitre/Irondriver, que van juntos bajo 'MEITRE_IRONDRIVER').
    """
    no_payout = fuente_detalle_sin_tipos_excluidos.copy()
    no_payout['_clave_local'] = no_payout['NOMBRE DE LOCAL']
    sin_local = no_payout['NOMBRE DE LOCAL'].isna()
    detalle_venta_norm = no_payout.loc[sin_local, 'DETALLE DE LA VENTA'].apply(_normalizar)
    es_meitre = detalle_venta_norm.str.contains('MEITRE', na=False) | detalle_venta_norm.str.contains('IRONDRIVER', na=False)
    no_payout.loc[sin_local, '_clave_local'] = detalle_venta_norm.where(~es_meitre, 'MEITRE_IRONDRIVER')
    return no_payout


# ─────────────────────────────────────────────
# CRUCE 1: MAYOR × DETALLE
# ─────────────────────────────────────────────

def cruce_mayor_detalle(mayor_mp: pd.DataFrame, detalle_mp: pd.DataFrame, tolerancia: float = 1.0,
                         detalle_mp_anterior: pd.DataFrame = None,
                         switch: str = None, mayor_rec: pd.DataFrame = None):
    """
    Primer cruce entre el Mayor de Mercado Pago (mayor_mp, ya depurado con
    depurar_mayores) y el Detalle de Mercado Pago (detalle_mp, ya depurado
    con depurar_mercado).

    detalle_mp_anterior : opcional, el Detalle de Mercado Pago del MES
        ANTERIOR (también ya depurado con depurar_mercado). Se usa solo para
        el paso de acreditaciones: si alguna línea de "Acreditación" del mes
        en curso no matchea contra detalle_mp, se reintenta el mismo
        mecanismo (agrupar por local/clave, tolerancia +/- $1) pero contra
        detalle_mp_anterior.

    switch : opcional. Si no se aclara (None), el cruce funciona de forma
        normal (todo contra mayor_mp). Si se pasa '9dD' (sin importar
        mayúsculas), las transacciones de detalle_mp cuya columna DETALLE DE
        LA VENTA contiene la palabra "meitre" se separan del resto, no se
        buscan contra mayor_mp, y en cambio se suman y se buscan contra las
        líneas de "Acreditación" de mayor_rec (tolerancia +/- $1). El resto
        del proceso opera exactamente igual que en el modo normal.
    mayor_rec : requerido solo cuando switch='9dD' (el segundo mayor, ya
        depurado con depurar_mayores).

    Devuelve
    --------
    match_mayor, match_mp, falta_mayor, falta_mp, acreditacion_pendiente,
    info (dict con diagnóstico: switch usado, mes detectado, cantidad de
    matches, etc. — para mostrar en la UI en vez de imprimir por consola).
    """

    switch_norm = (switch or '').strip().upper()
    es_9dd = switch_norm == '9DD'
    advertencias = []
    if es_9dd and mayor_rec is None:
        advertencias.append("Switch '9dD' solicitado pero no se pasó Mayor de Recaudación. Se procede en modo normal.")
        es_9dd = False

    mayor = mayor_mp.copy()
    mayor.columns = [c.strip() for c in mayor.columns]

    # Se ignoran del cruce 1 las líneas del mayor que sean Comisión o
    # Liquidación (sin importar mayúsculas/tildes/espacios): esos conceptos
    # se comparan aparte en cruce_comisiones_impuestos.
    _comentario_norm_tmp = mayor['Comentario'].apply(_normalizar)
    mask_excluir_mayor = (
        _comentario_norm_tmp.str.contains('COMISION', na=False)
        | _comentario_norm_tmp.str.contains('LIQUIDACION', na=False)
    )
    mayor = mayor.loc[~mask_excluir_mayor].reset_index(drop=True)
    mayor['id_mayor'] = mayor.index

    detalle = detalle_mp.copy()
    detalle.columns = [c.strip() for c in detalle.columns]
    detalle = detalle.reset_index(drop=True)
    detalle['id_mp'] = detalle.index

    mayor['matcheado'] = False
    detalle['matcheado'] = False

    mayor['_comentario_norm'] = mayor['Comentario'].apply(_normalizar)
    detalle['_tipo_norm'] = detalle['TIPO DE OPERACIÓN'].apply(_normalizar)
    detalle['_detalle_venta_norm'] = detalle['DETALLE DE LA VENTA'].apply(_normalizar)
    es_meitre_row = detalle['_detalle_venta_norm'].str.contains('MEITRE', na=False)

    detalle_anterior = None
    if detalle_mp_anterior is not None:
        detalle_anterior = detalle_mp_anterior.copy()
        detalle_anterior.columns = [c.strip() for c in detalle_anterior.columns]
        detalle_anterior = detalle_anterior.reset_index(drop=True)
        detalle_anterior['_tipo_norm'] = detalle_anterior['TIPO DE OPERACIÓN'].apply(_normalizar)

    match_id_counter = 0
    matches = []
    filas_extra_match_mp = []
    filas_extra_match_mayor = []

    # ---------------------------------------------------------------
    # 1) ACREDITACIONES
    # ---------------------------------------------------------------
    es_acred_mayor = mayor['_comentario_norm'].str.contains('ACREDITACION', na=False)

    periodos_mayor = pd.to_datetime(mayor['Fecha']).dt.to_period('M')
    periodos_validos = periodos_mayor.dropna()
    mes_actual = periodos_validos.mode().iloc[0] if len(periodos_validos) else None
    mes_siguiente = mes_actual + 1 if mes_actual is not None else None

    if mes_siguiente is not None:
        es_mes_siguiente = periodos_mayor == mes_siguiente
    else:
        es_mes_siguiente = pd.Series(False, index=mayor.index)

    mask_pendiente = es_acred_mayor & es_mes_siguiente
    acreditacion_pendiente = mayor.loc[mask_pendiente].drop(
        columns=[c for c in ['matcheado', '_comentario_norm'] if c in mayor.columns]
    ).copy()
    mayor.loc[mask_pendiente, 'matcheado'] = True

    mayor_acred_idx = mayor.index[es_acred_mayor & (~mayor['matcheado'])].tolist()

    mask_base_no_payout = ~detalle['_tipo_norm'].isin(TIPOS_EXCLUIDOS_ACREDITACION)
    if es_9dd:
        mask_base_no_payout = mask_base_no_payout & (~es_meitre_row)

    if len(mayor_acred_idx) == 1:
        no_payout = detalle[mask_base_no_payout]
        total = round(no_payout['VALOR DE LA COMPRA'].sum(), 2)
        idx_mayor = mayor_acred_idx[0]
        importe_mayor = mayor.at[idx_mayor, 'Importe']
        if abs(total - importe_mayor) <= tolerancia:
            match_id_counter += 1
            mayor.at[idx_mayor, 'matcheado'] = True
            for id_mp in no_payout['id_mp']:
                detalle.loc[detalle['id_mp'] == id_mp, 'matcheado'] = True
                matches.append({'match_id': match_id_counter, 'id_mayor': idx_mayor,
                                 'id_mp': id_mp, 'grupo': 'TOTAL'})

    elif len(mayor_acred_idx) > 1:
        no_payout = _clave_local(detalle[mask_base_no_payout])

        acred_local = (
            no_payout.groupby('_clave_local', dropna=False)['VALOR DE LA COMPRA']
            .sum().round(2).reset_index()
        )
        disponibles = list(mayor_acred_idx)
        for _, fila in acred_local.iterrows():
            clave = fila['_clave_local']
            total_local = fila['VALOR DE LA COMPRA']
            if not disponibles:
                break
            candidatos = mayor.loc[disponibles]
            diffs = (candidatos['Importe'] - total_local).abs()
            diffs_ok = diffs[diffs <= tolerancia]
            if len(diffs_ok) > 0:
                idx_mayor = diffs_ok.idxmin()
                disponibles.remove(idx_mayor)
                match_id_counter += 1
                mayor.at[idx_mayor, 'matcheado'] = True
                ids_mp_clave = no_payout.loc[no_payout['_clave_local'] == clave, 'id_mp']
                for id_mp in ids_mp_clave:
                    detalle.loc[detalle['id_mp'] == id_mp, 'matcheado'] = True
                    matches.append({'match_id': match_id_counter, 'id_mayor': idx_mayor,
                                     'id_mp': id_mp, 'grupo': str(clave)})

        if len(disponibles) == len(mayor_acred_idx):
            total_mayor = round(mayor.loc[mayor_acred_idx, 'Importe'].sum(), 2)
            total_mp = round(no_payout['VALOR DE LA COMPRA'].sum(), 2)
            if abs(total_mayor - total_mp) <= tolerancia:
                match_id_counter += 1
                for idx_mayor in mayor_acred_idx:
                    mayor.at[idx_mayor, 'matcheado'] = True
                    matches.append({'match_id': match_id_counter, 'id_mayor': idx_mayor,
                                     'id_mp': None, 'grupo': 'TOTAL'})
                for id_mp in no_payout['id_mp']:
                    detalle.loc[detalle['id_mp'] == id_mp, 'matcheado'] = True
                    matches.append({'match_id': match_id_counter, 'id_mayor': None,
                                     'id_mp': id_mp, 'grupo': 'TOTAL'})

    # ---- Fallback mes anterior ----
    disponibles_anterior = [idx for idx in mayor_acred_idx if not mayor.at[idx, 'matcheado']]
    if detalle_anterior is not None and disponibles_anterior:
        no_payout_ant = _clave_local(detalle_anterior[~detalle_anterior['_tipo_norm'].isin(TIPOS_EXCLUIDOS_ACREDITACION)])

        if len(disponibles_anterior) == 1:
            total_ant = round(no_payout_ant['VALOR DE LA COMPRA'].sum(), 2)
            idx_mayor = disponibles_anterior[0]
            importe_mayor = mayor.at[idx_mayor, 'Importe']
            if abs(total_ant - importe_mayor) <= tolerancia:
                match_id_counter += 1
                mayor.at[idx_mayor, 'matcheado'] = True
                matches.append({'match_id': match_id_counter, 'id_mayor': idx_mayor,
                                 'id_mp': None, 'grupo': 'TOTAL (MES ANTERIOR)'})
                extra = no_payout_ant.drop(columns=['_clave_local'], errors='ignore').copy()
                extra['match_id'] = match_id_counter
                extra['grupo'] = 'TOTAL (MES ANTERIOR)'
                filas_extra_match_mp.append(extra)

        else:
            acred_local_ant = (
                no_payout_ant.groupby('_clave_local', dropna=False)['VALOR DE LA COMPRA']
                .sum().round(2).reset_index()
            )
            disponibles_ant = list(disponibles_anterior)
            for _, fila in acred_local_ant.iterrows():
                clave = fila['_clave_local']
                total_local = fila['VALOR DE LA COMPRA']
                if not disponibles_ant:
                    break
                candidatos = mayor.loc[disponibles_ant]
                diffs = (candidatos['Importe'] - total_local).abs()
                diffs_ok = diffs[diffs <= tolerancia]
                if len(diffs_ok) > 0:
                    idx_mayor = diffs_ok.idxmin()
                    disponibles_ant.remove(idx_mayor)
                    match_id_counter += 1
                    mayor.at[idx_mayor, 'matcheado'] = True
                    matches.append({'match_id': match_id_counter, 'id_mayor': idx_mayor,
                                     'id_mp': None, 'grupo': f'{clave} (MES ANTERIOR)'})
                    extra = no_payout_ant.loc[no_payout_ant['_clave_local'] == clave].drop(
                        columns=['_clave_local'], errors='ignore'
                    ).copy()
                    extra['match_id'] = match_id_counter
                    extra['grupo'] = f'{clave} (MES ANTERIOR)'
                    filas_extra_match_mp.append(extra)

            if len(disponibles_ant) == len(disponibles_anterior):
                total_mayor_ant = round(mayor.loc[disponibles_anterior, 'Importe'].sum(), 2)
                total_mp_ant = round(no_payout_ant['VALOR DE LA COMPRA'].sum(), 2)
                if abs(total_mayor_ant - total_mp_ant) <= tolerancia:
                    match_id_counter += 1
                    for idx_mayor in disponibles_anterior:
                        mayor.at[idx_mayor, 'matcheado'] = True
                        matches.append({'match_id': match_id_counter, 'id_mayor': idx_mayor,
                                         'id_mp': None, 'grupo': 'TOTAL (MES ANTERIOR)'})
                    extra = no_payout_ant.drop(columns=['_clave_local'], errors='ignore').copy()
                    extra['match_id'] = match_id_counter
                    extra['grupo'] = 'TOTAL (MES ANTERIOR)'
                    filas_extra_match_mp.append(extra)

    # ---- Switch '9dD': acreditaciones de "meitre" contra mayor_rec ----
    if es_9dd:
        no_payout_meitre = detalle[
            (~detalle['_tipo_norm'].isin(TIPOS_EXCLUIDOS_ACREDITACION)) & es_meitre_row
        ]
        if len(no_payout_meitre) > 0:
            total_meitre = round(no_payout_meitre['VALOR DE LA COMPRA'].sum(), 2)

            rec = mayor_rec.copy()
            rec.columns = [c.strip() for c in rec.columns]
            rec['_comentario_norm'] = rec['Comentario'].apply(_normalizar)
            mask_acred_rec = rec['_comentario_norm'].str.contains('ACREDITACION', na=False)
            candidatos_rec = rec.loc[mask_acred_rec]

            diffs = (candidatos_rec['Importe'] - total_meitre).abs()
            diffs_ok = diffs[diffs <= tolerancia]
            if len(diffs_ok) > 0:
                idx_rec = diffs_ok.idxmin()
                match_id_counter += 1

                fila_rec = rec.loc[[idx_rec]].drop(columns=['_comentario_norm'], errors='ignore').copy()
                fila_rec['match_id'] = match_id_counter
                fila_rec['grupo'] = 'MEITRE (REC)'
                filas_extra_match_mayor.append(fila_rec)

                for id_mp in no_payout_meitre['id_mp']:
                    detalle.loc[detalle['id_mp'] == id_mp, 'matcheado'] = True
                    matches.append({'match_id': match_id_counter, 'id_mayor': None,
                                     'id_mp': id_mp, 'grupo': 'MEITRE (REC)'})
            # Si no matchea, las filas de "meitre" quedan sin marcar y siguen
            # el curso normal (Remanentes, o terminan en falta_mayor).

    # ---------------------------------------------------------------
    # 2) REMANENTES: Fecha + Importe, y luego solo Importe
    # ---------------------------------------------------------------
    remanentes_mp = detalle[~detalle['matcheado']]

    mayor['_fecha_pura'] = pd.to_datetime(mayor['Fecha']).dt.date
    detalle['_fecha_pura'] = pd.to_datetime(detalle['FECHA DE ORIGEN']).dt.date

    for id_mp in remanentes_mp['id_mp']:
        fila_mp = detalle.loc[detalle['id_mp'] == id_mp].iloc[0]
        disponibles = mayor.index[~mayor['matcheado']]
        candidatos = mayor.loc[disponibles]
        candidatos = candidatos[candidatos['_fecha_pura'] == fila_mp['_fecha_pura']]
        diffs = (candidatos['Importe'] - fila_mp['VALOR DE LA COMPRA']).abs()
        diffs_ok = diffs[diffs <= tolerancia]
        if len(diffs_ok) > 0:
            idx_mayor = diffs_ok.idxmin()
            match_id_counter += 1
            mayor.at[idx_mayor, 'matcheado'] = True
            detalle.loc[detalle['id_mp'] == id_mp, 'matcheado'] = True
            matches.append({'match_id': match_id_counter, 'id_mayor': idx_mayor,
                             'id_mp': id_mp, 'grupo': 'FECHA_IMPORTE'})

    remanentes_mp_rest = detalle[~detalle['matcheado']]
    for id_mp in remanentes_mp_rest['id_mp']:
        fila_mp = detalle.loc[detalle['id_mp'] == id_mp].iloc[0]
        disponibles = mayor.index[~mayor['matcheado']]
        candidatos = mayor.loc[disponibles]
        diffs = (candidatos['Importe'] - fila_mp['VALOR DE LA COMPRA']).abs()
        diffs_ok = diffs[diffs <= tolerancia]
        if len(diffs_ok) > 0:
            idx_mayor = diffs_ok.idxmin()
            match_id_counter += 1
            mayor.at[idx_mayor, 'matcheado'] = True
            detalle.loc[detalle['id_mp'] == id_mp, 'matcheado'] = True
            matches.append({'match_id': match_id_counter, 'id_mayor': idx_mayor,
                             'id_mp': id_mp, 'grupo': 'SOLO_IMPORTE'})

    # ---------------------------------------------------------------
    # Armado de resultados
    # ---------------------------------------------------------------
    matches_df = pd.DataFrame(matches)

    cols_aux = ['matcheado', '_comentario_norm', '_fecha_pura', 'id_mayor']
    cols_aux_mp = ['matcheado', '_tipo_norm', '_detalle_venta_norm', '_fecha_pura', 'id_mp']

    if not matches_df.empty:
        match_mayor = mayor.merge(
            matches_df[['match_id', 'id_mayor', 'grupo']].drop_duplicates('id_mayor'),
            on='id_mayor', how='inner'
        ).drop(columns=[c for c in cols_aux if c in mayor.columns])

        match_mp = detalle.merge(
            matches_df[['match_id', 'id_mp', 'grupo']],
            on='id_mp', how='inner'
        ).drop(columns=[c for c in cols_aux_mp if c in detalle.columns])
    else:
        match_mayor = mayor.iloc[0:0].drop(columns=[c for c in cols_aux if c in mayor.columns]).assign(match_id=[], grupo=[])
        match_mp = detalle.iloc[0:0].drop(columns=[c for c in cols_aux_mp if c in detalle.columns]).assign(match_id=[], grupo=[])

    if filas_extra_match_mp:
        extra_concat = pd.concat(filas_extra_match_mp, ignore_index=True)
        extra_concat = extra_concat.drop(columns=[c for c in cols_aux_mp if c in extra_concat.columns])
        match_mp = pd.concat([match_mp, extra_concat], ignore_index=True)

    if filas_extra_match_mayor:
        extra_concat_mayor = pd.concat(filas_extra_match_mayor, ignore_index=True)
        match_mayor = pd.concat([match_mayor, extra_concat_mayor], ignore_index=True)

    falta_mayor = detalle.loc[~detalle['matcheado']].drop(columns=[c for c in cols_aux_mp if c in detalle.columns])
    falta_mp = mayor.loc[~mayor['matcheado']].drop(columns=[c for c in cols_aux if c in mayor.columns])

    info = {
        'switch': '9dD' if es_9dd else 'normal',
        'mes_actual': str(mes_actual) if mes_actual is not None else None,
        'mes_siguiente': str(mes_siguiente) if mes_siguiente is not None else None,
        'n_acreditacion_pendiente': len(acreditacion_pendiente),
        'n_matches': int(matches_df['match_id'].nunique()) if not matches_df.empty else 0,
        'matches_por_grupo': matches_df['grupo'].value_counts().to_dict() if not matches_df.empty else {},
        'advertencias': advertencias,
    }

    return match_mayor, match_mp, falta_mayor, falta_mp, acreditacion_pendiente, info


# ─────────────────────────────────────────────
# CRUCE 2: COMISIONES E IMPUESTOS
# ─────────────────────────────────────────────

def _siguiente_match_id(match_mayor: pd.DataFrame, match_mp: pd.DataFrame) -> int:
    """Retoma la numeración de match_id donde la dejó el cruce 1."""
    max_ids = []
    if 'match_id' in match_mayor.columns and len(match_mayor):
        max_ids.append(match_mayor['match_id'].max())
    if 'match_id' in match_mp.columns and len(match_mp):
        max_ids.append(match_mp['match_id'].max())
    return int(max(max_ids)) + 1 if max_ids else 1


def cruce_comisiones_impuestos(mayor_mp: pd.DataFrame, mayor_rec: pd.DataFrame, detalle_mp: pd.DataFrame,
                                match_mayor: pd.DataFrame, match_mp: pd.DataFrame,
                                falta_mayor: pd.DataFrame, falta_mp: pd.DataFrame):
    """
    Segundo cruce: compara Comisiones e Impuestos entre los Mayores contables
    y el Detalle de Mercado Pago (detalle_mp, ya depurado con depurar_mercado).

    Las Comisiones pueden estar registradas en CUALQUIERA de los dos mayores
    según la sociedad (por eso se buscan en mayor_mp y mayor_rec juntos), pero
    los Impuestos (Liquidación) siempre se registran en el Mayor de
    Recaudación (mayor_rec).

    A diferencia del primer cruce, acá no se busca matchear filas 1 a 1; se
    compara el total registrado en el sistema (mayor) contra el total real de
    Mercado Pago (detalle) para cada concepto. Si hay diferencia, se agrega
    una fila a falta_mayor/falta_mp con esa diferencia; si hay actividad de
    Impuestos, además se agrega una fila resumen a match_mayor/match_mp.

    Devuelve
    --------
    comisiones_impuestos : DataFrame con columnas Concepto, Sistema, Mercado,
                            Diferencia (Sistema - Mercado), una fila por concepto.
    match_mayor, match_mp, falta_mayor, falta_mp : los mismos recibidos, con
        las filas correspondientes agregadas.
    """

    def _prep(df):
        df = df.copy()
        df.columns = [c.strip() for c in df.columns]
        df['_comentario_norm'] = df['Comentario'].apply(_normalizar_puntos)
        return df

    mp = _prep(mayor_mp)
    rec = _prep(mayor_rec)

    detalle = detalle_mp.copy()
    detalle.columns = [c.strip() for c in detalle.columns]

    match_mayor = match_mayor.copy()
    match_mp = match_mp.copy()
    falta_mayor = falta_mayor.copy()
    falta_mp = falta_mp.copy()

    # ---------------- Comisiones (busca en ambos mayores) ----------------
    mask_comision_mp = mp['_comentario_norm'].str.contains('COMISION', na=False)
    mask_comision_rec = rec['_comentario_norm'].str.contains('COMISION', na=False)
    hay_comision = mask_comision_mp.any() or mask_comision_rec.any()
    mercado_comisiones = round(detalle['COMISIÓN DE MERCADO LIBRE + IVA'].sum(), 2)

    if hay_comision:
        sistema_comisiones_num = round(
            mp.loc[mask_comision_mp, 'Importe'].sum() + rec.loc[mask_comision_rec, 'Importe'].sum(), 2
        )
        sistema_comisiones_mostrar = sistema_comisiones_num
    else:
        sistema_comisiones_num = 0.0
        sistema_comisiones_mostrar = 'No registradas'

    diferencia_comisiones = round(sistema_comisiones_num - mercado_comisiones, 2)

    # ---------------- Impuestos (siempre en mayor_rec) ----------------
    mask_liquidacion = rec['_comentario_norm'].str.contains('LIQUIDACION', na=False)
    sistema_impuestos = round(rec.loc[mask_liquidacion, 'Importe'].sum(), 2)
    mercado_impuestos = round(detalle['IMPUESTOS COBRADOS POR RETENCIONES DE IIBB'].sum(), 2)
    diferencia_impuestos = round(sistema_impuestos - mercado_impuestos, 2)

    comisiones_impuestos = pd.DataFrame([
        {
            'Concepto': 'Comisiones',
            'Sistema': sistema_comisiones_mostrar,
            'Mercado': mercado_comisiones,
            'Diferencia': diferencia_comisiones,
        },
        {
            'Concepto': 'Impuestos',
            'Sistema': sistema_impuestos,
            'Mercado': mercado_impuestos,
            'Diferencia': diferencia_impuestos,
        },
    ])

    # ---------------- Agregar diferencias a falta_mayor / falta_mp ----------------
    filas_falta_mp = []
    filas_falta_mayor = []

    if abs(diferencia_comisiones) > 0:
        filas_falta_mp.append({'Comentario': 'Diferencia Comisiones', 'Importe': diferencia_comisiones})
        filas_falta_mayor.append({'TIPO DE OPERACIÓN': 'Diferencia Comisiones',
                                   'VALOR DE LA COMPRA': -diferencia_comisiones})

    if abs(diferencia_impuestos) > 0:
        filas_falta_mp.append({'Comentario': 'Diferencia Impuestos', 'Importe': diferencia_impuestos})
        filas_falta_mayor.append({'TIPO DE OPERACIÓN': 'Diferencia Impuestos',
                                   'VALOR DE LA COMPRA': -diferencia_impuestos})

    if filas_falta_mp:
        falta_mp = pd.concat([falta_mp, pd.DataFrame(filas_falta_mp)], ignore_index=True)
    if filas_falta_mayor:
        falta_mayor = pd.concat([falta_mayor, pd.DataFrame(filas_falta_mayor)], ignore_index=True)

    # ---------------- Agregar el match resumido de Impuestos ----------------
    if mercado_impuestos != 0 or sistema_impuestos != 0:
        nuevo_match_id = _siguiente_match_id(match_mayor, match_mp)
        fila_match_mayor = {'Comentario': 'Impuestos', 'Importe': mercado_impuestos,
                             'match_id': nuevo_match_id, 'grupo': 'IMPUESTOS'}
        fila_match_mp = {'TIPO DE OPERACIÓN': 'Impuestos', 'VALOR DE LA COMPRA': mercado_impuestos,
                          'match_id': nuevo_match_id, 'grupo': 'IMPUESTOS'}
        match_mayor = pd.concat([match_mayor, pd.DataFrame([fila_match_mayor])], ignore_index=True)
        match_mp = pd.concat([match_mp, pd.DataFrame([fila_match_mp])], ignore_index=True)

    return comisiones_impuestos, match_mayor, match_mp, falta_mayor, falta_mp


# ─────────────────────────────────────────────
# DETALLE DE IMPUESTOS POR CATEGORÍA
# ─────────────────────────────────────────────

def generar_df_impuestos(mayor_rec: pd.DataFrame, detalle_mp: pd.DataFrame):
    """
    Genera dos DataFrames para comparar Impuestos por categoría entre el
    Mayor de Recaudación (mayor_rec) y el Detalle de Mercado Pago (detalle_mp).

    impuestos_sistema (lado mayor): todas las filas del mayor cuyo
        Comentario contenga "Ret" o "Liquidacion", cada una en su propia fila.
    impuestos_mercado (lado detalle): parsea IMPUESTOS DESAGREGADOS (JSON por
        fila con financial_entity/amount) y suma por financial_entity.
    """
    mayor = mayor_rec.copy()
    mayor.columns = [c.strip() for c in mayor.columns]
    mayor['_comentario_norm'] = mayor['Comentario'].apply(_normalizar_puntos)

    detalle = detalle_mp.copy()
    detalle.columns = [c.strip() for c in detalle.columns]

    mask = (
        mayor['_comentario_norm'].str.contains('RET', na=False)
        | mayor['_comentario_norm'].str.contains('LIQUIDACION', na=False)
    )
    impuestos_sistema = (
        mayor.loc[mask, ['Comentario', 'Importe']]
        .rename(columns={'Comentario': 'Concepto'})
        .sort_values('Concepto')
        .reset_index(drop=True)
    )

    registros = []
    for val in detalle['IMPUESTOS DESAGREGADOS'].dropna():
        try:
            items = json.loads(val)
        except (json.JSONDecodeError, TypeError):
            continue
        registros.extend(items)

    if registros:
        df_reg = pd.DataFrame(registros)
        df_reg['amount'] = pd.to_numeric(df_reg['amount'], errors='coerce')
        impuestos_mercado = (
            df_reg.groupby('financial_entity')['amount']
            .sum()
            .round(2)
            .reset_index()
            .rename(columns={'financial_entity': 'Categoria', 'amount': 'Importe'})
        )
        impuestos_mercado['Categoria'] = (
            impuestos_mercado['Categoria'].map(ETIQUETAS_CATEGORIA)
            .fillna(impuestos_mercado['Categoria'])
        )
    else:
        impuestos_mercado = pd.DataFrame(columns=['Categoria', 'Importe'])

    return impuestos_sistema, impuestos_mercado


# ─────────────────────────────────────────────
# EXPORTAR EN MEMORIA
# ─────────────────────────────────────────────

def exportar_cruce(match_mayor, match_mp, falta_mayor, falta_mp, acreditacion_pendiente,
                    comisiones_impuestos, impuestos_sistema, impuestos_mercado) -> bytes:
    """
    Exporta los DataFrames del cruce a un mismo libro de Excel en memoria:
    - match_mayor, match_mp, falta_mayor, falta_mp: una hoja por DataFrame.
    - acreditacion_pendiente: hoja 'Acreditacion_Pendiente'.
    - comisiones_impuestos: hoja 'Comisiones_Impuestos'.
    - impuestos_sistema / impuestos_mercado: hoja 'Impuestos', como dos
      cuadros separados (Sistema a la izquierda, Mercado a la derecha).
    """
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        match_mayor.to_excel(writer, sheet_name='match_mayor', index=False)
        match_mp.to_excel(writer, sheet_name='match_mp', index=False)
        falta_mayor.to_excel(writer, sheet_name='falta_mayor', index=False)
        falta_mp.to_excel(writer, sheet_name='falta_mp', index=False)

        acreditacion_pendiente.to_excel(writer, sheet_name='Acreditacion_Pendiente', index=False)

        comisiones_impuestos.to_excel(writer, sheet_name='Comisiones_Impuestos', index=False)

        impuestos_sistema.to_excel(writer, sheet_name='Impuestos', index=False, startrow=1, startcol=0)
        impuestos_mercado.to_excel(writer, sheet_name='Impuestos', index=False, startrow=1, startcol=4)

        ws = writer.sheets['Impuestos']
        ws['A1'] = 'SISTEMA'
        ws['E1'] = 'MERCADO'
        ws['A1'].font = Font(bold=True)
        ws['E1'].font = Font(bold=True)
        for col in ('A', 'B', 'E', 'F'):
            ws[f'{col}2'].font = Font(bold=True)
        for col, ancho in [('A', 32), ('B', 16), ('C', 3), ('D', 3), ('E', 28), ('F', 16)]:
            ws.column_dimensions[col].width = ancho

    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
# PIPELINE COMPLETO → devuelve Excel en memoria + stats
# ─────────────────────────────────────────────

def correr_conciliacion_mercadopago(archivo_mp, archivo_mayor_mp, archivo_mayor_rec,
                                     archivo_mp_anterior=None, switch=None, tolerancia: float = 1.0):
    """
    Pipeline de entrada para la app: recibe los archivos subidos (file-like
    de Streamlit), corre los dos cruces y devuelve (buf, stats) listos para
    mostrar/descargar.

    switch: None para modo normal, o '9dD' para activar el tratamiento
        especial de Meitre/Irondriver contra el Mayor de Recaudación.
    """
    df_mp = load_excel_file(archivo_mp)
    df_mayor_mp = load_excel_file(archivo_mayor_mp)
    df_mayor_rec = load_excel_file(archivo_mayor_rec)
    df_mp_anterior = load_excel_file(archivo_mp_anterior) if archivo_mp_anterior is not None else None

    detalle_mp = depurar_mercado(df_mp)
    detalle_mp_anterior = depurar_mercado(df_mp_anterior) if df_mp_anterior is not None else None
    mayor_mp = depurar_mayores(df_mayor_mp)
    mayor_rec = depurar_mayores(df_mayor_rec)

    match_mayor, match_mp, falta_mayor, falta_mp, acreditacion_pendiente, info = cruce_mayor_detalle(
        mayor_mp, detalle_mp, tolerancia=tolerancia,
        detalle_mp_anterior=detalle_mp_anterior, switch=switch, mayor_rec=mayor_rec,
    )

    comisiones_impuestos, match_mayor, match_mp, falta_mayor, falta_mp = cruce_comisiones_impuestos(
        mayor_mp, mayor_rec, detalle_mp, match_mayor, match_mp, falta_mayor, falta_mp
    )

    impuestos_sistema, impuestos_mercado = generar_df_impuestos(mayor_rec, detalle_mp)

    buf = exportar_cruce(
        match_mayor, match_mp, falta_mayor, falta_mp, acreditacion_pendiente,
        comisiones_impuestos, impuestos_sistema, impuestos_mercado,
    )

    fila_comisiones = comisiones_impuestos.loc[comisiones_impuestos['Concepto'] == 'Comisiones'].iloc[0]
    fila_impuestos = comisiones_impuestos.loc[comisiones_impuestos['Concepto'] == 'Impuestos'].iloc[0]

    stats = {
        'switch': info['switch'],
        'mes_actual': info['mes_actual'],
        'n_matches': info['n_matches'],
        'acreditacion_pendiente': len(acreditacion_pendiente),
        'falta_mayor': len(falta_mayor),
        'falta_mp': len(falta_mp),
        'diferencia_comisiones': float(fila_comisiones['Diferencia']),
        'diferencia_impuestos': float(fila_impuestos['Diferencia']),
        'advertencias': info['advertencias'],
    }

    return buf, stats
