import calendar
from io import BytesIO

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ======================================================================
# 1) DEPURACION
# ======================================================================

def depurar_ventas_hio(df_ventas_hio: pd.DataFrame) -> pd.DataFrame:
    """
    Depura el DataFrame df_ventas_hio:
    - Elimina las filas donde 'Fecha Z' o 'Fecha ARCA' sean NaN.
    - Convierte 'Pto Venta' y 'Comprobante' de float a objeto (string).
    - Redondea a 2 decimales las columnas float restantes.
    - Genera la columna 'Local' a partir de 'Establecimiento'.
    - Genera la columna 'Tipo' a partir de 'Tipo Comprobante'.
    """
    df = df_ventas_hio.copy()

    # Eliminar filas con NaN en Fecha Z o Fecha ARCA
    df = df.dropna(subset=['Fecha Z', 'Fecha ARCA'])

    # Convertir Pto Venta y Comprobante a objeto (string, sin decimales)
    for col in ['Pto Venta', 'Comprobante']:
        if col in df.columns:
            df[col] = df[col].apply(
                lambda x: str(int(x)) if pd.notnull(x) else x
            ).astype(object)

    # Redondear a 2 decimales las columnas float restantes
    float_cols = df.select_dtypes(include=['float64', 'float32']).columns
    df[float_cols] = df[float_cols].round(2)

    # Generar columna Local a partir de Establecimiento
    def clasificar_local(establecimiento):
        if pd.isnull(establecimiento):
            return np.nan
        texto = str(establecimiento)
        if "Banda" in texto or "Islas" in texto:
            return "9dD"
        elif "Ronda" in texto:
            return "Ronda"
        elif "Avellaneda" in texto:
            return "EASA"
        elif "Aribau" in texto or "Arcos" in texto or "Maldini" in texto:
            return "ERSA"
        else:
            return np.nan

    df['Local'] = df['Establecimiento'].apply(clasificar_local)

    # Generar columna Tipo a partir de Tipo Comprobante
    def clasificar_tipo(tipo_comprobante):
        if pd.isnull(tipo_comprobante):
            return np.nan
        texto = str(tipo_comprobante)
        if texto in ["FC A", "FC B", "NC B"]:
            return "B"
        elif texto == "FC N":
            return "N"
        else:
            return np.nan

    df['Tipo'] = df['Tipo Comprobante'].apply(clasificar_tipo)

    return df


# ======================================================================
# 2) EXPORTACION
# ======================================================================

FUENTE = "Calibri"
TAMANO = 10

FONT_NORMAL = Font(name=FUENTE, size=TAMANO)
FONT_HEADER = Font(name=FUENTE, size=TAMANO, bold=True, color="FFFFFF")
FONT_TITULO = Font(name=FUENTE, size=TAMANO + 1, bold=True)
FILL_HEADER = PatternFill("solid", fgColor="305496")
FILL_TOTAL = PatternFill("solid", fgColor="D9E1F2")
BORDE_FINO = Border(*[Side(style="thin", color="BFBFBF")] * 4)

FORMATO_IMPORTE = '#,##0.00'
FORMATO_FECHA = 'dd/mm/yyyy'
FORMATO_HORA = 'hh:mm:ss'

# Orden de columnas de la tabla cruda por tipo de hoja
COLS_LOCAL = ['Fecha Z', 'Fecha ARCA', 'Hora', 'Establecimiento', 'Tipo',
              'Serie / Número', 'Tipo Comprobante', 'Pto Venta', 'Comprobante',
              'Medio Pago', 'Neto', 'IVA', 'Total']
COLS_GENERAL = [c for c in COLS_LOCAL if c != 'Tipo']

COLS_IMPORTE = ['Neto', 'IVA', 'Total']
COLS_FECHA = ['Fecha Z', 'Fecha ARCA']
COLS_TEXTO = ['Pto Venta', 'Comprobante']

# Nombre de hoja -> valor esperado en la columna 'Local'
MAPA_LOCALES = {
    'ERSA': 'ERSA',
    'EASA': 'EASA',
    'RONDA': 'Ronda',
    '9DD': '9dD',
}


def _aplicar_estilo_header(ws, fila, col_inicio, col_fin):
    for c in range(col_inicio, col_fin + 1):
        cell = ws.cell(row=fila, column=c)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDE_FINO


def _escribir_tabla_cruda(ws, df, columnas):
    """Escribe la tabla de datos cruda a partir de la fila 1, columna A."""
    presentes = [c for c in columnas if c in df.columns]

    # Encabezado
    for j, col in enumerate(presentes, start=1):
        ws.cell(row=1, column=j, value=col)
    _aplicar_estilo_header(ws, 1, 1, len(presentes))
    ws.freeze_panes = "A2"

    # Datos
    for i, (_, fila) in enumerate(df[presentes].iterrows(), start=2):
        for j, col in enumerate(presentes, start=1):
            valor = fila[col]
            if pd.isnull(valor):
                valor = None
            cell = ws.cell(row=i, column=j, value=valor)
            cell.font = FONT_NORMAL
            if col in COLS_IMPORTE:
                cell.number_format = FORMATO_IMPORTE
                cell.alignment = Alignment(horizontal="right")
            elif col == 'Fecha Z' or col == 'Fecha ARCA':
                cell.number_format = FORMATO_FECHA
            elif col == 'Hora':
                cell.number_format = FORMATO_HORA
            elif col in COLS_TEXTO:
                cell.number_format = '@'
                cell.alignment = Alignment(horizontal="left")

    # Ancho de columnas aproximado
    for j, col in enumerate(presentes, start=1):
        letra = get_column_letter(j)
        ancho = max(12, min(28, len(col) + 4))
        ws.column_dimensions[letra].width = ancho

    return presentes, len(df) + 1  # ultima fila con datos (o 1 si vacio)


def _armar_cuadro_pivot(ws, df_mes, fila_inicio, col_inicio, titulo,
                          rango_datos, letras_col, primer_dia, ultimo_dia,
                          dentro_del_mes=True):
    """
    Arma un cuadro tipo tabla dinamica (Medio Pago en filas, Tipo en columnas,
    Neto/IVA/Total sumarizados) usando formulas SUMIFS que apuntan a la tabla
    cruda de la misma hoja. dentro_del_mes=True filtra por el mes de
    referencia; False suma todo lo que queda FUERA de ese mes.
    """
    tipos = sorted([t for t in df_mes['Tipo'].dropna().unique().tolist()]) if not df_mes.empty else ['B', 'N']
    medios = sorted(df_mes['Medio Pago'].dropna().unique().tolist()) if not df_mes.empty else []

    col_medio, col_fecha_arca, col_tipo, col_neto, col_iva, col_total = letras_col
    ultima_fila_datos = rango_datos

    # Titulo
    ws.cell(row=fila_inicio, column=col_inicio, value=titulo).font = FONT_TITULO

    fila = fila_inicio + 1
    # Encabezados: Medio Pago | por cada tipo: Neto/IVA/Total | Totales generales
    ws.cell(row=fila, column=col_inicio, value="Medio Pago")
    c = col_inicio + 1
    for tipo in tipos:
        ws.cell(row=fila - 0, column=c, value=f"Tipo {tipo}")
        ws.merge_cells(start_row=fila, start_column=c, end_row=fila, end_column=c + 2)
        for k, nombre in enumerate(['Neto', 'IVA', 'Total']):
            ws.cell(row=fila + 1, column=c + k, value=nombre)
        c += 3
    ws.cell(row=fila, column=c, value="Total general")
    ws.merge_cells(start_row=fila, start_column=c, end_row=fila, end_column=c + 2)
    for k, nombre in enumerate(['Neto', 'IVA', 'Total']):
        ws.cell(row=fila + 1, column=c + k, value=nombre)

    col_fin = c + 2
    _aplicar_estilo_header(ws, fila, col_inicio, col_fin)
    _aplicar_estilo_header(ws, fila + 1, col_inicio, col_fin)

    fila_datos_ini = fila + 2
    fila_actual = fila_datos_ini

    def formula_sumifs(col_valor_letra, medio_pago_cell, tipo_valor=None):
        base = (f'=SUMIFS(${col_valor_letra}$2:${col_valor_letra}${ultima_fila_datos},'
                f'${col_medio}$2:${col_medio}${ultima_fila_datos},{medio_pago_cell}')
        if tipo_valor is not None:
            base += f',${col_tipo}$2:${col_tipo}${ultima_fila_datos},"{tipo_valor}"'
        if dentro_del_mes:
            base += (f',${col_fecha_arca}$2:${col_fecha_arca}${ultima_fila_datos},">="&DATE({primer_dia.year},{primer_dia.month},{primer_dia.day})'
                      f',${col_fecha_arca}$2:${col_fecha_arca}${ultima_fila_datos},"<="&DATE({ultimo_dia.year},{ultimo_dia.month},{ultimo_dia.day})')
            base += ')'
            return base
        else:
            # Solo lo POSTERIOR al mes de referencia (ARCA pendiente de imputar).
            # Lo anterior al mes de referencia es remanente de meses ya
            # reportados y no debe volver a sumarse aca.
            base += f',${col_fecha_arca}$2:${col_fecha_arca}${ultima_fila_datos},">"&DATE({ultimo_dia.year},{ultimo_dia.month},{ultimo_dia.day}))'
            return base

    for medio in medios:
        medio_cell = f"${get_column_letter(col_inicio)}${fila_actual}"
        ws.cell(row=fila_actual, column=col_inicio, value=medio).font = FONT_NORMAL

        c = col_inicio + 1
        for tipo in tipos:
            for col_letra in [col_neto, col_iva, col_total]:
                cell = ws.cell(row=fila_actual, column=c,
                                value=formula_sumifs(col_letra, medio_cell, tipo))
                cell.number_format = FORMATO_IMPORTE
                cell.font = FONT_NORMAL
                c += 1
        # total general (sin filtro de tipo) por medio de pago
        for col_letra in [col_neto, col_iva, col_total]:
            cell = ws.cell(row=fila_actual, column=c,
                            value=formula_sumifs(col_letra, medio_cell, tipo_valor=None))
            cell.number_format = FORMATO_IMPORTE
            cell.font = FONT_NORMAL
            c += 1
        fila_actual += 1

    fila_total = fila_actual
    ws.cell(row=fila_total, column=col_inicio, value="Total general").font = Font(name=FUENTE, size=TAMANO, bold=True)
    for c in range(col_inicio + 1, col_fin + 1):
        letra = get_column_letter(c)
        cell = ws.cell(row=fila_total, column=c,
                        value=f"=SUM({letra}{fila_datos_ini}:{letra}{fila_actual - 1})" if fila_actual > fila_datos_ini else 0)
        cell.number_format = FORMATO_IMPORTE
        cell.font = Font(name=FUENTE, size=TAMANO, bold=True)
        cell.fill = FILL_TOTAL

    for c in range(col_inicio, col_fin + 1):
        for f in range(fila, fila_total + 1):
            ws.cell(row=f, column=c).border = BORDE_FINO

    return fila_total + 1  # primera fila libre despues del cuadro


def detectar_mes_referencia(df: pd.DataFrame, col_fecha: str = 'Fecha ARCA') -> tuple[int, int]:
    """
    Detecta el mes de referencia (mes corriente del reporte) como el mes/año
    con mayor cantidad de registros en la columna col_fecha (por defecto
    'Fecha ARCA'). Devuelve una tupla (anio, mes).
    """
    fechas = df[col_fecha].dropna()
    periodos = fechas.dt.to_period('M')
    mes_mas_frecuente = periodos.value_counts().idxmax()
    return mes_mas_frecuente.year, mes_mas_frecuente.month


def exportar_ventas_hio_buffer(
    df: pd.DataFrame,
    mes_referencia: tuple[int, int] | None = None,
) -> tuple[bytes, tuple[int, int]]:
    """
    Exporta un DataFrame ya depurado (df_dep) a un Excel en memoria con:
      - Hoja 'Reporte Gral': todos los registros, todas las fechas.
      - Una hoja por Local (ERSA, EASA, RONDA, 9DD) con:
          * tabla de datos cruda (todas las fechas, importes en formato miles,
            misma tipografia/tamaño en toda la hoja)
          * cuadro tipo tabla dinamica (Medio Pago x Tipo, Neto/IVA/Total)
            SOLO para filas con Fecha ARCA dentro del mes de referencia
          * debajo, un cuadro resumen (mismo agrupamiento) con todo lo que
            quedo FUERA del mes de referencia

    Devuelve (bytes_del_excel, (anio, mes)) con el mes de referencia usado.
    """
    df = df.copy()

    if mes_referencia is None:
        anio, mes = detectar_mes_referencia(df)
    else:
        anio, mes = mes_referencia

    primer_dia = pd.Timestamp(anio, mes, 1)
    ultimo_dia = pd.Timestamp(anio, mes, calendar.monthrange(anio, mes)[1])

    wb = Workbook()
    ws_gral = wb.active
    ws_gral.title = "Reporte Gral"
    _escribir_tabla_cruda(ws_gral, df, COLS_GENERAL)

    for nombre_hoja, valor_local in MAPA_LOCALES.items():
        df_local = df[df['Local'].astype(str).str.upper() == valor_local.upper()].copy()
        ws = wb.create_sheet(nombre_hoja)

        presentes, ultima_fila = _escribir_tabla_cruda(ws, df_local, COLS_LOCAL)

        # letras de columnas relevantes dentro de la tabla cruda de esta hoja
        idx = {c: presentes.index(c) + 1 for c in presentes}
        letras_col = tuple(get_column_letter(idx[c]) for c in
                            ['Medio Pago', 'Fecha ARCA', 'Tipo', 'Neto', 'IVA', 'Total'])

        col_inicio_pivot = len(presentes) + 3

        df_mes = df_local[df_local['Fecha ARCA'].between(primer_dia, ultimo_dia)]
        df_resto = df_local[df_local['Fecha ARCA'] > ultimo_dia]

        siguiente_fila = _armar_cuadro_pivot(
            ws, df_mes, 1, col_inicio_pivot,
            f"Resumen mes {mes:02d}/{anio} (Medio Pago x Tipo)",
            ultima_fila, letras_col, primer_dia, ultimo_dia, dentro_del_mes=True)

        _armar_cuadro_pivot(
            ws, df_resto, siguiente_fila + 1, col_inicio_pivot,
            "Resumen ARCA pendiente (fecha posterior al mes de referencia)",
            ultima_fila, letras_col, primer_dia, ultimo_dia, dentro_del_mes=False)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), (anio, mes)


# ======================================================================
# 3) PIPELINE COMPLETO
# ======================================================================

def correr_registro_ventas(
    archivo_hio,
    mes_referencia: tuple[int, int] | None = None,
) -> tuple[bytes, dict]:
    df_raw = pd.read_excel(archivo_hio)
    df_dep = depurar_ventas_hio(df_raw)

    buf, (anio, mes) = exportar_ventas_hio_buffer(df_dep, mes_referencia=mes_referencia)

    stats = {
        "total":     int(len(df_dep)),
        "sin_local": int(df_dep["Local"].isna().sum()),
        "sin_tipo":  int(df_dep["Tipo"].isna().sum()),
        "anio":      anio,
        "mes":       mes,
    }

    return buf, stats
