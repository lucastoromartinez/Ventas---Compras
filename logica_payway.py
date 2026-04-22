"""
Lógica de lectura de liquidaciones Payway desde PDFs (v2)
Usa pdfplumber en lugar de PyMuPDF.
"""
import re
import io
from io import BytesIO
from typing import List

import pandas as pd


# ─────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────

# Mapeo AG.RET.ING.BRUTOS → tarjeta
# No se usa texto libre porque todos los PDFs Galicia incluyen
# "SOLICITUD DE ADHESION AL SISTEMA VISA" en el pie legal → falsos positivos
AGRET_TARJETA = {
    "100": "master",
    "900": "visa",
    "391": "cabal",
    "810": "american",
}

NUM_FINAL_RE = re.compile(r'\$\s*([\d.,]+)\s*$')
NUM_RE       = re.compile(r'(\d{1,3}(?:\.\d{3})*,\d{2})')


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def _leer_pdf(pdf_bytes: bytes) -> str:
    import pdfplumber
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        return "\n".join(p.extract_text() or "" for p in pdf.pages)


def _a_float(s: str) -> float:
    return float(s.replace(".", "").replace(",", ".").strip())


# ─────────────────────────────────────────────
# EXTRACCIÓN BASE
# ─────────────────────────────────────────────

def extraer_datos_base_pdf(pdf_bytes: bytes, nombre_archivo: str) -> dict:
    """
    Extrae: ID (tarjeta + nro resumen), Total Presentado, Total Descuentos, Total Neto.
    """
    texto  = _leer_pdf(pdf_bytes)
    lineas = texto.splitlines()
    nombre_lower = nombre_archivo.lower()

    # Tarjeta: AG.RET.ING.BRUTOS como fuente primaria
    tarjeta = ""
    m = re.search(r"AG\.RET\.ING\.BRUTOS[:\s]+([\d]+)", texto, re.IGNORECASE)
    if m:
        tarjeta = AGRET_TARJETA.get(m.group(1), "")

    # Fallback por nombre de archivo
    if not tarjeta:
        for kw in ["amex", "american", "master", "cabal", "visa"]:
            if kw in nombre_lower:
                tarjeta = "american" if kw in ("amex", "american") else kw
                break

    # Número de resumen
    nro_resumen = ""
    for patron in [
        r"n[°ºo]?\s*de\s*resumen[:\s]+([\d]+)",
        r"resumen[:\s]+([\d]+)",
    ]:
        m = re.search(patron, texto, re.IGNORECASE)
        if m:
            nro_resumen = str(int(m.group(1)))
            break

    if not nro_resumen:
        m = re.search(r"(?:amex|american|masterd?|visa|cabal)\s+(\d+)", nombre_lower)
        if m:
            nro_resumen = m.group(1)

    # Totales
    def primer_numero_ventana(desde, ventana=6):
        for l in lineas[desde: desde + ventana]:
            m2 = NUM_RE.search(l)
            if m2:
                return _a_float(m2.group(1))
        return 0.0

    idx_pres  = next((i for i, l in enumerate(lineas) if "TOTAL PRESENTADO $" in l), None)
    idx_desc  = next((i for i, l in enumerate(lineas) if "TOTAL DESCUENTO $"  in l), None)
    idx_saldo = next((i for i, l in enumerate(lineas) if re.match(r"SALDO \$", l.strip())), None)

    total_presentado = primer_numero_ventana(idx_pres + 1) if idx_pres is not None else 0.0

    total_descuentos = 0.0
    if idx_desc is not None:
        for l in lineas[idx_desc: idx_desc + 6]:
            nums = NUM_RE.findall(l)
            if nums and l.strip() != lineas[idx_desc].strip():
                total_descuentos = _a_float(nums[0])
                break

    total_neto = primer_numero_ventana(idx_saldo + 1) if idx_saldo is not None else 0.0

    return {
        "ID": f"{tarjeta}{nro_resumen}",
        "Total Presentado": total_presentado,
        "Total Descuentos": total_descuentos,
        "Total Neto": total_neto,
    }


# ─────────────────────────────────────────────
# EXTRACCIÓN DETALLE DESCUENTOS
# ─────────────────────────────────────────────

def extraer_detalle_descuentos_pdf(pdf_bytes: bytes) -> dict:
    """
    Extrae impuestos, retenciones y percepciones del bloque 'Deducciones Impositivas'.
    Columnas fijas: IVA 21%, IVA 10,5%, Retenciones Totales, Percepciones Totales,
                    Monto Gravado IVA 21%, Monto Gravado IVA 10,5%, Base Exenta
    Columnas dinámicas: prefijo R. (retenciones IIBB) o P. (percepciones/IVA)
    """
    texto = _leer_pdf(pdf_bytes)

    datos = {
        "IVA 21%": 0.0,
        "IVA 10,5%": 0.0,
        "Retenciones Totales": 0.0,
        "Percepciones Totales": 0.0,
        "Monto Gravado IVA 21%": 0.0,
        "Monto Gravado IVA 10,5%": 0.0,
        "Base Exenta": 0.0,
    }

    # IVA fijo
    m = re.search(r"IVA\s*21,00\s*%\s*\$\s*([\d.,]+)", texto, re.IGNORECASE)
    if m:
        datos["IVA 21%"] = _a_float(m.group(1))

    m = re.search(r"IVA\s*10,50?\s*%\s*Ley\s*25\.?063\s*\$\s*([\d.,]+)", texto, re.IGNORECASE)
    if m:
        datos["IVA 10,5%"] = _a_float(m.group(1))

    # Bases imponibles
    m = re.search(r"Tasa\s*21,00\s*%\s*\$\s*([\d.,]+)", texto, re.IGNORECASE)
    if m:
        datos["Monto Gravado IVA 21%"] = _a_float(m.group(1))

    m = re.search(r"Tasa\s*10,50\s*%\s*\$\s*([\d.,]+)", texto, re.IGNORECASE)
    if m:
        datos["Monto Gravado IVA 10,5%"] = _a_float(m.group(1))

    m = re.search(r"Base\s*Exenta\s*\$\s*([\d.,]+)", texto, re.IGNORECASE)
    if m:
        datos["Base Exenta"] = _a_float(m.group(1))

    # Columnas dinámicas: bloque Deducciones Impositivas
    bloque_m = re.search(
        r'Deducciones Impositivas\n(.*?)'
        r'(?:_{5,}|Continúa en|Fin de la|SE RECUERDA|ADEMAS|DE ACUERDO|SR\.\s*COMERCIANTE|\Z)',
        texto, re.DOTALL | re.IGNORECASE
    )
    if bloque_m:
        IGNORAR = re.compile(r'^\(ver|^en hoja aparte', re.IGNORECASE)
        for linea in bloque_m.group(1).splitlines():
            linea = linea.strip()
            if not linea or IGNORAR.match(linea):
                continue
            m2 = NUM_FINAL_RE.search(linea)
            if not m2:
                continue
            valor = _a_float(m2.group(1))
            desc  = re.sub(r'\s*\$\s*[\d.,]+\s*$', '', linea).strip()
            if not desc:
                continue
            if re.search(r'ret\.ib|retenc.*ib|iibb|ing\.?\s*brutos', desc, re.IGNORECASE):
                col = f"R. {desc}"
                datos[col] = datos.get(col, 0.0) + valor
                datos["Retenciones Totales"] += valor
            else:
                col = f"P. {desc}"
                datos[col] = datos.get(col, 0.0) + valor
                datos["Percepciones Totales"] += valor

    return datos


# ─────────────────────────────────────────────
# PIPELINE COMPLETO
# ─────────────────────────────────────────────

def procesar_pdfs_payway(archivos_subidos) -> bytes:
    """
    Recibe lista de archivos subidos por Streamlit, procesa cada PDF
    y devuelve un Excel con formato en memoria.
    """
    COLUMNAS_FIJAS = [
        "ID", "Total Neto", "Total Presentado", "Total Descuentos",
        "IVA 10,5%", "IVA 21%", "Retenciones Totales", "Percepciones Totales",
        "Monto Gravado IVA 21%", "Monto Gravado IVA 10,5%", "Base Exenta",
    ]

    filas             = []
    columnas_dinamicas = set()

    for archivo in archivos_subidos:
        pdf_bytes = archivo.read()
        nombre    = archivo.name

        base    = extraer_datos_base_pdf(pdf_bytes, nombre)
        detalle = extraer_detalle_descuentos_pdf(pdf_bytes)
        fila    = {**detalle, **base}
        filas.append(fila)

        for k in fila:
            if k not in COLUMNAS_FIJAS:
                columnas_dinamicas.add(k)

    columnas_finales = COLUMNAS_FIJAS + sorted(columnas_dinamicas)

    df = pd.DataFrame(filas)
    for col in columnas_finales:
        if col not in df.columns:
            df[col] = "" if col == "ID" else 0.0

    df = df[columnas_finales]
    cols_num = [c for c in df.columns if c != "ID"]
    df[cols_num] = df[cols_num].fillna(0.0)

    fila_total = {"ID": "TOTAL"}
    for col in cols_num:
        fila_total[col] = round(df[col].sum(), 2)

    df = pd.concat([df, pd.DataFrame([fila_total])], ignore_index=True)

    # Generar Excel con formato
    buf = BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)

    # Aplicar estilos
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(buf)
    ws = wb.active

    THIN  = Side(border_style="thin", color="AAAAAA")
    BORDE = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    n_rows = ws.max_row
    n_cols = ws.max_column

    # Encabezado azul
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.fill      = PatternFill("solid", fgColor="1F4E79")
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = BORDE
    ws.row_dimensions[1].height = 40

    # Filas de datos
    for row in range(2, n_rows + 1):
        is_total = ws.cell(row=row, column=1).value == "TOTAL"
        for col in range(1, n_cols + 1):
            c = ws.cell(row=row, column=col)
            c.border = BORDE
            if is_total:
                c.fill = PatternFill("solid", fgColor="D6E4F0")
                c.font = Font(name="Arial", bold=True, size=10)
            else:
                c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(
                horizontal="center" if col == 1 else "right",
                vertical="center"
            )
            if col > 1 and c.value is not None:
                c.number_format = "#,##0.00"

    # Anchos de columna
    ws.column_dimensions["A"].width = 22
    for col in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    ws.freeze_panes = "A2"

    buf_out = BytesIO()
    wb.save(buf_out)
    return buf_out.getvalue()
