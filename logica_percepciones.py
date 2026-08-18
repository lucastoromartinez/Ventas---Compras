import re
from io import BytesIO

import pandas as pd
from rapidfuzz import fuzz
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from proveedores import PADRON_PROVEEDORES


# ─────────────────────────────────────────────
# CARGA
# ─────────────────────────────────────────────

def load_excel_sistema(file) -> pd.DataFrame:
    file.seek(0)
    return pd.read_excel(file)


def load_excel_arca(file, filas_a_revisar: int = 5) -> pd.DataFrame:
    """
    Lee el Excel de percepciones de ARCA detectando automáticamente si los
    encabezados arrancan en la primera o en la segunda columna.
    """
    file.seek(0)
    raw = pd.read_excel(file, header=None, nrows=filas_a_revisar)

    fila_header = None
    for i in range(len(raw)):
        if raw.iloc[i].notna().sum() > 0:
            fila_header = i
            break

    if fila_header is None:
        raise ValueError("No se pudo detectar una fila de encabezados válida en el Excel de ARCA.")

    primera_col_vacia = pd.isna(raw.iloc[fila_header].iloc[0])

    file.seek(0)
    if primera_col_vacia:
        df = pd.read_excel(file, header=fila_header, usecols=lambda c: c != raw.columns[0])
    else:
        df = pd.read_excel(file, header=fila_header)

    df = df.loc[:, ~df.columns.astype(str).str.contains('^Unnamed', na=False)]
    df = df.dropna(axis=1, how='all')

    return df


# ─────────────────────────────────────────────
# DEPURACIÓN SISTEMA
# ─────────────────────────────────────────────

def _parse_nro(s: str) -> tuple[str, str]:
    s = str(s).strip()
    s = s.replace('−', '-').replace('–', '-')  # em dash / en dash
    s = re.sub(r'\s*-\s*', '-', s)
    count = s.count('-')
    if count == 0:
        left, right = s[:4], s[4:]
    elif count == 1:
        left, right = s.split('-', 1)
    else:
        idx = s.index('-')
        left = s[:idx]
        right = s[idx + 1:].replace('-', '')
    pto = left.lstrip('0')
    comp = right.lstrip('0')
    if pto == '' and len(right) >= 3 and right[0] != '0' and right[1:3] == '00':
        pto = right[0]
        comp = right[1:].lstrip('0')
    return pto, comp


def depurar_sistema_percepciones(df: pd.DataFrame, padron: dict | None = None) -> pd.DataFrame:
    """
    Depura el reporte del sistema:
    - Calcula 'Importe' = Debe - Haber.
    - A partir de 'Su Factura' genera 'Pto. Venta' y 'N°Comprobante'.
    - Genera 'CUIT' buscando, para cada 'Tercero', qué nombre del padrón de
      proveedores del repo (proveedores.py) coincide (comparación exacta sin
      espacios ni mayúsculas), y le asigna el primer CUIT encontrado para
      ese nombre (si el mismo nombre tiene más de un CUIT cargado en el
      padrón, se queda con el primero).
    """
    if padron is None:
        padron = PADRON_PROVEEDORES

    df = df.copy()
    df.columns = df.columns.astype(str).str.strip()

    norm_cols = {c.lower(): c for c in df.columns}

    def _resolver(candidatos: list[str]) -> str:
        for cand in candidatos:
            k = cand.lower()
            if k in norm_cols:
                return norm_cols[k]
        raise KeyError(f"No se encontró ninguna columna de {candidatos}. Disponibles: {list(df.columns)}")

    c_debe = _resolver(["Debe"])
    c_haber = _resolver(["Haber"])
    c_factura = _resolver(["Su Factura", "Su factura"])
    c_tercero = _resolver(["Tercero"])

    df[c_debe] = pd.to_numeric(df[c_debe], errors="coerce").fillna(0).round(2)
    df[c_haber] = pd.to_numeric(df[c_haber], errors="coerce").fillna(0).round(2)
    df["Importe"] = (df[c_debe] - df[c_haber]).round(2)

    df[["Pto. Venta", "N°Comprobante"]] = pd.DataFrame(
        df[c_factura].apply(_parse_nro).tolist(), index=df.index
    )

    nombre_a_cuit = {}
    for cuit, nombre in padron.items():
        clave = str(nombre).strip().upper()
        nombre_a_cuit.setdefault(clave, cuit)

    df["CUIT"] = df[c_tercero].astype(str).str.strip().str.upper().map(nombre_a_cuit)

    return df


# ─────────────────────────────────────────────
# DEPURACIÓN ARCA
# ─────────────────────────────────────────────

def depurar_arca_percepciones(df: pd.DataFrame) -> pd.DataFrame:
    """
    Depura el reporte de percepciones de ARCA:
    - Normaliza 'CUIT' a texto.
    - Redondea 'Monto Percibido' a 2 decimales.
    - Separa 'N° Comprobante' en 'Pto. Venta' y 'N°Comprobante' (8 dígitos).
    """
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip()

    norm_cols = {c.lower(): c for c in df.columns}

    def _resolver(candidatos: list[str]) -> str:
        for cand in candidatos:
            k = cand.lower()
            if k in norm_cols:
                return norm_cols[k]
        raise KeyError(f"No se encontró ninguna columna de {candidatos}. Disponibles: {list(df.columns)}")

    c_cuit = _resolver(["CUIT"])
    c_monto = _resolver(["Monto Percibido"])
    c_comp = _resolver(["N° Comprobante", "N°Comprobante", "Nro Comprobante", "Numero Comprobante"])

    if pd.api.types.is_numeric_dtype(df[c_cuit]):
        df[c_cuit] = df[c_cuit].astype("Int64").astype(str)
    else:
        df[c_cuit] = df[c_cuit].astype(str).str.replace(r"[^0-9]", "", regex=True)
    if c_cuit != "CUIT":
        df = df.rename(columns={c_cuit: "CUIT"})

    df[c_monto] = pd.to_numeric(df[c_monto], errors="coerce").round(2)
    if c_monto != "Monto Percibido":
        df = df.rename(columns={c_monto: "Monto Percibido"})

    def _separar_comprobante(valor):
        if pd.isna(valor):
            return pd.Series([pd.NA, pd.NA])

        s = str(valor).strip()
        if s.endswith(".0"):
            s = s[:-2]
        if s[:3] == "A0 ":
            s = s[3:]
        s = re.sub(r"\D", "", s)

        if s == "":
            return pd.Series([pd.NA, pd.NA])

        if len(s) < 9:
            comprobante = s.zfill(8)
            pto_venta = pd.NA
        else:
            comprobante = s[-8:]
            pto_venta = s[:-8]

        return pd.Series([pto_venta, comprobante])

    df[["Pto. Venta", "N°Comprobante"]] = df[c_comp].apply(_separar_comprobante)
    df = df.drop(columns=[c_comp])

    return df


# ─────────────────────────────────────────────
# CRUCE
# ─────────────────────────────────────────────

def cruce_percepciones(
    df_arca: pd.DataFrame,
    df_sistema: pd.DataFrame,
    tolerancia_importe: float = 1.0,
    score_nombre_min: int = 70,
    padron: dict | None = None,
):
    """
    Cruza percepciones de ARCA contra el sistema en tres fases:

    PASO 1 - CUIT + Importe: candidatos con el mismo CUIT y el mismo Importe
    (con tolerancia). Si hay más de uno, desempata por Pto. Venta / N°Comprobante.

    PASO 2 - Importe + CUIT vía padrón: para lo que no matcheó por CUIT
    literal (puede pasar si el mismo proveedor quedó cargado con más de un
    CUIT en el padrón), busca en proveedores.py el nombre correspondiente al
    CUIT de ARCA y lo compara contra 'Tercero' del sistema.

    PASO 3 - Importe + Nombre (fuzzy), para el remanente que no matcheó ni
    por CUIT ni por nombre vía padrón.

    Retorna:
        tuple de 4 pd.DataFrame:
            - df_match_sistema, df_match_arca (ligados por 'id_match', con
              'match_tipo' = 'cuit' / 'cuit_padron' / 'nombre', y 'match_score')
            - df_falta_sistema (líneas de ARCA sin match)
            - df_falta_arca (líneas de sistema sin match)
    """
    if padron is None:
        padron = PADRON_PROVEEDORES

    df_arca = df_arca.reset_index(drop=True).copy()
    df_sistema = df_sistema.reset_index(drop=True).copy()

    def _buscar_columna(df, nombre):
        return next(
            (c for c in df.columns if str(c).replace(".", "").strip().upper() == nombre),
            None
        )

    def _norm_digitos(serie):
        def limpiar(x):
            if pd.isna(x):
                return ""
            s = re.sub(r"\D", "", str(x))
            return str(int(s)) if s else ""
        return serie.apply(limpiar)

    def _normalizar_nombre(nombre):
        return re.sub(r"\s+", "", str(nombre)).strip().upper()

    col_cuit_arca = _buscar_columna(df_arca, "CUIT")
    col_cuit_sistema = _buscar_columna(df_sistema, "CUIT")

    cuit_arca_norm = df_arca[col_cuit_arca].astype(str).str.replace(r"[^0-9]", "", regex=True)
    cuit_sistema_norm = df_sistema[col_cuit_sistema].astype(str).str.replace(r"[^0-9]", "", regex=True)

    col_pv_arca = _buscar_columna(df_arca, "PTO VENTA")
    col_nc_arca = _buscar_columna(df_arca, "N°COMPROBANTE")
    col_pv_sistema = _buscar_columna(df_sistema, "PTO VENTA")
    col_nc_sistema = _buscar_columna(df_sistema, "N°COMPROBANTE")

    usar_comprobante = all([col_pv_arca, col_nc_arca, col_pv_sistema, col_nc_sistema])

    if usar_comprobante:
        pv_arca_norm = _norm_digitos(df_arca[col_pv_arca])
        nc_arca_norm = _norm_digitos(df_arca[col_nc_arca])
        pv_sistema_norm = _norm_digitos(df_sistema[col_pv_sistema])
        nc_sistema_norm = _norm_digitos(df_sistema[col_nc_sistema])

    def desempatar_por_comprobante(arca_idx, candidatos_idx):
        if len(candidatos_idx) <= 1 or not usar_comprobante:
            return candidatos_idx

        pv_a = pv_arca_norm.loc[arca_idx]
        nc_a = nc_arca_norm.loc[arca_idx]

        nivel_a = [i for i in candidatos_idx
                   if pv_sistema_norm.loc[i] == pv_a and nc_sistema_norm.loc[i] == nc_a
                   and nc_a != ""]
        if nivel_a:
            return nivel_a

        nivel_b = [i for i in candidatos_idx if nc_sistema_norm.loc[i] == nc_a and nc_a != ""]
        if nivel_b:
            return nivel_b

        nivel_c = [i for i in candidatos_idx if pv_sistema_norm.loc[i] == pv_a and pv_a != ""]
        if nivel_c:
            return nivel_c

        concat_a = f"{pv_a}{nc_a}"
        mejor_score = -1
        ganadores = []
        for i in candidatos_idx:
            concat_s = f"{pv_sistema_norm.loc[i]}{nc_sistema_norm.loc[i]}"
            score = fuzz.ratio(concat_a, concat_s)
            if score > mejor_score:
                mejor_score = score
                ganadores = [i]
            elif score == mejor_score:
                ganadores.append(i)
        return ganadores

    matches = []  # (a_idx, s_idx, tipo, score)
    sistema_disponibles = set(df_sistema.index)

    # PASO 1: CUIT + Importe
    for a_idx in df_arca.index:
        cuit_a = cuit_arca_norm.loc[a_idx]
        monto_a = df_arca.loc[a_idx, "Monto Percibido"]

        candidatos_idx = [
            i for i in sistema_disponibles
            if cuit_sistema_norm.loc[i] == cuit_a
            and abs(df_sistema.loc[i, "Importe"] - monto_a) <= tolerancia_importe
        ]
        if not candidatos_idx:
            continue

        candidatos_idx = desempatar_por_comprobante(a_idx, candidatos_idx)
        s_idx = candidatos_idx[0]

        matches.append((a_idx, s_idx, "cuit", 100))
        sistema_disponibles.remove(s_idx)

    # PASO 2: Importe + CUIT vía padrón (mismo proveedor, CUIT distinto por duplicados)
    arca_matcheados = {a_idx for a_idx, s_idx, tipo, score in matches}
    arca_pendientes = [i for i in df_arca.index if i not in arca_matcheados]

    for a_idx in arca_pendientes:
        cuit_a = cuit_arca_norm.loc[a_idx]
        nombre_padron = padron.get(cuit_a)
        if not nombre_padron:
            continue

        monto_a = df_arca.loc[a_idx, "Monto Percibido"]
        candidatos_idx = [
            i for i in sistema_disponibles
            if abs(df_sistema.loc[i, "Importe"] - monto_a) <= tolerancia_importe
            and _normalizar_nombre(df_sistema.loc[i, "Tercero"]) == _normalizar_nombre(nombre_padron)
        ]
        if not candidatos_idx:
            continue

        candidatos_idx = desempatar_por_comprobante(a_idx, candidatos_idx)
        s_idx = candidatos_idx[0]

        matches.append((a_idx, s_idx, "cuit_padron", 100))
        sistema_disponibles.remove(s_idx)
        arca_matcheados.add(a_idx)

    # PASO 3: Importe + Nombre (fuzzy), para el remanente
    arca_pendientes = [i for i in df_arca.index if i not in arca_matcheados]

    for a_idx in arca_pendientes:
        monto_a = df_arca.loc[a_idx, "Monto Percibido"]
        nombre_a = str(df_arca.loc[a_idx, "Razon Social"])

        candidatos_idx = [
            i for i in sistema_disponibles
            if abs(df_sistema.loc[i, "Importe"] - monto_a) <= tolerancia_importe
        ]
        if not candidatos_idx:
            continue

        puntajes = [(i, fuzz.token_sort_ratio(nombre_a, str(df_sistema.loc[i, "Tercero"])))
                    for i in candidatos_idx]
        mejor_score = max(p[1] for p in puntajes)

        if mejor_score < score_nombre_min:
            continue

        empatados = [i for i, sc in puntajes if sc == mejor_score]
        empatados = desempatar_por_comprobante(a_idx, empatados)
        s_idx = empatados[0]

        matches.append((a_idx, s_idx, "nombre", mejor_score))
        sistema_disponibles.remove(s_idx)

    arca_matcheados = {a_idx for a_idx, s_idx, tipo, score in matches}
    sistema_matcheados = {s_idx for a_idx, s_idx, tipo, score in matches}

    filas_sistema, filas_arca = [], []
    for id_match, (a_idx, s_idx, tipo, score) in enumerate(matches, start=1):
        fila_s = df_sistema.loc[s_idx].copy()
        fila_s["id_match"] = id_match
        fila_s["match_tipo"] = tipo
        fila_s["match_score"] = score
        filas_sistema.append(fila_s)

        fila_a = df_arca.loc[a_idx].copy()
        fila_a["id_match"] = id_match
        fila_a["match_tipo"] = tipo
        fila_a["match_score"] = score
        filas_arca.append(fila_a)

    cols_extra = ["id_match", "match_tipo", "match_score"]

    if filas_sistema:
        df_match_sistema = pd.DataFrame(filas_sistema).reset_index(drop=True)
        df_match_sistema = df_match_sistema[[c for c in df_match_sistema.columns if c not in cols_extra] + cols_extra]
    else:
        df_match_sistema = pd.DataFrame(columns=list(df_sistema.columns) + cols_extra)

    if filas_arca:
        df_match_arca = pd.DataFrame(filas_arca).reset_index(drop=True)
        df_match_arca = df_match_arca[[c for c in df_match_arca.columns if c not in cols_extra] + cols_extra]
    else:
        df_match_arca = pd.DataFrame(columns=list(df_arca.columns) + cols_extra)

    idx_falta_sistema = [i for i in df_arca.index if i not in arca_matcheados]
    df_falta_sistema = df_arca.loc[idx_falta_sistema].reset_index(drop=True)

    idx_falta_arca = [i for i in df_sistema.index if i not in sistema_matcheados]
    df_falta_arca = df_sistema.loc[idx_falta_arca].reset_index(drop=True)

    return df_match_sistema, df_match_arca, df_falta_sistema, df_falta_arca


# ─────────────────────────────────────────────
# EXPORTAR A BUFFER EN MEMORIA (descargable único)
# ─────────────────────────────────────────────

def generar_excel_percepciones(
    df_match_sistema: pd.DataFrame,
    df_match_arca: pd.DataFrame,
    df_falta_sistema: pd.DataFrame,
    df_falta_arca: pd.DataFrame,
) -> bytes:
    DATE_FORMAT = "DD/MM/YYYY"
    DATE_COLS = {"Fecha", "Fecha factura", "Fecha Percepcion", "Fecha Comprobante"}

    def _style_sheet(ws) -> None:
        thin = Side(style="thin")
        header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        header_font = Font(bold=True)
        header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, cell in enumerate(ws[1], start=1):
            if str(cell.value).strip() in DATE_COLS:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col_idx).number_format = DATE_FORMAT

        for col in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0 for c in col),
                default=10,
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)

    hojas = {
        "Match_Sistema": df_match_sistema,
        "Match_Arca": df_match_arca,
        "Falta_Sistema": df_falta_sistema,
        "Falta_Arca": df_falta_arca,
    }

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for nombre_hoja, df in hojas.items():
            df.to_excel(writer, sheet_name=nombre_hoja, index=False)

        wb = writer.book
        for nombre_hoja in hojas:
            _style_sheet(wb[nombre_hoja])

    return buf.getvalue()


# ─────────────────────────────────────────────
# PIPELINE COMPLETO
# ─────────────────────────────────────────────

def correr_cruce_percepciones(
    archivo_arca, archivo_sistema, tolerancia_importe: float = 1.0, score_nombre_min: int = 70
):
    df_sistema = load_excel_sistema(archivo_sistema)
    df_arca = load_excel_arca(archivo_arca)

    df_sistema_dep = depurar_sistema_percepciones(df_sistema)
    df_arca_dep = depurar_arca_percepciones(df_arca)

    df_match_sistema, df_match_arca, df_falta_sistema, df_falta_arca = cruce_percepciones(
        df_arca_dep, df_sistema_dep, tolerancia_importe=tolerancia_importe, score_nombre_min=score_nombre_min
    )

    stats = {
        "match": len(df_match_arca),
        "faltante_sistema": len(df_falta_sistema),
        "faltante_arca": len(df_falta_arca),
    }

    buf_reporte = generar_excel_percepciones(
        df_match_sistema, df_match_arca, df_falta_sistema, df_falta_arca
    )

    return buf_reporte, stats
