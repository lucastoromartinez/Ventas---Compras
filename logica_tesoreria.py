"""
logica_tesoreria.py — Conciliación Tesorería (Caja Central vs Contabilidad)

Cruza los movimientos de la Caja Central (un Excel con una hoja por día,
cargado por Tesorería) contra el mayor contable unificado del sistema:
  1. Carga la Caja Central hoja por hoja, ubicando SALDO INICIAL, encabezado
     Detalle/Monto, SALDO FINAL y la fila de DIFERENCIA de arqueo.
  2. Depura la Caja Unificada del sistema (Monto = Debe - Haber).
  3. Cruza ambos lados en 4 pasos: ingresos agrupados, agrupamiento por
     nombre, uno a uno por monto+fecha, y combinaciones (varias líneas de
     un lado suman una del otro).
  4. Exporta el resultado (match y faltantes de cada lado) a un Excel en
     memoria, con formato numérico/fecha y ancho de columna autoajustado.
"""

import datetime
import re
import unicodedata
from io import BytesIO
from itertools import combinations

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
# CARGA
# ─────────────────────────────────────────────

def cargar_caja_central(archivo) -> pd.DataFrame:
    """
    Lee el Excel de Caja Central (una hoja por día) y arma un único
    DataFrame con columnas: Fecha, Detalle, Monto.

    Por hoja:
      - Busca la celda "SALDO INICIAL" y toma la fecha de esa misma fila.
      - Busca el encabezado "Detalle" / "Monto" para saber las columnas.
      - Ubica la fila de "DIFERENCIA" de arqueo tomando la ÚLTIMA fila que
        menciona esa palabra (siempre al final, tras SALDO FINAL y ARQUEO),
        para no confundirla con un movimiento real (ej. "Diferencia de
        sueldos").
      - Toma como movimientos las filas entre el encabezado y "SALDO
        FINAL", excluyendo por número de fila la de "DIFERENCIA" ya
        identificada, y las que digan SALDO o ARQUEO.
      - Si esa fila de diferencia tiene un monto distinto de cero, agrega
        una fila extra "Diferencia Arqueo" con ese valor.
    """
    wb = openpyxl.load_workbook(archivo, data_only=True)
    registros = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column

        fecha_inicial = None
        header_row = None
        col_detalle = None
        col_monto = None
        saldo_final_row = None
        diferencia_row = None

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                val = ws.cell(row=r, column=c).value
                if not isinstance(val, str):
                    continue
                texto = val.strip().upper()

                if fecha_inicial is None and "SALDO INICIAL" in texto:
                    for cc in range(c + 1, max_col + 1):
                        v2 = ws.cell(row=r, column=cc).value
                        if isinstance(v2, (datetime.datetime, datetime.date)):
                            fecha_inicial = v2
                            break

                elif header_row is None and texto == "DETALLE":
                    header_row = r
                    col_detalle = c
                    for cc in range(c + 1, max_col + 1):
                        v2 = ws.cell(row=r, column=cc).value
                        if isinstance(v2, str) and v2.strip().upper() == "MONTO":
                            col_monto = cc
                            break

                elif saldo_final_row is None and "SALDO FINAL" in texto:
                    saldo_final_row = r

                elif "DIFERENCIA" in texto:
                    diferencia_row = r  # se queda con la ÚLTIMA ocurrencia

        if fecha_inicial is None or header_row is None or col_monto is None:
            continue

        fin = saldo_final_row if saldo_final_row is not None else max_row

        for r in range(header_row + 1, fin):
            if r == diferencia_row:
                continue

            detalle = ws.cell(row=r, column=col_detalle).value
            monto = ws.cell(row=r, column=col_monto).value

            if not isinstance(detalle, str) or monto is None:
                continue
            if not isinstance(monto, (int, float)):
                continue

            texto_up = detalle.strip().upper()
            if texto_up == "" or "SALDO" in texto_up or "ARQUEO" in texto_up:
                continue

            registros.append({
                "Fecha": fecha_inicial,
                "Detalle": detalle.strip(),
                "Monto": monto,
            })

        if diferencia_row is not None:
            monto_dif = ws.cell(row=diferencia_row, column=col_monto).value
            if monto_dif is not None and monto_dif != 0:
                registros.append({
                    "Fecha": fecha_inicial,
                    "Detalle": "Diferencia Arqueo",
                    "Monto": monto_dif,
                })

    df = pd.DataFrame(registros, columns=["Fecha", "Detalle", "Monto"])
    df["Fecha"] = pd.to_datetime(df["Fecha"])
    df = df.sort_values("Fecha", kind="stable").reset_index(drop=True)
    return df


def cargar_caja_central_multiple(archivos) -> pd.DataFrame:
    """
    Igual que cargar_caja_central pero acepta varios archivos (uno por
    mes) y devuelve un único DataFrame con todos los meses juntos,
    ordenado por Fecha.
    """
    if not isinstance(archivos, (list, tuple)):
        archivos = [archivos]

    dfs = [cargar_caja_central(archivo) for archivo in archivos]
    df = pd.concat(dfs, ignore_index=True)
    df = df.sort_values("Fecha", kind="stable").reset_index(drop=True)
    return df


def load_excel_file(archivo) -> pd.DataFrame:
    return pd.read_excel(archivo)


# ─────────────────────────────────────────────
# DEPURACIÓN
# ─────────────────────────────────────────────

def depurar_caja_unificada(df: pd.DataFrame) -> pd.DataFrame:
    """
    Crea la columna "Monto" = Debe - Haber (NaN/vacío se toman como 0).
    """
    df = df.copy()

    def a_numero(serie):
        if serie.dtype == object:
            serie = (
                serie.astype(str)
                .str.strip()
                .str.replace(".", "", regex=False)   # separador de miles
                .str.replace(",", ".", regex=False)  # coma decimal -> punto
            )
            serie = serie.replace({"": None, "nan": None, "None": None})
        return pd.to_numeric(serie, errors="coerce")

    debe = a_numero(df["Debe"]).fillna(0)
    haber = a_numero(df["Haber"]).fillna(0)

    df["Monto"] = (debe - haber).astype("float64").round(2)

    return df


# ─────────────────────────────────────────────
# CRUCE
# ─────────────────────────────────────────────

PALABRAS_EXCLUIDAS_NOMBRE = ("pago", "devolucion", "ingreso")


def _normalizar_texto(texto):
    if not isinstance(texto, str):
        return ""
    t = unicodedata.normalize("NFKD", texto)
    t = "".join(c for c in t if not unicodedata.combining(c))
    t = t.lower()
    t = re.sub(r"[^a-z0-9\s]", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _es_nombre(texto_normalizado):
    if texto_normalizado == "":
        return False
    return not any(p in texto_normalizado for p in PALABRAS_EXCLUIDAS_NOMBRE)


def _tolerancia_dinamica(monto, tolerancia_pesos, tolerancia_pct):
    return max(tolerancia_pesos, abs(monto) * tolerancia_pct / 100)


def _buscar_combinacion(monto_objetivo, fecha_objetivo, pool_df, max_combinacion,
                         ventana_dias, tolerancia_pesos, tolerancia_pct):
    """
    Busca en pool_df un subconjunto de filas (2 a max_combinacion) cuya
    suma coincida con monto_objetivo (con tolerancia), restringido a filas
    a lo sumo a ventana_dias de fecha_objetivo. Devuelve los índices de la
    PRIMER combinación válida (se prueban primero las más chicas), o None.
    """
    candidatos = pool_df[
        pool_df["_fecha_norm"].apply(lambda f: abs((f - fecha_objetivo).days) <= ventana_dias)
    ]
    if len(candidatos) < 2:
        return None

    tol = _tolerancia_dinamica(monto_objetivo, tolerancia_pesos, tolerancia_pct)
    indices = candidatos.index.tolist()

    for k in range(2, max_combinacion + 1):
        if k > len(indices):
            break
        for combo in combinations(indices, k):
            suma = candidatos.loc[list(combo), "_monto_norm"].sum()
            if abs(suma - monto_objetivo) <= tol:
                return list(combo)
    return None


def cruzar_caja(
    df_caja_unificada,
    df_caja_michu,
    col_fecha="Fecha",
    col_monto="Monto",
    col_detalle_unificada="Comentario",
    col_detalle_michu="Detalle",
    tolerancia_pesos=5,
    tolerancia_pct=0.001,
    tolerancia_dias=3,
    max_combinacion=4,
    ventana_dias_combinacion=5,
):
    """
    Cruza df_caja_unificada (sistema) contra df_caja_michu (tesorería).

    PASO 1 - Ingresos agrupados: suma de "ingreso" (tesorería) vs "ingreso
    efectivo" (sistema).
    PASO 2 - Agrupamiento por nombre: filas cuyo detalle no dice "pago",
    "devolucion" ni "ingreso" se agrupan por nombre normalizado y se
    cruzan sus sumas.
    PASO 3 - Uno a uno: por Monto (con tolerancia) + Fecha (con
    tolerancia_dias).
    PASO 4 - Combinaciones: varias líneas de un lado (hasta
    max_combinacion) que suman el monto de una línea del otro lado,
    dentro de una ventana de ventana_dias_combinacion días. Es greedy y
    más propenso a falsos positivos, por eso queda taggeado aparte como
    "Agrupado (Combinación)".

    Returns
    -------
    dict con match_caja_unificada, match_tesoreria, falta_unificada,
    falta_tesoreria y warnings (lista de strings con inconsistencias
    detectadas, ej. ingresos que no cuadran).
    """

    unif = df_caja_unificada.copy().reset_index(drop=True)
    michu = df_caja_michu.copy().reset_index(drop=True)

    unif["_id"] = unif.index
    michu["_id"] = michu.index

    unif["_matched"] = False
    michu["_matched"] = False
    unif["id"] = pd.NA
    michu["id"] = pd.NA

    unif["_fecha_norm"] = pd.to_datetime(unif[col_fecha]).dt.date
    michu["_fecha_norm"] = pd.to_datetime(michu[col_fecha]).dt.date
    unif["_monto_norm"] = unif[col_monto].astype(float).round(2)
    michu["_monto_norm"] = michu[col_monto].astype(float).round(2)

    match_id_counter = 1
    tipo_por_id = {}
    warnings = []

    # ------------------------------------------------------------------
    # PASO 1: bloque de "ingresos" agrupados
    # ------------------------------------------------------------------
    def tiene_ingreso_y_mas_de_un_numero(texto):
        if not isinstance(texto, str):
            return False
        if "ingreso" not in texto.lower():
            return False
        numeros = re.findall(r"\d+", texto)
        return len(numeros) > 1

    mask_michu_ingreso = michu[col_detalle_michu].apply(tiene_ingreso_y_mas_de_un_numero)
    mask_unif_ingreso_efectivo = (
        unif[col_detalle_unificada].astype(str).str.lower().str.contains("ingreso efectivo", na=False)
    )

    grupo_michu = michu[mask_michu_ingreso]
    grupo_unif = unif[mask_unif_ingreso_efectivo]

    suma_michu = grupo_michu["_monto_norm"].sum()
    suma_unif = grupo_unif["_monto_norm"].sum()

    if len(grupo_michu) > 0 and len(grupo_unif) > 0:
        tol = _tolerancia_dinamica(max(abs(suma_michu), abs(suma_unif)), tolerancia_pesos, tolerancia_pct)
        if abs(suma_michu - suma_unif) <= tol:
            michu.loc[grupo_michu.index, "_matched"] = True
            michu.loc[grupo_michu.index, "id"] = match_id_counter
            unif.loc[grupo_unif.index, "_matched"] = True
            unif.loc[grupo_unif.index, "id"] = match_id_counter
            tipo_por_id[match_id_counter] = "Agrupado (Ingreso)"
            match_id_counter += 1
        else:
            warnings.append(
                f"Ingresos no cuadran: tesorería suma {suma_michu:.2f} vs sistema "
                f"'Ingreso efectivo' suma {suma_unif:.2f} (diferencia "
                f"{suma_michu - suma_unif:.2f}, tolerancia {tol:.2f}). No se marcaron como matcheados."
            )

    # ------------------------------------------------------------------
    # PASO 2: agrupamiento por nombre, en ambos lados
    # ------------------------------------------------------------------
    unif["_nombre_norm"] = unif[col_detalle_unificada].apply(_normalizar_texto)
    michu["_nombre_norm"] = michu[col_detalle_michu].apply(_normalizar_texto)

    restante_unif = unif[~unif["_matched"]]
    restante_michu = michu[~michu["_matched"]]

    nombres_unif = set(
        restante_unif.loc[restante_unif["_nombre_norm"].apply(_es_nombre), "_nombre_norm"]
    )
    nombres_michu = set(
        restante_michu.loc[restante_michu["_nombre_norm"].apply(_es_nombre), "_nombre_norm"]
    )
    nombres_comunes = nombres_unif & nombres_michu

    for nombre in nombres_comunes:
        grupo_u = unif[(~unif["_matched"]) & (unif["_nombre_norm"] == nombre)]
        grupo_m = michu[(~michu["_matched"]) & (michu["_nombre_norm"] == nombre)]
        if grupo_u.empty or grupo_m.empty:
            continue

        suma_u = grupo_u["_monto_norm"].sum()
        suma_m = grupo_m["_monto_norm"].sum()
        tol = _tolerancia_dinamica(max(abs(suma_u), abs(suma_m)), tolerancia_pesos, tolerancia_pct)
        if abs(suma_u - suma_m) > tol:
            continue

        fechas = pd.concat([grupo_u["_fecha_norm"], grupo_m["_fecha_norm"]])
        rango_dias = (fechas.max() - fechas.min()).days
        if rango_dias > tolerancia_dias:
            continue

        unif.loc[grupo_u.index, "_matched"] = True
        unif.loc[grupo_u.index, "id"] = match_id_counter
        michu.loc[grupo_m.index, "_matched"] = True
        michu.loc[grupo_m.index, "id"] = match_id_counter
        tipo_por_id[match_id_counter] = "Agrupado (Nombre)"
        match_id_counter += 1

    # ------------------------------------------------------------------
    # PASO 3: resto de conceptos, uno a uno por Monto (con tolerancia) +
    #         Fecha (con tolerancia_dias)
    # ------------------------------------------------------------------
    restante_unif = unif[~unif["_matched"]].sort_values("_fecha_norm")
    restante_michu = michu[~michu["_matched"]].copy()

    for _, fila_u in restante_unif.iterrows():
        if restante_michu.empty:
            break

        monto_u = fila_u["_monto_norm"]
        fecha_u = fila_u["_fecha_norm"]
        tol = _tolerancia_dinamica(monto_u, tolerancia_pesos, tolerancia_pct)

        candidatos = restante_michu[(restante_michu["_monto_norm"] - monto_u).abs() <= tol]
        if candidatos.empty:
            continue

        dif_dias = candidatos["_fecha_norm"].apply(lambda f: abs((f - fecha_u).days))
        candidatos = candidatos[dif_dias <= tolerancia_dias]
        if candidatos.empty:
            continue

        dif_dias = candidatos["_fecha_norm"].apply(lambda f: abs((f - fecha_u).days))
        dif_monto = (candidatos["_monto_norm"] - monto_u).abs()
        orden = pd.DataFrame({"dif_dias": dif_dias, "dif_monto": dif_monto}).sort_values(
            ["dif_dias", "dif_monto"]
        )
        fila_m = candidatos.loc[orden.index[0]]
        id_michu_elegido = fila_m["_id"]

        dm = round(abs(monto_u - fila_m["_monto_norm"]), 2)
        dd = abs((fecha_u - fila_m["_fecha_norm"]).days)
        if dm == 0 and dd == 0:
            tipo = "Exacto"
        elif dm > 0 and dd == 0:
            tipo = "Tolerancia Importe"
        elif dm == 0 and dd > 0:
            tipo = "Tolerancia Fecha"
        else:
            tipo = "Tolerancia Importe y Fecha"

        unif.loc[unif["_id"] == fila_u["_id"], "_matched"] = True
        unif.loc[unif["_id"] == fila_u["_id"], "id"] = match_id_counter
        michu.loc[michu["_id"] == id_michu_elegido, "_matched"] = True
        michu.loc[michu["_id"] == id_michu_elegido, "id"] = match_id_counter
        tipo_por_id[match_id_counter] = tipo
        match_id_counter += 1

        restante_michu = restante_michu[restante_michu["_id"] != id_michu_elegido]

    # ------------------------------------------------------------------
    # PASO 4: combinaciones (una línea de un lado = varias del otro)
    # ------------------------------------------------------------------
    restante_unif = unif[~unif["_matched"]].sort_values("_fecha_norm")
    for _, fila_u in restante_unif.iterrows():
        if unif.loc[unif["_id"] == fila_u["_id"], "_matched"].iloc[0]:
            continue
        pool_michu = michu[~michu["_matched"]]
        combo_idx = _buscar_combinacion(
            fila_u["_monto_norm"], fila_u["_fecha_norm"], pool_michu,
            max_combinacion, ventana_dias_combinacion, tolerancia_pesos, tolerancia_pct,
        )
        if combo_idx is None:
            continue

        unif.loc[unif["_id"] == fila_u["_id"], "_matched"] = True
        unif.loc[unif["_id"] == fila_u["_id"], "id"] = match_id_counter
        michu.loc[combo_idx, "_matched"] = True
        michu.loc[combo_idx, "id"] = match_id_counter
        tipo_por_id[match_id_counter] = "Agrupado (Combinación)"
        match_id_counter += 1

    restante_michu = michu[~michu["_matched"]].sort_values("_fecha_norm")
    for _, fila_m in restante_michu.iterrows():
        if michu.loc[michu["_id"] == fila_m["_id"], "_matched"].iloc[0]:
            continue
        pool_unif = unif[~unif["_matched"]]
        combo_idx = _buscar_combinacion(
            fila_m["_monto_norm"], fila_m["_fecha_norm"], pool_unif,
            max_combinacion, ventana_dias_combinacion, tolerancia_pesos, tolerancia_pct,
        )
        if combo_idx is None:
            continue

        michu.loc[michu["_id"] == fila_m["_id"], "_matched"] = True
        michu.loc[michu["_id"] == fila_m["_id"], "id"] = match_id_counter
        unif.loc[combo_idx, "_matched"] = True
        unif.loc[combo_idx, "id"] = match_id_counter
        tipo_por_id[match_id_counter] = "Agrupado (Combinación)"
        match_id_counter += 1

    unif["Tipo Match"] = unif["id"].map(tipo_por_id)
    michu["Tipo Match"] = michu["id"].map(tipo_por_id)

    # ------------------------------------------------------------------
    # Armar resultados finales
    # ------------------------------------------------------------------
    cols_out_unif = list(df_caja_unificada.columns) + ["id", "Tipo Match"]
    cols_out_michu = list(df_caja_michu.columns) + ["id", "Tipo Match"]

    match_caja_unificada = unif[unif["_matched"]][cols_out_unif].sort_values("id").reset_index(drop=True)
    match_tesoreria = michu[michu["_matched"]][cols_out_michu].sort_values("id").reset_index(drop=True)
    falta_unificada = unif[~unif["_matched"]][list(df_caja_unificada.columns) + ["id"]].reset_index(drop=True)
    falta_tesoreria = michu[~michu["_matched"]][list(df_caja_michu.columns) + ["id"]].reset_index(drop=True)

    return {
        "match_caja_unificada": match_caja_unificada,
        "match_tesoreria": match_tesoreria,
        "falta_unificada": falta_unificada,
        "falta_tesoreria": falta_tesoreria,
        "warnings": warnings,
    }


# ─────────────────────────────────────────────
# EXPORTAR EN MEMORIA
# ─────────────────────────────────────────────

def generar_excel_en_memoria_tesoreria(match_caja_unificada, match_tesoreria,
                                        falta_unificada, falta_tesoreria) -> bytes:
    """
    Exporta los 4 DataFrames del cruce a un Excel en memoria, con:
      - columnas float con formato numérico (separador de miles, 2 decimales)
      - columnas de fecha con formato dd/mm/aaaa
      - ancho de columna autoajustado al contenido
    """
    buf = BytesIO()

    hojas = {
        "Match Contabilidad": match_caja_unificada,
        "Match Tesoreria": match_tesoreria,
        "Falta Contabilidad": falta_unificada,
        "Falta Tesoreria": falta_tesoreria,
    }

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for nombre, df in hojas.items():
            df.to_excel(writer, sheet_name=nombre, index=False)

        workbook = writer.book

        for nombre, df in hojas.items():
            ws = workbook[nombre]

            for col_idx, col_name in enumerate(df.columns, start=1):
                letra = get_column_letter(col_idx)
                serie = df[col_name]

                if pd.api.types.is_float_dtype(serie):
                    formato = "#,##0.00"
                elif pd.api.types.is_datetime64_any_dtype(serie):
                    formato = "dd/mm/yyyy"
                else:
                    formato = None

                if formato:
                    for celda in ws[letra][1:]:  # saltea el encabezado
                        celda.number_format = formato

                largos = [len(str(col_name))]
                if len(serie):
                    largos += [len(str(v)) for v in serie.dropna()]
                ws.column_dimensions[letra].width = max(largos) + 3

    return buf.getvalue()


# ─────────────────────────────────────────────
# PIPELINE COMPLETO
# ─────────────────────────────────────────────

def correr_conciliacion_tesoreria(
    archivos_caja_central,
    archivo_caja_unificada,
    col_detalle_unificada="Comentario",
    tolerancia_pesos=5,
    tolerancia_pct=0.001,
    tolerancia_dias=3,
):
    """
    archivos_caja_central: uno o varios archivos de Caja Central (uno por
    mes, cada uno con una hoja por día). archivo_caja_unificada: un único
    archivo del sistema que ya puede traer varios meses juntos.
    """
    df_caja_michu = cargar_caja_central_multiple(archivos_caja_central)

    df_caja_unificada = load_excel_file(archivo_caja_unificada)
    df_caja_unificada = depurar_caja_unificada(df_caja_unificada)

    resultado = cruzar_caja(
        df_caja_unificada,
        df_caja_michu,
        col_detalle_unificada=col_detalle_unificada,
        col_detalle_michu="Detalle",
        tolerancia_pesos=tolerancia_pesos,
        tolerancia_pct=tolerancia_pct,
        tolerancia_dias=tolerancia_dias,
    )

    buf = generar_excel_en_memoria_tesoreria(
        resultado["match_caja_unificada"],
        resultado["match_tesoreria"],
        resultado["falta_unificada"],
        resultado["falta_tesoreria"],
    )

    stats = {
        "match": len(resultado["match_caja_unificada"]),
        "falta_contabilidad": len(resultado["falta_unificada"]),
        "falta_tesoreria": len(resultado["falta_tesoreria"]),
        "warnings": resultado["warnings"],
    }

    return buf, stats
