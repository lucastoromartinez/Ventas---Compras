"""
logica_cupones.py — Conciliación Cupones (Nave vs Extracto Banco)
Cruza el reporte de acreditaciones de Nave contra el extracto bancario (que
ya llega normalizado y categorizado, con la columna "conciliacion"
cargada), y arma el reporte en el mismo formato, hojas, encabezados y
fórmulas que "Plantilla_Automatizacion_Conciliacion_Nave_FINAL":

  - "BANCO ACTUAL" / "NAVE ACTUAL": extracto y reporte del mes actual,
    reducidos a las columnas de la plantilla.
  - "BANCO <mes anterior>" / "NAVE <mes anterior>": ídem para el mes
    anterior (solo si se cargó).
  - "PENDIENTES ANTERIORES": grupos/cupones del mes anterior que Nave
    registró pero el banco todavía no había acreditado a su cierre.
  - "CRUCE AUTOMÁTICO": una fila por "Clave de cruce" (Grupo de
    acreditación de Nave, o N° de operación si Nave no agrupó) con Banco,
    Nave, Diferencia y la categoría de conciliación.
  - "CONCILIACIÓN DIFERENCIA": cascada que explica la diferencia total
    Banco vs Nave categoría por categoría hasta llegar a $0 (CONCILIADO).

correr_conciliacion_cupones_plantilla() es el pipeline de entrada para la
app (Streamlit pasa file-like en vez de Path).
"""

from io import BytesIO
from itertools import combinations

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# CARGA
# ─────────────────────────────────────────────────────────────────────────────

def importar_extracto_banco(archivo) -> pd.DataFrame:
    """
    El extracto de banco ya llega normalizado y categorizado (fecha,
    debitos/creditos/saldo/importe tipados y columna "conciliacion"
    cargada), así que se importa tal cual.
    """
    return pd.read_excel(archivo)


def importar_reporte_nave(
    archivo,
    hoja: str | int = 0,
    columna_busqueda: int = 0,
    texto_encabezado: str = "Fecha de operación",
    max_filas_busqueda: int = 50,
) -> pd.DataFrame:
    raw = pd.read_excel(archivo, sheet_name=hoja, header=None, nrows=max_filas_busqueda)

    mascara = raw.iloc[:, columna_busqueda].astype(str).str.strip() == texto_encabezado
    filas_encontradas = raw.index[mascara].tolist()
    if not filas_encontradas:
        raise ValueError(
            f"No se encontró '{texto_encabezado}' en las primeras "
            f"{max_filas_busqueda} filas de la columna {columna_busqueda}."
        )
    fila_header = filas_encontradas[0]

    df = pd.read_excel(archivo, sheet_name=hoja, header=fila_header)
    df.dropna(how="all", inplace=True)
    df.reset_index(drop=True, inplace=True)

    mascara_total = df.iloc[:, 0].astype(str).str.strip() == "Total"
    filas_total = df.index[mascara_total].tolist()
    if filas_total:
        df = df.loc[:filas_total[0] - 1]
    df.reset_index(drop=True, inplace=True)

    return df


# ─────────────────────────────────────────────────────────────────────────────
# DEPURACIÓN
# ─────────────────────────────────────────────────────────────────────────────

def depurar_leyenda(df: pd.DataFrame) -> pd.DataFrame:
    """
    Deja "leyenda adicional1" reducida al código de operación puro, sacando
    los prefijos que agrega el banco ("Devolución - ", "Operación ",
    "Grupo de acreditación ") y los sufijos "PCT"/"TCTD". Solo se saca el
    guion pegado a "Devolución": el resto de los guiones se respeta porque
    los códigos de grupo también los usan (p.ej. "LS-21825349").
    """
    df = df.copy()
    df["leyenda adicional1"] = (
        df["leyenda adicional1"]
        .astype(str)
        .str.replace(r"^Devolución\s*-\s*", "", regex=True)
        .str.replace(r"^Operación\s+", "", regex=True)
        .str.replace(r"^Grupo de acreditación\s+", "", regex=True)
        .str.replace(" PCT", "", regex=False)
        .str.replace(" TCTD", "", regex=False)
        .str.strip()
    )
    return df


def depurar_nave(df: pd.DataFrame) -> pd.DataFrame:
    """
    Además de tipar fechas y montos, deja tres columnas de identificación
    separadas (usadas por el reporte tipo Plantilla):
      - "N° operación": col_operacion limpio (sin ".0" cuando es numérico).
      - "Grupo de acreditación": tal cual lo informó Nave, crudo (vacío/None
        cuando Nave nunca agrupó la operación).
      - "Clave de cruce": la clave para matchear contra el banco — Grupo de
        acreditación si existe, si no N° operación.
    """
    df = df.copy()

    # --- Detectar nombre de columna de operación ---
    if "Número de operación" in df.columns:
        col_operacion = "Número de operación"
    elif "Código de operación" in df.columns:
        col_operacion = "Código de operación"
    else:
        raise ValueError("No se encontró 'Número de operación' ni 'Código de operación' en el DataFrame.")

    # --- Fecha de operación ---
    df["Fecha de operación"] = pd.to_datetime(
        df["Fecha de operación"].astype(str).str.strip(),
        format="%d/%m/%Y %H:%M",
        errors="coerce"
    )

    # --- Fecha de acreditación ---
    df["Fecha de acreditación"] = pd.to_datetime(
        df["Fecha de acreditación"].astype(str).str.strip(),
        format="%d/%m/%Y",
        errors="coerce"
    )

    # --- Grupo de acreditación / N° operación / Clave de cruce ---
    grupo_original = df["Grupo de acreditación"]
    mask_vacio = (
        grupo_original.isna() |
        grupo_original.astype(str).str.strip().isin(["", "nan", "None", "-"])
    )
    # Tipo de acreditación: si Nave nunca agrupó la operación (grupo vacío/"-")
    # vs. si vino con un código de lote real.
    df["Tipo de acreditación"] = mask_vacio.map({True: "Individual", False: "Grupo"})

    # col_operacion puede ser numérico (arrastra ".0", ej. 127585397.0) o alfanumérico
    # (ej. "JIU213094047", formato que Nave empezó a usar en meses más recientes): si
    # es numérico, se limpia el ".0"; si no, se usa el texto tal cual.
    raw_str = df[col_operacion].astype(str).str.strip()
    numeros_operacion = pd.to_numeric(df[col_operacion], errors="coerce")
    limpio = numeros_operacion.astype("Int64").astype(str)
    limpio = limpio.where(numeros_operacion.notna(), raw_str)
    df["N° operación"] = limpio

    grupo_limpio = grupo_original.astype(str).str.strip()
    df["Clave de cruce"] = grupo_limpio.where(~mask_vacio, limpio)
    # "Grupo de acreditación" queda crudo (vacío de verdad, no relleno) para reporte.
    df["Grupo de acreditación"] = grupo_limpio.where(~mask_vacio, None)

    # --- Redondeo de columnas float (algunos meses no traen todas) ---
    cols_float = [
        "Monto bruto",
        "Costo del servicio",
        "IVA del costo",
        "Retención IIBB CABA",
        "Monto neto",
        "Propina",
    ]
    for col in cols_float:
        if col not in df.columns:
            df[col] = 0.0
        df[col] = df[col].round(2)

    return df



# ─────────────────────────────────────────────────────────────────────────────
# ETIQUETAS DE MES
# ─────────────────────────────────────────────────────────────────────────────

_MESES_ABREV = {
    1: "ENE", 2: "FEB", 3: "MAR", 4: "ABR", 5: "MAY", 6: "JUN",
    7: "JUL", 8: "AGO", 9: "SEP", 10: "OCT", 11: "NOV", 12: "DIC",
}
_MESES_COMPLETOS = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


def _mes_predominante(df: pd.DataFrame, columna_fecha: str) -> tuple[int, int]:
    """Año y mes con más filas en `columna_fecha` (para titular las hojas)."""
    fechas = pd.to_datetime(df[columna_fecha], errors="coerce").dropna()
    if fechas.empty:
        raise ValueError(f"No se pudo determinar el mes: '{columna_fecha}' no tiene fechas válidas.")
    periodo = fechas.dt.to_period("M").mode().iloc[0]
    return periodo.year, periodo.month


def _etiqueta_abreviada(anio: int, mes: int) -> str:
    return f"{_MESES_ABREV[mes]} {anio}"


def _etiqueta_completa(anio: int, mes: int) -> str:
    return f"{_MESES_COMPLETOS[mes]} {anio}"


def _mes_bare(etiqueta_completa: str | None) -> str:
    """De "Diciembre 2025" saca "diciembre" (para textos tipo "Pendiente diciembre")."""
    if not etiqueta_completa:
        return "mes anterior"
    return etiqueta_completa.split()[0].lower()


def _formato_ars(valor: float) -> str:
    signo = "-" if valor < 0 else ""
    texto = f"{abs(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{signo}${texto}"


def _clave_orden(clave) -> tuple[int, float | str]:
    try:
        return (0, float(clave))
    except (TypeError, ValueError):
        return (1, str(clave))


# ─────────────────────────────────────────────────────────────────────────────
# PREPARACIÓN DE HOJAS "CRUDAS" (BANCO / NAVE)
# ─────────────────────────────────────────────────────────────────────────────

def preparar_banco_reporte(df_banco_acred: pd.DataFrame) -> pd.DataFrame:
    """Reduce el extracto de banco depurado a las columnas de la plantilla."""
    fecha = pd.to_datetime(df_banco_acred["fecha"], errors="coerce")
    fecha_txt = fecha.apply(lambda f: f.isoformat() if pd.notna(f) else "")
    return pd.DataFrame({
        "Fecha": fecha_txt,
        "Descripción": df_banco_acred["descripcion"],
        "Conciliación": df_banco_acred["conciliacion"],
        "Clave de cruce": df_banco_acred["leyenda adicional1"],
        "Importe neto": df_banco_acred["importe"],
    })


def preparar_nave_reporte(df_nave_dep: pd.DataFrame) -> pd.DataFrame:
    """Reduce el reporte de Nave depurado a las columnas de la plantilla."""
    fecha_op = df_nave_dep["Fecha de operación"].dt.strftime("%d/%m/%Y %H:%M").fillna("")
    fecha_acred = df_nave_dep["Fecha de acreditación"].dt.strftime("%d/%m/%Y").fillna("")
    return pd.DataFrame({
        "Fecha de operación": fecha_op,
        "Fecha de acreditación": fecha_acred,
        "N° operación": df_nave_dep["N° operación"],
        "Grupo acreditación": df_nave_dep["Grupo de acreditación"],
        "Clave de cruce": df_nave_dep["Clave de cruce"],
        "Monto neto": df_nave_dep["Monto neto"],
        "Estado Nave": df_nave_dep["Estado"].fillna("").astype(str).str.strip(),
    })


# ─────────────────────────────────────────────────────────────────────────────
# PENDIENTES ANTERIORES
# ─────────────────────────────────────────────────────────────────────────────

def calcular_pendientes_anteriores(
    df_nave_anterior_dep: pd.DataFrame,
    df_banco_anterior_acred: pd.DataFrame,
    tolerancia: float = 1.0,
) -> pd.DataFrame:
    """
    Claves que Nave registró en el mes anterior pero el banco todavía no
    había acreditado a su cierre (Banco ≈ 0 para esa clave, Nave != 0).
    """
    nave_sum = df_nave_anterior_dep.groupby("Clave de cruce")["Monto neto"].sum()
    banco_sum = df_banco_anterior_acred.groupby("leyenda adicional1")["importe"].sum()
    banco_alineado = banco_sum.reindex(nave_sum.index).fillna(0.0)

    mask_pendiente = (banco_alineado.abs() <= tolerancia) & (nave_sum.abs() > tolerancia)
    pendientes = nave_sum[mask_pendiente]

    df = pd.DataFrame({
        "Clave": pendientes.index.astype(str),
        "Importe Nave pendiente": pendientes.values.round(2),
        "Observación manual": "",
    })
    df = df.assign(_orden=df["Clave"].map(_clave_orden)).sort_values("_orden")
    df = df.drop(columns="_orden").reset_index(drop=True)
    return df


def remapear_pendientes_via_acreditacion(
    pendientes_anteriores: pd.DataFrame,
    df_nave_acred_mes_dep: pd.DataFrame | None,
) -> pd.DataFrame:
    """
    Un cupón individual (sin grupo) que quedó pendiente el mes anterior a
    veces el banco lo termina acreditando adentro de un grupo nuevo,
    armado recién este mes — un grupo que no existía el mes anterior, así
    que el cruce por clave no lo encuentra ahí (pendientes_anteriores lo
    tiene bajo su viejo N° de operación individual, pero CRUCE AUTOMÁTICO
    de este mes solo conoce la clave del grupo nuevo) y aparece una
    "Diferencia Banco vs Nave" en el grupo que en realidad no es tal: el
    cupón sí se cobró, solo que mezclado en un grupo que no existía antes.

    Si se pasa el extracto de Nave "por fecha de acreditación" del mes
    actual (trae, agrupado, todo lo que se acreditó este mes sin importar
    cuándo pasó la operación), buscamos ahí, por N° de operación, a qué
    grupo terminó perteneciendo cada pendiente individual y reasignamos su
    "Clave" a la del grupo nuevo — así CRUCE AUTOMÁTICO lo suma dentro de
    ese grupo en vez de dejarlo perdido bajo el número de operación viejo.
    Si no se pasa el archivo, o un pendiente no aparece ahí, queda como
    estaba (bajo su propio N° de operación).
    """
    pendientes = pendientes_anteriores.copy()
    if df_nave_acred_mes_dep is None or pendientes.empty:
        return pendientes

    mapa_clave = (
        df_nave_acred_mes_dep
        .drop_duplicates(subset="N° operación")
        .set_index("N° operación")["Clave de cruce"]
    )
    pendientes["Clave"] = pendientes["Clave"].map(lambda c: mapa_clave.get(c, c))

    # Varios pendientes individuales pueden terminar en el mismo grupo nuevo: sumarlos.
    pendientes = pendientes.groupby("Clave", as_index=False).agg(
        **{
            "Importe Nave pendiente": ("Importe Nave pendiente", "sum"),
            "Observación manual": ("Observación manual", lambda s: "; ".join(x for x in s if x)),
        }
    )
    pendientes = pendientes.assign(_orden=pendientes["Clave"].map(_clave_orden)).sort_values("_orden")
    pendientes = pendientes.drop(columns="_orden").reset_index(drop=True)
    return pendientes


def completar_pendientes_via_acreditacion(
    pendientes_anteriores: pd.DataFrame,
    df_nave_acred_mes_dep: pd.DataFrame | None,
    anio_anterior: int | None,
    mes_anterior_num: int | None,
) -> pd.DataFrame:
    """
    Cupones que salieron el mes anterior, sin grupo (individuales), y el
    banco los termina acreditando dentro de un grupo armado recién este
    mes — un grupo que el mes anterior ni siquiera existía. Nave tampoco
    llegó a marcarles una Fecha de acreditación en el mes anterior (el
    grupo no estaba cerrado todavía), así que ningún extracto de Nave del
    mes anterior — ni "por fecha de cobro" ni "por fecha de acreditación"
    — los puede mostrar como pendientes: simplemente no existen ahí. No
    son diferencias reales, son cupones que cambiaron de clave.

    La única forma de encontrarlos es mirando el extracto de Nave "por
    fecha de acreditación" DE ESTE MES: como trae agrupado todo lo que se
    acreditó este mes sin importar cuándo pasó la operación, filtrando
    las filas cuya Fecha de operación cae en el mes anterior se obtiene,
    directamente y ya resuelta a la clave del grupo nuevo, la plata que
    el mes anterior dejó pendiente sin que ese pendiente hubiera podido
    detectarse en su momento.

    Primero remapea (remapear_pendientes_via_acreditacion) los pendientes
    que ya se habían detectado el mes anterior y el banco terminó
    acreditando en un grupo nuevo; después agrega, como pendientes
    nuevos, las claves que salgan de ese filtro y todavía no estén
    contempladas (para no duplicar plata ya explicada por otro lado).
    """
    pendientes = remapear_pendientes_via_acreditacion(pendientes_anteriores, df_nave_acred_mes_dep)
    if df_nave_acred_mes_dep is None or anio_anterior is None or mes_anterior_num is None:
        return pendientes

    mask_mes_anterior = (
        (df_nave_acred_mes_dep["Fecha de operación"].dt.year == anio_anterior) &
        (df_nave_acred_mes_dep["Fecha de operación"].dt.month == mes_anterior_num)
    )
    derivadas = df_nave_acred_mes_dep[mask_mes_anterior].groupby("Clave de cruce")["Monto neto"].sum()

    claves_existentes = set(pendientes["Clave"])
    nuevas = derivadas[~derivadas.index.isin(claves_existentes)]
    if len(nuevas):
        extra = pd.DataFrame({
            "Clave": nuevas.index,
            "Importe Nave pendiente": nuevas.values.round(2),
            "Observación manual": "Detectado vía Nave por fecha de acreditación",
        })
        pendientes = pd.concat([pendientes, extra], ignore_index=True)

    pendientes = pendientes.assign(_orden=pendientes["Clave"].map(_clave_orden)).sort_values("_orden")
    pendientes = pendientes.drop(columns="_orden").reset_index(drop=True)
    return pendientes


# ─────────────────────────────────────────────────────────────────────────────
# CRUCE AUTOMÁTICO
# ─────────────────────────────────────────────────────────────────────────────

def calcular_cruce_automatico(
    df_nave_actual_dep: pd.DataFrame,
    df_banco_actual_acred: pd.DataFrame,
    pendientes_anteriores: pd.DataFrame,
    tolerancia: float = 1.0,
) -> pd.DataFrame:
    """
    Una fila por "Clave de cruce" (unión Nave/Banco del mes actual) con
    Banco, Nave, Diferencia (Banco - Nave), Pendiente mes anterior, "mes
    anterior acreditado en mes actual" (= Pendiente), Diferencia residual
    (lo que le queda a la Diferencia después de aplicar el pendiente) y la
    Categoría de conciliación:
        - "Cupón mes anterior acreditado en banco actual": la clave venía
          pendiente del mes anterior (haya cerrado del todo o no).
        - "Conciliado": Banco == Nave.
        - "Pendiente actual por acreditar": Nave sin banco todavía.
        - "Banco sin Nave / revisar": banco sin Nave (a investigar).
        - "Diferencia Banco vs Nave": ambos lados con datos pero no cierran.
    """
    nave_sum = df_nave_actual_dep.groupby("Clave de cruce")["Monto neto"].sum()
    banco_sum = df_banco_actual_acred.groupby("leyenda adicional1")["importe"].sum()

    pendientes_map = dict(zip(
        pendientes_anteriores["Clave"].astype(str),
        pendientes_anteriores["Importe Nave pendiente"],
    ))

    claves = sorted(set(nave_sum.index) | set(banco_sum.index), key=_clave_orden)

    filas = []
    for clave in claves:
        banco_val = float(banco_sum.get(clave, 0.0))
        nave_val = float(nave_sum.get(clave, 0.0))
        diferencia = banco_val - nave_val
        pendiente = float(pendientes_map.get(clave, 0.0))

        if clave in pendientes_map:
            categoria = "Cupón mes anterior acreditado en banco actual"
            residual = diferencia - pendiente
        elif abs(diferencia) <= tolerancia:
            categoria = "Conciliado"
            residual = 0.0
        elif abs(banco_val) <= tolerancia:
            categoria = "Pendiente actual por acreditar"
            residual = 0.0
        elif abs(nave_val) <= tolerancia:
            categoria = "Banco sin Nave / revisar"
            residual = 0.0
        else:
            categoria = "Diferencia Banco vs Nave"
            residual = 0.0

        filas.append({
            "Clave de cruce": clave,
            "Banco": round(banco_val, 2),
            "Nave": round(nave_val, 2),
            "Diferencia": round(diferencia, 2),
            "Pendiente mes anterior": round(pendiente, 2),
            "mes anterior acreditado en mes actual": round(pendiente, 2),
            "Diferencia residual": round(residual, 2),
            "Categoría conciliación": categoria,
        })

    return pd.DataFrame(filas)


_COLUMNAS_AJUSTE_DIAGNOSTICO = ["Retención IIBB CABA", "IVA del costo", "Costo del servicio", "Propina"]


def agregar_diagnostico_diferencia(
    cruce: pd.DataFrame,
    df_nave_actual_dep: pd.DataFrame,
    tolerancia: float = 1.0,
) -> pd.DataFrame:
    """
    Agrega la columna "Diagnóstico diferencia" a CRUCE AUTOMÁTICO: para las
    claves en "Diferencia Banco vs Nave" (Banco y Nave presentes pero no
    cierran), prueba si restarle al Monto neto de Nave alguna combinación de
    columnas de ajuste (Retención IIBB CABA, IVA del costo, Costo del
    servicio, Propina — la misma lógica de ajuste que usaba el cruce viejo)
    hace que cierre contra el banco dentro de la tolerancia.

    Es solo diagnóstico manual, no cambia la Categoría conciliación: si
    encuentra una combinación que explica la diferencia, la deja como
    comentario (p.ej. "Retención IIBB CABA + IVA del costo"); si ninguna
    combinación la explica, deja la celda en blanco.

    No aplica lo mismo para "Cupones individuales" vs "Grupo": la búsqueda
    de ajuste es idéntica para ambos, por clave, sin distinguir tipo.
    """
    cruce = cruce.copy()
    cruce["Diagnóstico diferencia"] = ""

    columnas = [c for c in _COLUMNAS_AJUSTE_DIAGNOSTICO if c in df_nave_actual_dep.columns]
    if not columnas:
        return cruce

    ajustes_sum = df_nave_actual_dep.groupby("Clave de cruce")[columnas].sum()
    candidatas = cruce[cruce["Categoría conciliación"] == "Diferencia Banco vs Nave"]

    diagnosticos = {}
    for _, fila in candidatas.iterrows():
        clave = fila["Clave de cruce"]
        if clave not in ajustes_sum.index:
            continue
        valores_ajuste = ajustes_sum.loc[clave]
        nave_val = fila["Nave"]
        banco_val = fila["Banco"]

        for r in range(1, len(columnas) + 1):
            encontrado = False
            for combo in combinations(columnas, r):
                ajustado = nave_val - sum(valores_ajuste[c] for c in combo)
                if abs(ajustado - banco_val) <= tolerancia:
                    diagnosticos[clave] = " + ".join(combo)
                    encontrado = True
                    break
            if encontrado:
                break

    if diagnosticos:
        cruce["Diagnóstico diferencia"] = cruce["Clave de cruce"].map(diagnosticos).fillna("")
    return cruce


def _tipo_por_conteo(df_nave_dep: pd.DataFrame) -> pd.Series:
    """"Grupo" si la clave tiene más de una fila de Nave, "Individual" si tiene una sola."""
    conteo = df_nave_dep.groupby("Clave de cruce").size()
    return conteo.map(lambda c: "Grupo" if c > 1 else "Individual")


# ─────────────────────────────────────────────────────────────────────────────
# CONCILIACIÓN DIFERENCIA
# ─────────────────────────────────────────────────────────────────────────────

def calcular_conciliacion_diferencia(
    cruce: pd.DataFrame,
    df_nave_actual_dep: pd.DataFrame,
    df_nave_anterior_dep: pd.DataFrame | None,
    mes_anterior: str | None,
    mes_actual: str,
    df_nave_acred_mes_dep: pd.DataFrame | None = None,
) -> dict:
    """
    Importes de la cascada "CONCILIACIÓN DIFERENCIA": parte de la
    diferencia total Banco - Nave y la va explicando categoría por
    categoría (de CRUCE AUTOMÁTICO) hasta llegar a $0.

    df_nave_acred_mes_dep (opcional): el extracto de Nave "por fecha de
    acreditación" del mes actual. Un grupo que se formó recién este mes
    (ver completar_pendientes_via_acreditacion) no tiene ninguna fila en
    df_nave_anterior_dep, así que el conteo Grupo/Individual para el
    texto "$X grupos + $Y individuales" no lo puede clasificar ahí — se
    usa este extracto como respaldo para esos casos.
    """
    banco_total = cruce["Banco"].sum()
    nave_total = cruce["Nave"].sum()
    diferencia_total = banco_total - nave_total

    cat_mes_anterior = cruce[cruce["Categoría conciliación"] == "Cupón mes anterior acreditado en banco actual"]
    cat_pendiente = cruce[cruce["Categoría conciliación"] == "Pendiente actual por acreditar"]
    cat_diferencias = cruce[cruce["Categoría conciliación"] == "Diferencia Banco vs Nave"]
    cat_banco_sin_nave = cruce[cruce["Categoría conciliación"] == "Banco sin Nave / revisar"]

    importe_acreditado_mes_anterior = cat_mes_anterior["mes anterior acreditado en mes actual"].sum()
    importe_residual_grupos_mixtos = cat_mes_anterior["Diferencia residual"].sum()

    if df_nave_anterior_dep is not None and len(cat_mes_anterior):
        tipo_anterior = _tipo_por_conteo(df_nave_anterior_dep)
        tipo_por_clave = cat_mes_anterior["Clave de cruce"].map(tipo_anterior)
        if df_nave_acred_mes_dep is not None:
            tipo_acred_mes = _tipo_por_conteo(df_nave_acred_mes_dep)
            tipo_por_clave = tipo_por_clave.fillna(cat_mes_anterior["Clave de cruce"].map(tipo_acred_mes))
        tipo_por_clave = tipo_por_clave.fillna("Individual")
    else:
        tipo_por_clave = pd.Series("Individual", index=cat_mes_anterior.index)
    monto_grupos_acreditados = cat_mes_anterior.loc[
        tipo_por_clave == "Grupo", "mes anterior acreditado en mes actual"
    ].sum()
    monto_individuales_acreditados = cat_mes_anterior.loc[
        tipo_por_clave == "Individual", "mes anterior acreditado en mes actual"
    ].sum()

    tipo_actual = _tipo_por_conteo(df_nave_actual_dep)
    tipo_por_clave_pendiente = cat_pendiente["Clave de cruce"].map(tipo_actual).fillna("Individual")
    importe_pendiente_grupos = cat_pendiente.loc[tipo_por_clave_pendiente == "Grupo", "Diferencia"].sum()
    importe_pendiente_individuales = cat_pendiente.loc[tipo_por_clave_pendiente == "Individual", "Diferencia"].sum()

    importe_diferencias = cat_diferencias["Diferencia"].sum()
    importe_banco_sin_nave = cat_banco_sin_nave["Diferencia"].sum()

    mes_ant_bare = _mes_bare(mes_anterior)
    mes_act_bare = _mes_bare(mes_actual)

    return {
        "diferencia_total": diferencia_total,
        "importe_acreditado_mes_anterior": importe_acreditado_mes_anterior,
        "importe_residual_grupos_mixtos": importe_residual_grupos_mixtos,
        "desc_acreditado_mes_anterior": (
            f"{_formato_ars(monto_grupos_acreditados)} grupos + "
            f"{_formato_ars(monto_individuales_acreditados)} individuales"
        ),
        "desc_residual": f"Diferencia que pertenece a {mes_act_bare}",
        "importe_pendiente_grupos": importe_pendiente_grupos,
        "importe_pendiente_individuales": importe_pendiente_individuales,
        "importe_diferencias": importe_diferencias,
        "importe_banco_sin_nave": importe_banco_sin_nave,
        "mes_anterior_bare": mes_ant_bare,
        "mes_actual_bare": mes_act_bare,
    }


# ─────────────────────────────────────────────────────────────────────────────
# ESTILOS
# ─────────────────────────────────────────────────────────────────────────────

_COLOR_TITULO = "FF17365D"
_FUENTE_TITULO = Font(name="Carlito", size=16, bold=True, color="FFFFFFFF")
_FUENTE_HEADER = Font(name="Carlito", size=11, bold=True, color="FFFFFFFF")
_FUENTE_DATO = Font(name="Carlito", size=11)
_FUENTE_TOTAL = Font(name="Carlito", size=11, bold=True)
_RELLENO_TITULO = PatternFill(start_color=_COLOR_TITULO, end_color=_COLOR_TITULO, fill_type="solid")
_LADO_FINO = Side(style="thin")
_BORDE_FINO = Border(left=_LADO_FINO, right=_LADO_FINO, top=_LADO_FINO, bottom=_LADO_FINO)


def _escribir_encabezado_hoja(ws, titulo: str, headers: list[str]) -> None:
    n_cols = len(headers)
    ws.sheet_view.showGridLines = False
    ws.append([titulo] + [None] * (n_cols - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    for celda in ws[1]:
        celda.font = _FUENTE_TITULO
        celda.fill = _RELLENO_TITULO
    ws.row_dimensions[1].height = 28
    ws.append([None] * n_cols)
    ws.append(headers)
    for celda in ws[3]:
        celda.font = _FUENTE_HEADER
        celda.fill = _RELLENO_TITULO
        celda.border = _BORDE_FINO
        celda.alignment = Alignment(vertical="center")
    ws.row_dimensions[3].height = 27.6


def _escribir_tabla(
    wb, nombre_hoja: str, titulo: str, df: pd.DataFrame,
    formatos: dict[str, str] | None = None, anchos: dict[str, float] | None = None,
):
    ws = wb.create_sheet(nombre_hoja)
    headers = list(df.columns)
    _escribir_encabezado_hoja(ws, titulo, headers)

    for _, fila in df.iterrows():
        valores = ["" if pd.isna(fila[col]) else fila[col] for col in headers]
        ws.append(valores)

    formatos = formatos or {}
    fila_inicio, fila_fin = 4, 3 + len(df)
    for i, col in enumerate(headers, start=1):
        fmt = formatos.get(col)
        for fila_idx in range(fila_inicio, fila_fin + 1):
            celda = ws.cell(row=fila_idx, column=i)
            celda.border = _BORDE_FINO
            celda.font = _FUENTE_DATO
            if fmt:
                celda.number_format = fmt

    anchos = anchos or {}
    for i, col in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = anchos.get(col, 18)
    return ws


def _escribir_cruce_automatico(wb, cruce: pd.DataFrame, mes_anterior: str | None, mes_actual: str) -> None:
    ws = wb.create_sheet("CRUCE AUTOMÁTICO")

    mes_ant_bare = _mes_bare(mes_anterior)
    mes_act_bare = _mes_bare(mes_actual)
    col_pendiente = f"Pendiente {mes_ant_bare}"
    col_acreditado = f"{mes_ant_bare.capitalize()} acreditado en {mes_act_bare}"

    headers = [
        "Clave de cruce", "Banco", "Nave", "Diferencia",
        col_pendiente, col_acreditado, "Diferencia residual", "Categoría conciliación",
        "Diagnóstico diferencia",
    ]
    _escribir_encabezado_hoja(ws, f"Cruce automático — {mes_actual}", headers)

    tiene_diagnostico = "Diagnóstico diferencia" in cruce.columns
    for _, fila in cruce.iterrows():
        ws.append([
            fila["Clave de cruce"], fila["Banco"], fila["Nave"], fila["Diferencia"],
            fila["Pendiente mes anterior"], fila["mes anterior acreditado en mes actual"],
            fila["Diferencia residual"], fila["Categoría conciliación"],
            fila["Diagnóstico diferencia"] if tiene_diagnostico else "",
        ])

    n = len(cruce)
    fila_inicio, fila_fin = 4, 3 + n
    for fila_idx in range(fila_inicio, fila_fin + 1):
        for col_idx in range(1, 10):
            celda = ws.cell(row=fila_idx, column=col_idx)
            celda.border = _BORDE_FINO
            celda.font = _FUENTE_DATO
            if 2 <= col_idx <= 7:
                celda.number_format = "#,##0.00"

    fila_total = fila_fin + 2  # una fila en blanco antes del total
    ws.cell(row=fila_total, column=1, value="TOTAL VISIBLE")
    for col_idx in range(2, 8):
        letra = get_column_letter(col_idx)
        celda = ws.cell(row=fila_total, column=col_idx, value=f"=SUBTOTAL(109,{letra}{fila_inicio}:{letra}{fila_fin})")
        celda.number_format = "#,##0.00"
    ws.cell(row=fila_total, column=8, value="Se actualiza al filtrar")
    for col_idx in range(1, 10):
        celda = ws.cell(row=fila_total, column=col_idx)
        celda.font = _FUENTE_TOTAL
        celda.border = _BORDE_FINO

    if n:
        rango_cf = f"A{fila_inicio}:I{fila_fin}"
        columna_cat = f"$H{fila_inicio}"
        reglas = [
            ("mes anterior acreditado", "FFD9EAF7"),
            ("Pendiente actual", "FFF4CCCC"),
            ("Diferencia", "FFFFF2CC"),
            ("Conciliado", "FFE2F0D9"),
        ]
        for texto, color in reglas:
            relleno = PatternFill(start_color=color, end_color=color, fill_type="solid")
            formula = f'ISNUMBER(SEARCH("{texto}",{columna_cat}))'
            ws.conditional_formatting.add(rango_cf, FormulaRule(formula=[formula], fill=relleno))

    anchos = {1: 20, 2: 18, 3: 18, 4: 18, 5: 22, 6: 26, 7: 20, 8: 46, 9: 40}
    for col_idx, ancho in anchos.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho


def _escribir_conciliacion_diferencia(wb, datos: dict, mes_anterior: str | None, mes_actual: str) -> None:
    ws = wb.create_sheet("CONCILIACIÓN DIFERENCIA")
    headers = ["Categoría", "Importe", "Signo", "Aplicación", "Control acumulado", "Resultado"]
    _escribir_encabezado_hoja(ws, "Explicación completa de la diferencia Banco vs Nave", headers)

    mes_ant_bare = datos["mes_anterior_bare"]
    mes_act_bare = datos["mes_actual_bare"]

    filas = [
        ("Diferencia Banco - Nave", datos["diferencia_total"], -1, "Punto de partida"),
        (f"Cupones {mes_ant_bare} acreditados en {mes_act_bare}",
         datos["importe_acreditado_mes_anterior"], 1, datos["desc_acreditado_mes_anterior"]),
        ("Residual de grupos mixtos", datos["importe_residual_grupos_mixtos"], 1, datos["desc_residual"]),
        (f"Cupones {mes_act_bare} por acreditar - grupos", datos["importe_pendiente_grupos"], 1, "Resta"),
        ("Diferencias Banco vs Nave", datos["importe_diferencias"], 1, "Suma"),
        ("Banco sin Nave", datos["importe_banco_sin_nave"], 1, "Suma"),
        (f"Cupones {mes_act_bare} por acreditar - individuales", datos["importe_pendiente_individuales"], 1, "Resta"),
    ]

    fila_inicio = 4
    for offset, (categoria, importe, signo, aplicacion) in enumerate(filas):
        fila = fila_inicio + offset
        ws.cell(row=fila, column=1, value=categoria)
        ws.cell(row=fila, column=2, value=round(float(importe), 2))
        ws.cell(row=fila, column=3, value=signo)
        ws.cell(row=fila, column=4, value=aplicacion)
        formula_control = f"=C{fila}*B{fila}" if fila == fila_inicio else f"=E{fila - 1}+C{fila}*B{fila}"
        ws.cell(row=fila, column=5, value=formula_control)

    fila_final = fila_inicio + len(filas) - 1
    ws.cell(
        row=fila_final, column=6,
        value=f'=IF(ROUND(E{fila_final},2)=0,"CONCILIADO","REVISAR: "&TEXT(E{fila_final},"#,##0.00"))',
    )

    for fila_idx in range(fila_inicio, fila_final + 1):
        for col_idx in range(1, 7):
            celda = ws.cell(row=fila_idx, column=col_idx)
            celda.border = _BORDE_FINO
            celda.font = _FUENTE_DATO
            if col_idx in (2, 5):
                celda.number_format = "#,##0.00"

    # Casillero de control: retoma los totales por categoría de CRUCE AUTOMÁTICO
    fila_pie = fila_final + 5
    pie = [
        ("Pendiente actual por acreditar", f"=B{fila_inicio + 3}+B{fila_final}"),
        ("Diferencia Banco vs Nave", f"=B{fila_inicio + 4}"),
        ("Banco sin Nave / revisar", f"=B{fila_inicio + 5}"),
        ("Cupón mes anterior acreditado en banco actual", f"=B{fila_inicio + 2}"),
        (None, f"=B{fila_inicio + 1}"),
    ]
    for offset, (etiqueta, formula) in enumerate(pie):
        fila = fila_pie + offset
        if etiqueta:
            celda_a = ws.cell(row=fila, column=1, value=etiqueta)
            celda_a.font = _FUENTE_DATO
        celda_b = ws.cell(row=fila, column=2, value=formula)
        celda_b.font = _FUENTE_DATO
        celda_b.number_format = "#,##0.00"

    anchos = {1: 43, 2: 20, 3: 12, 4: 60, 5: 22, 6: 18}
    for col_idx, ancho in anchos.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho


# ─────────────────────────────────────────────────────────────────────────────
# ARMADO DEL EXCEL COMPLETO
# ─────────────────────────────────────────────────────────────────────────────

_FORMATOS_BANCO = {"Importe neto": "#,##0.00"}
_ANCHOS_BANCO = {"Fecha": 14, "Descripción": 34, "Conciliación": 20, "Clave de cruce": 16, "Importe neto": 18}

_FORMATOS_NAVE = {"Monto neto": "#,##0.00"}
_ANCHOS_NAVE = {
    "Fecha de operación": 21, "Fecha de acreditación": 18, "N° operación": 16,
    "Grupo acreditación": 20, "Clave de cruce": 16, "Monto neto": 18, "Estado Nave": 16,
}

_FORMATOS_PENDIENTES = {"Importe Nave pendiente": "#,##0.00"}
_ANCHOS_PENDIENTES = {"Clave": 20, "Importe Nave pendiente": 22, "Observación manual": 48}


def generar_excel_plantilla_cupones(
    df_banco_actual_acred: pd.DataFrame,
    df_nave_actual_dep: pd.DataFrame,
    mes_actual: str,
    df_banco_anterior_acred: pd.DataFrame | None = None,
    df_nave_anterior_dep: pd.DataFrame | None = None,
    mes_anterior: str | None = None,
    mes_anterior_hoja: str | None = None,
    df_nave_acred_mes_dep: pd.DataFrame | None = None,
    tolerancia: float = 1.0,
) -> tuple[bytes, pd.DataFrame]:
    """
    Arma el workbook con el mismo formato, hojas, encabezados y fórmulas que
    "Plantilla_Automatizacion_Conciliacion_Nave_FINAL". Retorna (bytes del
    excel, DataFrame de CRUCE AUTOMÁTICO — para las stats de la UI).

    df_nave_acred_mes_dep (opcional): el extracto de Nave "por fecha de
    acreditación" del mes actual — se usa para completar los pendientes
    del mes anterior con cupones que el banco terminó acreditando dentro
    de un grupo nuevo (ver completar_pendientes_via_acreditacion), y se
    vuelca tal cual en la hoja "Nave x Acred Mes" si se pasa.
    """
    con_anterior = df_banco_anterior_acred is not None and df_nave_anterior_dep is not None

    if con_anterior:
        pendientes = calcular_pendientes_anteriores(df_nave_anterior_dep, df_banco_anterior_acred, tolerancia)
        anio_ant, mes_ant = _mes_predominante(df_nave_anterior_dep, "Fecha de operación")
        pendientes = completar_pendientes_via_acreditacion(
            pendientes, df_nave_acred_mes_dep, anio_ant, mes_ant,
        )
    else:
        pendientes = pd.DataFrame(columns=["Clave", "Importe Nave pendiente", "Observación manual"])

    cruce = calcular_cruce_automatico(df_nave_actual_dep, df_banco_actual_acred, pendientes, tolerancia)
    cruce = agregar_diagnostico_diferencia(cruce, df_nave_actual_dep, tolerancia)
    datos_conciliacion = calcular_conciliacion_diferencia(
        cruce, df_nave_actual_dep, df_nave_anterior_dep if con_anterior else None,
        mes_anterior, mes_actual, df_nave_acred_mes_dep,
    )

    wb = Workbook()
    wb.remove(wb.active)

    if con_anterior:
        etiqueta_hoja_ant = mes_anterior_hoja or mes_anterior
        _escribir_tabla(
            wb, f"BANCO {etiqueta_hoja_ant}", f"BANCO {etiqueta_hoja_ant}",
            preparar_banco_reporte(df_banco_anterior_acred), _FORMATOS_BANCO, _ANCHOS_BANCO,
        )
        _escribir_tabla(
            wb, f"NAVE {etiqueta_hoja_ant}", f"NAVE {etiqueta_hoja_ant}",
            preparar_nave_reporte(df_nave_anterior_dep), _FORMATOS_NAVE, _ANCHOS_NAVE,
        )

    _escribir_tabla(wb, "BANCO ACTUAL", "BANCO ACTUAL",
                     preparar_banco_reporte(df_banco_actual_acred), _FORMATOS_BANCO, _ANCHOS_BANCO)
    _escribir_tabla(wb, "NAVE ACTUAL", "NAVE ACTUAL",
                     preparar_nave_reporte(df_nave_actual_dep), _FORMATOS_NAVE, _ANCHOS_NAVE)

    if df_nave_acred_mes_dep is not None:
        _escribir_tabla(wb, "Nave x Acred Mes", "NAVE POR FECHA DE ACREDITACIÓN — MES ACTUAL",
                         preparar_nave_reporte(df_nave_acred_mes_dep), _FORMATOS_NAVE, _ANCHOS_NAVE)

    if con_anterior:
        _escribir_tabla(wb, "PENDIENTES ANTERIORES", "Pendientes del mes anterior",
                         pendientes, _FORMATOS_PENDIENTES, _ANCHOS_PENDIENTES)

    _escribir_cruce_automatico(wb, cruce, mes_anterior, mes_actual)
    _escribir_conciliacion_diferencia(wb, datos_conciliacion, mes_anterior, mes_actual)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), cruce


# ─────────────────────────────────────────────────────────────────────────────
# PIPELINE COMPLETO
# ─────────────────────────────────────────────────────────────────────────────

def correr_conciliacion_cupones_plantilla(
    archivo_banco_actual,
    archivo_nave_actual,
    archivo_banco_anterior=None,
    archivo_nave_anterior=None,
    archivo_nave_acred_mes=None,
    tolerancia: float = 1.0,
) -> tuple[bytes, dict]:
    """
    archivo_nave_acred_mes (opcional): el extracto de Nave "por fecha de
    acreditación" del mes actual. Se usa para resolver cupones individuales
    que quedaron pendientes el mes anterior y el banco terminó acreditando
    adentro de un grupo nuevo (armado recién este mes) — ver
    remapear_pendientes_via_acreditacion(). También se vuelca tal cual en
    la hoja "Nave x Acred Mes" si se pasa.
    """
    df_banco_actual_acred = depurar_leyenda(importar_extracto_banco(archivo_banco_actual))
    df_nave_actual_dep = depurar_nave(importar_reporte_nave(archivo_nave_actual))
    anio_act, mes_act = _mes_predominante(df_nave_actual_dep, "Fecha de operación")
    mes_actual = _etiqueta_completa(anio_act, mes_act)

    con_anterior = archivo_banco_anterior is not None and archivo_nave_anterior is not None
    df_banco_anterior_acred = df_nave_anterior_dep = None
    mes_anterior = mes_anterior_hoja = None
    if con_anterior:
        df_banco_anterior_acred = depurar_leyenda(importar_extracto_banco(archivo_banco_anterior))
        df_nave_anterior_dep = depurar_nave(importar_reporte_nave(archivo_nave_anterior))
        anio_ant, mes_ant = _mes_predominante(df_nave_anterior_dep, "Fecha de operación")
        mes_anterior = _etiqueta_completa(anio_ant, mes_ant)
        mes_anterior_hoja = _etiqueta_abreviada(anio_ant, mes_ant)

    df_nave_acred_mes_dep = None
    if archivo_nave_acred_mes is not None:
        df_nave_acred_mes_dep = depurar_nave(importar_reporte_nave(archivo_nave_acred_mes))

    buf, cruce = generar_excel_plantilla_cupones(
        df_banco_actual_acred, df_nave_actual_dep, mes_actual,
        df_banco_anterior_acred, df_nave_anterior_dep, mes_anterior, mes_anterior_hoja,
        df_nave_acred_mes_dep, tolerancia=tolerancia,
    )

    conteo = cruce["Categoría conciliación"].value_counts().to_dict()
    stats = {
        "mes_actual": mes_actual,
        "mes_anterior": mes_anterior,
        "conciliado": conteo.get("Conciliado", 0),
        "mes_anterior_acreditado": conteo.get("Cupón mes anterior acreditado en banco actual", 0),
        "pendiente_actual": conteo.get("Pendiente actual por acreditar", 0),
        "diferencia_banco_vs_nave": conteo.get("Diferencia Banco vs Nave", 0),
        "banco_sin_nave": conteo.get("Banco sin Nave / revisar", 0),
        "banco_total": round(float(cruce["Banco"].sum()), 2),
        "nave_total": round(float(cruce["Nave"].sum()), 2),
    }
    return buf, stats
