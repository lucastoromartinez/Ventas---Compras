import re
from io import BytesIO

import pandas as pd
from rapidfuzz import fuzz, process
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from logica_percepciones import load_excel_sistema, load_excel_arca, _parse_nro
from proveedores import PADRON_PROVEEDORES


# ─────────────────────────────────────────────
# DEPURACIÓN SISTEMA
# ─────────────────────────────────────────────

def depurar_sistema_retenciones(df: pd.DataFrame, padron: dict | None = None) -> pd.DataFrame:
    """
    Depura el reporte del sistema:
    - Calcula 'Importe' = Debe - Haber.
    - A partir de 'Su Factura' genera 'Pto. Venta' y 'N°Comprobante'.
    - Genera 'CUIT' buscando, para cada 'Tercero', qué nombre del padrón de
      proveedores del repo (proveedores.py) coincide (comparación exacta sin
      espacios ni mayúsculas), y le asigna el primer CUIT encontrado para
      ese nombre.
    - Conserva 'Serie': define si la fila se cruza línea a línea (Serie
      cargada) o de forma sumarizada por proveedor (Serie vacía).
    """
    if padron is None:
        padron = PADRON_PROVEEDORES

    df = df.copy()
    df.columns = df.columns.astype(str).str.strip()

    norm_cols = {c.lower(): c for c in df.columns}

    def _resolver(candidatos: list[str]) -> str:
        for cand in candidatos:
            k = cand.lower()
            if k in norm_cols:
                return norm_cols[k]
        raise KeyError(f"No se encontró ninguna columna de {candidatos}. Disponibles: {list(df.columns)}")

    c_debe = _resolver(["Debe"])
    c_haber = _resolver(["Haber"])
    c_factura = _resolver(["Su Factura", "Su factura"])
    c_tercero = _resolver(["Tercero"])
    c_serie = _resolver(["Serie"])

    df[c_debe] = pd.to_numeric(df[c_debe], errors="coerce").fillna(0).round(2)
    df[c_haber] = pd.to_numeric(df[c_haber], errors="coerce").fillna(0).round(2)
    df["Importe"] = (df[c_debe] - df[c_haber]).round(2)

    df[["Pto. Venta", "N°Comprobante"]] = pd.DataFrame(
        df[c_factura].apply(_parse_nro).tolist(), index=df.index
    )

    nombre_a_cuit = {}
    for cuit, nombre in padron.items():
        clave = str(nombre).strip().upper()
        nombre_a_cuit.setdefault(clave, cuit)

    df["CUIT"] = df[c_tercero].astype(str).str.strip().str.upper().map(nombre_a_cuit)

    if c_tercero != "Tercero":
        df = df.rename(columns={c_tercero: "Tercero"})
    if c_serie != "Serie":
        df = df.rename(columns={c_serie: "Serie"})

    return df


# ─────────────────────────────────────────────
# DEPURACIÓN ARCA
# ─────────────────────────────────────────────

def depurar_arca_retenciones(df: pd.DataFrame) -> pd.DataFrame:
    """
    Depura el reporte de retenciones de ARCA:
    - Normaliza 'CUIT' a texto.
    - Redondea 'Monto Retenido' a 2 decimales.
    - Separa 'N° Comprobante' en 'Pto. Venta' y 'N°Comprobante' (8 dígitos).
    """
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip()

    norm_cols = {c.lower(): c for c in df.columns}

    def _resolver(candidatos: list[str]) -> str:
        for cand in candidatos:
            k = cand.lower()
            if k in norm_cols:
                return norm_cols[k]
        raise KeyError(f"No se encontró ninguna columna de {candidatos}. Disponibles: {list(df.columns)}")

    c_cuit = _resolver(["CUIT"])
    c_monto = _resolver(["Monto Retenido"])
    c_comp = _resolver(["N° Comprobante", "N°Comprobante", "Nro Comprobante", "Numero Comprobante"])

    if pd.api.types.is_numeric_dtype(df[c_cuit]):
        df[c_cuit] = df[c_cuit].astype("Int64").astype(str)
    else:
        df[c_cuit] = df[c_cuit].astype(str).str.replace(r"[^0-9]", "", regex=True)
    if c_cuit != "CUIT":
        df = df.rename(columns={c_cuit: "CUIT"})

    df[c_monto] = pd.to_numeric(df[c_monto], errors="coerce").round(2)
    if c_monto != "Monto Retenido":
        df = df.rename(columns={c_monto: "Monto Retenido"})

    def _separar_comprobante(valor):
        if pd.isna(valor):
            return pd.Series([pd.NA, pd.NA])

        s = str(valor).strip()
        if s.endswith(".0"):
            s = s[:-2]
        if s[:3] == "A0 ":
            s = s[3:]
        s = re.sub(r"\D", "", s)

        if s == "":
            return pd.Series([pd.NA, pd.NA])

        if len(s) < 9:
            comprobante = s.zfill(8)
            pto_venta = pd.NA
        else:
            comprobante = s[-8:]
            pto_venta = s[:-8]

        return pd.Series([pto_venta, comprobante])

    df[["Pto. Venta", "N°Comprobante"]] = df[c_comp].apply(_separar_comprobante)
    df = df.drop(columns=[c_comp])

    return df


# ─────────────────────────────────────────────
# CRUCE
# ─────────────────────────────────────────────

def cruce_retenciones(
    df_arca: pd.DataFrame,
    df_sistema: pd.DataFrame,
    tolerancia_importe: float = 1.0,
    score_nombre_min: int = 70,
    padron: dict | None = None,
):
    """
    Cruza retenciones de ARCA contra el sistema. A diferencia de percepciones,
    las retenciones no siempre tienen comprobante asociado: las filas del
    sistema con 'Serie' cargada se cruzan línea a línea; las filas con
    'Serie' vacía se cruzan de forma sumarizada, agrupando varias líneas de
    ARCA del mismo proveedor contra una sola línea del sistema cuya suma de
    'Monto Retenido' cierra el 'Importe'.

    PASO 1 (Serie no nula, línea a línea):
      1a. CUIT + Importe exacto.
      1b. CUIT vía padrón + Importe (mismo proveedor con más de un CUIT
          cargado en proveedores.py).
      1c. Fuzzy por nombre + Importe, para el remanente (score >= score_nombre_min).

    PASO 2 (Serie nula, sumarizado por proveedor):
      2a. Agrupa ARCA libres por CUIT real; matchea contra el CUIT (generado
          vía padrón) de la fila del sistema si la suma cierra.
      2b. Si no hay grupo por CUIT, busca en el padrón el nombre asociado al
          CUIT de cada grupo de ARCA y lo compara contra 'Tercero' del sistema.
      2c. Si tampoco, fuzzy por nombre entre los grupos de ARCA restantes y
          'Tercero' del sistema (probando los 4 mejores candidatos).

    Retorna:
        tuple de 4 pd.DataFrame:
            - df_match_sistema, df_match_arca (ligados por 'id_match', con
              'match_tipo' y 'match_score')
            - df_falta_sistema (líneas de ARCA sin match)
            - df_falta_arca (líneas de sistema sin match)
    """
    if padron is None:
        padron = PADRON_PROVEEDORES

    df_arca = df_arca.reset_index(drop=True).copy()
    df_sistema = df_sistema.reset_index(drop=True).copy()

    def _buscar_columna(df, nombre):
        return next(
            (c for c in df.columns if str(c).replace(".", "").strip().upper() == nombre),
            None
        )

    def _normalizar_nombre(nombre):
        return re.sub(r"\s+", "", str(nombre)).strip().upper()

    col_cuit_arca = _buscar_columna(df_arca, "CUIT")
    col_cuit_sistema = _buscar_columna(df_sistema, "CUIT")
    col_serie = _buscar_columna(df_sistema, "SERIE")
    if col_serie is None:
        raise ValueError("No se encontró la columna 'Serie' en df_sistema.")

    cuit_arca_norm = df_arca[col_cuit_arca].astype(str).str.replace(r"[^0-9]", "", regex=True)
    cuit_sistema_norm = df_sistema[col_cuit_sistema].astype(str).str.replace(r"[^0-9]", "", regex=True)

    idx_serie_con = set(df_sistema.index[df_sistema[col_serie].notna()])
    idx_serie_vacia = set(df_sistema.index[df_sistema[col_serie].isna()])

    matches = []  # (a_idx, s_idx, tipo, score) -- PASO 1, línea a línea
    matches_grupo = []  # (lista_a_idx, s_idx, tipo, score) -- PASO 2, sumarizado

    # ============================================================
    # PASO 1: Serie no nula (línea a línea)
    # ============================================================
    sistema_disponibles = set(idx_serie_con)

    # 1a: CUIT + Importe
    for a_idx in df_arca.index:
        cuit_a = cuit_arca_norm.loc[a_idx]
        monto_a = df_arca.loc[a_idx, "Monto Retenido"]

        candidatos_idx = [
            i for i in sistema_disponibles
            if cuit_sistema_norm.loc[i] == cuit_a
            and abs(df_sistema.loc[i, "Importe"] - monto_a) <= tolerancia_importe
        ]
        if not candidatos_idx:
            continue

        s_idx = candidatos_idx[0]
        matches.append((a_idx, s_idx, "cuit", 100))
        sistema_disponibles.remove(s_idx)

    # 1b: CUIT vía padrón + Importe (duplicados de CUIT para el mismo nombre)
    arca_matcheados = {a_idx for a_idx, s_idx, tipo, score in matches}
    arca_pendientes = [i for i in df_arca.index if i not in arca_matcheados]

    for a_idx in arca_pendientes:
        cuit_a = cuit_arca_norm.loc[a_idx]
        nombre_padron = padron.get(cuit_a)
        if not nombre_padron:
            continue

        monto_a = df_arca.loc[a_idx, "Monto Retenido"]
        candidatos_idx = [
            i for i in sistema_disponibles
            if abs(df_sistema.loc[i, "Importe"] - monto_a) <= tolerancia_importe
            and _normalizar_nombre(df_sistema.loc[i, "Tercero"]) == _normalizar_nombre(nombre_padron)
        ]
        if not candidatos_idx:
            continue

        s_idx = candidatos_idx[0]
        matches.append((a_idx, s_idx, "cuit_padron", 100))
        sistema_disponibles.remove(s_idx)
        arca_matcheados.add(a_idx)

    # 1c: Fuzzy por nombre + Importe, para el remanente
    arca_pendientes = [i for i in df_arca.index if i not in arca_matcheados]

    for a_idx in arca_pendientes:
        monto_a = df_arca.loc[a_idx, "Monto Retenido"]
        nombre_a = str(df_arca.loc[a_idx, "Razon Social"])

        candidatos_idx = [
            i for i in sistema_disponibles
            if abs(df_sistema.loc[i, "Importe"] - monto_a) <= tolerancia_importe
        ]
        if not candidatos_idx:
            continue

        puntajes = [(i, fuzz.token_sort_ratio(nombre_a, str(df_sistema.loc[i, "Tercero"])))
                    for i in candidatos_idx]
        mejor_score = max(p[1] for p in puntajes)
        if mejor_score < score_nombre_min:
            continue

        s_idx = next(i for i, sc in puntajes if sc == mejor_score)
        matches.append((a_idx, s_idx, "nombre", mejor_score))
        sistema_disponibles.remove(s_idx)
        arca_matcheados.add(a_idx)

    # ============================================================
    # PASO 2: Serie vacía (sumarizado por proveedor)
    # ============================================================
    arca_libres = set(df_arca.index) - arca_matcheados

    def _grupos_por_cuit(indices):
        grupos = {}
        for a_idx in indices:
            grupos.setdefault(cuit_arca_norm.loc[a_idx], []).append(a_idx)
        return grupos

    def _grupos_por_nombre(indices):
        grupos = {}
        for a_idx in indices:
            grupos.setdefault(str(df_arca.loc[a_idx, "Razon Social"]), []).append(a_idx)
        return grupos

    for s_idx in idx_serie_vacia:
        importe_sistema = df_sistema.loc[s_idx, "Importe"]

        # 2a: agrupado por CUIT exacto
        grupos_cuit = _grupos_por_cuit(arca_libres)
        cuit_sistema_val = cuit_sistema_norm.loc[s_idx]
        grupo_arca = grupos_cuit.get(cuit_sistema_val, []) if cuit_sistema_val else []

        if grupo_arca:
            suma_arca = sum(df_arca.loc[i, "Monto Retenido"] for i in grupo_arca)
            if abs(suma_arca - importe_sistema) <= tolerancia_importe:
                matches_grupo.append((grupo_arca, s_idx, "cuit", 100))
                arca_libres -= set(grupo_arca)
                continue

        # 2b: CUIT vía padrón (mismo proveedor, CUIT distinto por duplicados)
        tercero_sistema_norm = _normalizar_nombre(df_sistema.loc[s_idx, "Tercero"])
        grupos_cuit_libres = _grupos_por_cuit(arca_libres)
        grupo_padron = None
        for cuit_grupo, indices_grupo in grupos_cuit_libres.items():
            nombre_padron = padron.get(cuit_grupo)
            if nombre_padron and _normalizar_nombre(nombre_padron) == tercero_sistema_norm:
                grupo_padron = indices_grupo
                break

        if grupo_padron:
            suma_arca = sum(df_arca.loc[i, "Monto Retenido"] for i in grupo_padron)
            if abs(suma_arca - importe_sistema) <= tolerancia_importe:
                matches_grupo.append((grupo_padron, s_idx, "cuit_padron", 100))
                arca_libres -= set(grupo_padron)
                continue

        # 2c: fuzzy por nombre entre los grupos de ARCA restantes
        grupos_nombre = _grupos_por_nombre(arca_libres)
        if not grupos_nombre:
            continue

        nombres_grupos = list(grupos_nombre.keys())
        candidatos_nombre = process.extract(
            str(df_sistema.loc[s_idx, "Tercero"]), nombres_grupos,
            scorer=fuzz.token_sort_ratio, limit=4
        )

        for nombre_grupo, score, _ in candidatos_nombre:
            if score < score_nombre_min:
                continue
            grupo_arca_c = grupos_nombre[nombre_grupo]
            suma_arca = sum(df_arca.loc[i, "Monto Retenido"] for i in grupo_arca_c)
            if abs(suma_arca - importe_sistema) <= tolerancia_importe:
                matches_grupo.append((grupo_arca_c, s_idx, "nombre", score))
                arca_libres -= set(grupo_arca_c)
                break

    # ============================================================
    # Armado de resultados finales
    # ============================================================
    filas_sistema, filas_arca = [], []
    id_match = 0

    for a_idx, s_idx, tipo, score in matches:
        id_match += 1
        fila_s = df_sistema.loc[s_idx].copy()
        fila_s["id_match"] = id_match
        fila_s["match_tipo"] = tipo
        fila_s["match_score"] = score
        filas_sistema.append(fila_s)

        fila_a = df_arca.loc[a_idx].copy()
        fila_a["id_match"] = id_match
        fila_a["match_tipo"] = tipo
        fila_a["match_score"] = score
        filas_arca.append(fila_a)

    for grupo_arca, s_idx, tipo, score in matches_grupo:
        id_match += 1
        tipo_grupo = f"{tipo}_sumarizado"
        fila_s = df_sistema.loc[s_idx].copy()
        fila_s["id_match"] = id_match
        fila_s["match_tipo"] = tipo_grupo
        fila_s["match_score"] = score
        filas_sistema.append(fila_s)

        for a_idx in grupo_arca:
            fila_a = df_arca.loc[a_idx].copy()
            fila_a["id_match"] = id_match
            fila_a["match_tipo"] = tipo_grupo
            fila_a["match_score"] = score
            filas_arca.append(fila_a)

    cols_extra = ["id_match", "match_tipo", "match_score"]

    if filas_sistema:
        df_match_sistema = pd.DataFrame(filas_sistema).reset_index(drop=True)
        df_match_sistema = df_match_sistema[[c for c in df_match_sistema.columns if c not in cols_extra] + cols_extra]
    else:
        df_match_sistema = pd.DataFrame(columns=list(df_sistema.columns) + cols_extra)

    if filas_arca:
        df_match_arca = pd.DataFrame(filas_arca).reset_index(drop=True)
        df_match_arca = df_match_arca[[c for c in df_match_arca.columns if c not in cols_extra] + cols_extra]
    else:
        df_match_arca = pd.DataFrame(columns=list(df_arca.columns) + cols_extra)

    arca_matcheados_final = arca_matcheados | {a_idx for grupo, _, _, _ in matches_grupo for a_idx in grupo}
    sistema_matcheados_final = {s_idx for _, s_idx, _, _ in matches} | {s_idx for _, s_idx, _, _ in matches_grupo}

    idx_falta_sistema = [i for i in df_arca.index if i not in arca_matcheados_final]
    df_falta_sistema = df_arca.loc[idx_falta_sistema].reset_index(drop=True)

    idx_falta_arca = [i for i in df_sistema.index if i not in sistema_matcheados_final]
    df_falta_arca = df_sistema.loc[idx_falta_arca].reset_index(drop=True)

    return df_match_sistema, df_match_arca, df_falta_sistema, df_falta_arca


# ─────────────────────────────────────────────
# PROVEEDORES NUEVOS (matchearon por nombre, no están en el padrón)
# ─────────────────────────────────────────────

def detectar_proveedores_nuevos(
    df_match_arca: pd.DataFrame,
    padron: dict | None = None,
    umbral_nombre: int = 80,
) -> pd.DataFrame:
    """
    Devuelve los CUIT que matchearon por nombre (no por CUIT) con score >=
    umbral_nombre y que todavía no están en el padrón de proveedores del repo.
    """
    if padron is None:
        padron = PADRON_PROVEEDORES

    candidatos = df_match_arca[
        df_match_arca["match_tipo"].isin(["nombre", "nombre_sumarizado"])
        & (df_match_arca["match_score"] >= umbral_nombre)
    ]

    nuevos = []
    vistos = set()
    for _, fila in candidatos.iterrows():
        cuit = fila["CUIT"]
        if cuit in padron or cuit in vistos:
            continue
        vistos.add(cuit)
        nuevos.append({"CUIT": cuit, "Nombre": fila["Razon Social"], "score": fila["match_score"]})

    return pd.DataFrame(nuevos, columns=["CUIT", "Nombre", "score"])


# ─────────────────────────────────────────────
# EXPORTAR A BUFFER EN MEMORIA (descargable único)
# ─────────────────────────────────────────────

def generar_excel_retenciones(
    df_match_sistema: pd.DataFrame,
    df_match_arca: pd.DataFrame,
    df_falta_sistema: pd.DataFrame,
    df_falta_arca: pd.DataFrame,
    df_proveedores_nuevos: pd.DataFrame | None = None,
) -> bytes:
    DATE_FORMAT = "DD/MM/YYYY"
    DATE_COLS = {"Fecha", "Fecha factura", "Fecha Percepcion", "Fecha Comprobante"}

    def _style_sheet(ws) -> None:
        thin = Side(style="thin")
        header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        header_font = Font(bold=True)
        header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, cell in enumerate(ws[1], start=1):
            if str(cell.value).strip() in DATE_COLS:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col_idx).number_format = DATE_FORMAT

        for col in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0 for c in col),
                default=10,
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)

    hojas = {
        "Match_Sistema": df_match_sistema,
        "Match_Arca": df_match_arca,
        "Falta_Sistema": df_falta_sistema,
        "Falta_Arca": df_falta_arca,
    }
    if df_proveedores_nuevos is not None and not df_proveedores_nuevos.empty:
        hojas["Proveedores_Nuevos"] = df_proveedores_nuevos

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for nombre_hoja, df in hojas.items():
            df.to_excel(writer, sheet_name=nombre_hoja, index=False)

        wb = writer.book
        for nombre_hoja in hojas:
            _style_sheet(wb[nombre_hoja])

    return buf.getvalue()


# ─────────────────────────────────────────────
# PIPELINE COMPLETO
# ─────────────────────────────────────────────

def correr_cruce_retenciones(
    archivo_arca, archivo_sistema, tolerancia_importe: float = 1.0, score_nombre_min: int = 70
):
    df_sistema = load_excel_sistema(archivo_sistema)
    df_arca = load_excel_arca(archivo_arca)

    df_sistema_dep = depurar_sistema_retenciones(df_sistema)
    df_arca_dep = depurar_arca_retenciones(df_arca)

    df_match_sistema, df_match_arca, df_falta_sistema, df_falta_arca = cruce_retenciones(
        df_arca_dep, df_sistema_dep, tolerancia_importe=tolerancia_importe, score_nombre_min=score_nombre_min
    )

    df_proveedores_nuevos = detectar_proveedores_nuevos(df_match_arca, umbral_nombre=max(score_nombre_min, 80))

    stats = {
        "match": len(df_match_arca),
        "faltante_sistema": len(df_falta_sistema),
        "faltante_arca": len(df_falta_arca),
        "proveedores_nuevos": len(df_proveedores_nuevos),
    }

    buf_reporte = generar_excel_retenciones(
        df_match_sistema, df_match_arca, df_falta_sistema, df_falta_arca, df_proveedores_nuevos
    )

    return buf_reporte, stats
