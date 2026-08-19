import re
from io import BytesIO

import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from logica_percepciones import load_excel_sistema
from proveedores import PADRON_PROVEEDORES


# ─────────────────────────────────────────────
# CARGA
# ─────────────────────────────────────────────

_PATRON_TXT_SIRCREB = re.compile(
    r'^.{22}'                          # prefijo: se descarta
    r'(?P<cuit>\d{2}-\d+-\d)'          # CUIT: nn-nnnnnnnn-n
    r'(?P<fecha>\d{2}/\d{2}/\d{4})'    # Fecha: dd/mm/aaaa
    r'(?P<codigo>\d{6})'               # código: se descarta
    r'(?P<importe_ent>\d{7})'          # importe entero (7 dígitos, con ceros)
    r',(?P<dec>\d{2})$'                # decimales del importe
)


def cargar_txt_sircreb(file, padron: dict | None = None) -> pd.DataFrame:
    """
    Parsea el archivo de texto de SIRCREB, con formato:
    PREFIJO(22, se descarta) + CUIT(13) + FECHA(10, dd/mm/aaaa) +
    CODIGO(6, se descarta) + IMPORTE(7 dígitos + coma + 2 decimales).

    - CUIT: sin guiones, como texto.
    - Nombre: se busca el CUIT en el padrón de proveedores del repo
      (proveedores.py). Queda en None si el CUIT no está cargado.

    Retorna:
        pd.DataFrame con columnas ['cuit', 'fecha', 'importe', 'Nombre'].
    """
    if padron is None:
        padron = PADRON_PROVEEDORES

    file.seek(0)
    contenido = file.read()
    if isinstance(contenido, bytes):
        contenido = contenido.decode("utf-8")

    rows = []
    for line in contenido.splitlines():
        line = line.rstrip()
        if not line.strip():
            continue
        m = _PATRON_TXT_SIRCREB.match(line)
        if m:
            rows.append(m.groupdict())

    if not rows:
        raise ValueError("No se pudo interpretar ninguna línea del archivo SIRCREB (formato inesperado).")

    df = pd.DataFrame(rows)

    df["fecha"] = pd.to_datetime(df["fecha"], format="%d/%m/%Y")
    df["importe"] = (df["importe_ent"] + "." + df["dec"]).astype(float)
    df["cuit"] = df["cuit"].str.replace("-", "", regex=False)
    df["Nombre"] = df["cuit"].map(padron)

    return df[["cuit", "fecha", "importe", "Nombre"]]


# ─────────────────────────────────────────────
# DEPURACIÓN SISTEMA
# ─────────────────────────────────────────────

def depurar_sistema_sircreb(
    df_sircreb: pd.DataFrame, df_ret: pd.DataFrame, padron: dict | None = None
) -> pd.DataFrame:
    """
    Combina y depura los reportes del sistema (SIRCREB + retenciones IIBB):
    - Descarta filas donde 'Asiento', 'Apunte' o 'Fecha' estén vacías.
    - Combina ambos reportes en uno solo.
    - Calcula 'Importe' = Debe - Haber.
    - Si el reporte combinado ya trae una columna CUIT (sin importar
      mayúsculas o minúsculas), la depura como siempre (sin guiones ni
      espacios) y la usa directamente, sin pasar por el padrón. Si no viene,
      genera 'CUIT' a partir de 'Tercero', buscando en el padrón de
      proveedores del repo (proveedores.py) qué nombre coincide (comparación
      exacta sin espacios ni mayúsculas), y le asigna el primer CUIT
      encontrado para ese nombre.
    """
    if padron is None:
        padron = PADRON_PROVEEDORES

    def _normalizar(nombre):
        return re.sub(r"\s+", "", str(nombre)).replace(".", "").strip().upper()

    def _buscar_columna(df, objetivo):
        return next((c for c in df.columns if _normalizar(c) == _normalizar(objetivo)), None)

    def _dropear_vacios(df):
        df = df.copy()
        col_asiento = _buscar_columna(df, "Asiento")
        col_apunte = _buscar_columna(df, "Apunte")
        col_fecha = _buscar_columna(df, "Fecha")

        columnas_encontradas = {"Asiento": col_asiento, "Apunte": col_apunte, "Fecha": col_fecha}
        faltantes = {n for n, c in columnas_encontradas.items() if c is None}
        if faltantes:
            raise ValueError(f"Faltan las siguientes columnas: {faltantes}")

        return df.dropna(subset=[col_asiento, col_apunte, col_fecha])

    df_sircreb = _dropear_vacios(df_sircreb)
    df_ret = _dropear_vacios(df_ret)

    df = pd.concat([df_sircreb, df_ret], ignore_index=True)

    col_debe = _buscar_columna(df, "Debe")
    col_haber = _buscar_columna(df, "Haber")
    col_tercero = _buscar_columna(df, "Tercero")

    columnas_encontradas = {"Debe": col_debe, "Haber": col_haber, "Tercero": col_tercero}
    faltantes = {n for n, c in columnas_encontradas.items() if c is None}
    if faltantes:
        raise ValueError(f"Faltan las siguientes columnas en el reporte combinado: {faltantes}")

    df[col_debe] = pd.to_numeric(df[col_debe], errors="coerce").fillna(0)
    df[col_haber] = pd.to_numeric(df[col_haber], errors="coerce").fillna(0)
    df["Importe"] = (df[col_debe] - df[col_haber]).astype("float64").round(2)

    col_cuit = _buscar_columna(df, "CUIT")
    if col_cuit is not None:
        if pd.api.types.is_numeric_dtype(df[col_cuit]):
            df[col_cuit] = df[col_cuit].astype("Int64").astype(str)
        else:
            df[col_cuit] = df[col_cuit].astype(str)
        df[col_cuit] = (
            df[col_cuit]
            .str.replace("-", "", regex=False)
            .str.replace(" ", "", regex=False)
        )
        if col_cuit != "CUIT":
            df = df.rename(columns={col_cuit: "CUIT"})
    else:
        nombre_a_cuit = {}
        for cuit, nombre in padron.items():
            clave = str(nombre).strip().upper()
            nombre_a_cuit.setdefault(clave, cuit)

        df["CUIT"] = df[col_tercero].astype(str).str.strip().str.upper().map(nombre_a_cuit)

    return df.reset_index(drop=True)


# ─────────────────────────────────────────────
# CRUCE
# ─────────────────────────────────────────────

def cruce_sircreb(
    df_arca: pd.DataFrame,
    df_sistema: pd.DataFrame,
    tolerancia_importe: float = 1.0,
):
    """
    Cruza SIRCREB contra el sistema en tres pasadas:

    PASO 1 - CUIT + Importe, uno a uno (tolerancia): para cada línea de
    arca, busca en sistema una línea con el mismo CUIT y el mismo Importe.

    PASO 2 - Sumarizado por CUIT del lado del sistema, contra arca
    remanente: agrupa las líneas de sistema por CUIT, sumando Importe, y
    cruza ese importe sumarizado contra arca por CUIT + Importe.

    PASO 3 - Sumarizado por CUIT de arca, contra sistema remanente: mismo
    criterio que el paso 2 pero a la inversa.

    Retorna:
        tuple de 4 pd.DataFrame:
            - df_match_sistema, df_match_arca (ligados por 'id_match', con
              'match_tipo' = 'directo' / 'sumarizado_sistema' / 'sumarizado_arca')
            - df_falta_sistema (líneas de arca sin match)
            - df_falta_arca (líneas de sistema sin match)
    """
    df_arca = df_arca.reset_index(drop=True).copy()
    df_sistema = df_sistema.reset_index(drop=True).copy()

    def _normalizar(nombre):
        return re.sub(r"\s+", "", str(nombre)).replace(".", "").strip().upper()

    def _buscar_columna(df, *nombres):
        objetivo = {_normalizar(n) for n in nombres}
        return next((c for c in df.columns if _normalizar(c) in objetivo), None)

    col_cuit_arca = _buscar_columna(df_arca, "cuit")
    col_importe_arca = _buscar_columna(df_arca, "importe")
    col_cuit_sistema = _buscar_columna(df_sistema, "CUIT")
    col_importe_sistema = _buscar_columna(df_sistema, "Importe")

    if col_cuit_arca is None or col_importe_arca is None:
        raise ValueError("df_arca necesita columnas de CUIT e Importe.")
    if col_cuit_sistema is None or col_importe_sistema is None:
        raise ValueError("df_sistema necesita columnas de CUIT e Importe.")

    cuit_arca = df_arca[col_cuit_arca].astype(str).str.replace(r"[^0-9]", "", regex=True)
    cuit_sistema = df_sistema[col_cuit_sistema].astype(str).str.replace(r"[^0-9]", "", regex=True)

    matches = []  # (idxs_arca, idxs_sistema, tipo)
    arca_disponibles = set(df_arca.index)
    sistema_disponibles = set(df_sistema.index)

    # PASO 1: CUIT + Importe, 1 a 1
    for a_idx in list(arca_disponibles):
        cuit_a = cuit_arca.loc[a_idx]
        monto_a = df_arca.loc[a_idx, col_importe_arca]

        candidatos = [
            i for i in sistema_disponibles
            if cuit_sistema.loc[i] == cuit_a
            and abs(df_sistema.loc[i, col_importe_sistema] - monto_a) <= tolerancia_importe
        ]
        if not candidatos:
            continue

        s_idx = candidatos[0]
        matches.append(([a_idx], [s_idx], "directo"))
        arca_disponibles.discard(a_idx)
        sistema_disponibles.discard(s_idx)

    # PASO 2: Sumarizado por CUIT del sistema, contra arca remanente
    if sistema_disponibles:
        remanente_sistema = df_sistema.loc[list(sistema_disponibles)]
        sumarizado_sistema = remanente_sistema.groupby(
            cuit_sistema.loc[remanente_sistema.index]
        )[col_importe_sistema].sum()

        for cuit_val, monto_sum in sumarizado_sistema.items():
            candidatos = [
                i for i in arca_disponibles
                if cuit_arca.loc[i] == cuit_val
                and abs(df_arca.loc[i, col_importe_arca] - monto_sum) <= tolerancia_importe
            ]
            if not candidatos:
                continue

            a_idx = candidatos[0]
            idxs_sistema_cuit = [i for i in sistema_disponibles if cuit_sistema.loc[i] == cuit_val]

            matches.append(([a_idx], idxs_sistema_cuit, "sumarizado_sistema"))
            arca_disponibles.discard(a_idx)
            for i in idxs_sistema_cuit:
                sistema_disponibles.discard(i)

    # PASO 3: Sumarizado por CUIT de arca, contra sistema remanente
    if arca_disponibles:
        remanente_arca = df_arca.loc[list(arca_disponibles)]
        sumarizado_arca = remanente_arca.groupby(cuit_arca.loc[remanente_arca.index])[col_importe_arca].sum()

        for cuit_val, monto_sum in sumarizado_arca.items():
            candidatos = [
                i for i in sistema_disponibles
                if cuit_sistema.loc[i] == cuit_val
                and abs(df_sistema.loc[i, col_importe_sistema] - monto_sum) <= tolerancia_importe
            ]
            if not candidatos:
                continue

            s_idx = candidatos[0]
            idxs_arca_cuit = [i for i in arca_disponibles if cuit_arca.loc[i] == cuit_val]

            matches.append((idxs_arca_cuit, [s_idx], "sumarizado_arca"))
            for i in idxs_arca_cuit:
                arca_disponibles.discard(i)
            sistema_disponibles.discard(s_idx)

    # Armado de resultados
    filas_sistema, filas_arca = [], []
    for id_match, (idxs_arca, idxs_sistema, tipo) in enumerate(matches, start=1):
        for i in idxs_sistema:
            fila_sis = df_sistema.loc[i].copy()
            fila_sis["id_match"] = id_match
            fila_sis["match_tipo"] = tipo
            filas_sistema.append(fila_sis)

        for i in idxs_arca:
            fila_a = df_arca.loc[i].copy()
            fila_a["id_match"] = id_match
            fila_a["match_tipo"] = tipo
            filas_arca.append(fila_a)

    cols_extra = ["id_match", "match_tipo"]

    if filas_sistema:
        df_match_sistema = pd.DataFrame(filas_sistema).reset_index(drop=True)
    else:
        df_match_sistema = pd.DataFrame(columns=list(df_sistema.columns) + cols_extra)

    if filas_arca:
        df_match_arca = pd.DataFrame(filas_arca).reset_index(drop=True)
    else:
        df_match_arca = pd.DataFrame(columns=list(df_arca.columns) + cols_extra)

    df_falta_sistema = df_arca.loc[list(arca_disponibles)].reset_index(drop=True)
    df_falta_arca = df_sistema.loc[list(sistema_disponibles)].reset_index(drop=True)

    return df_match_sistema, df_match_arca, df_falta_sistema, df_falta_arca


# ─────────────────────────────────────────────
# EXPORTAR A BUFFER EN MEMORIA (descargable único)
# ─────────────────────────────────────────────

def generar_excel_sircreb(
    df_match_sistema: pd.DataFrame,
    df_match_arca: pd.DataFrame,
    df_falta_sistema: pd.DataFrame,
    df_falta_arca: pd.DataFrame,
) -> bytes:
    DATE_FORMAT = "DD/MM/YYYY"
    DATE_COLS = {"Fecha", "fecha", "Fecha factura"}

    def _style_sheet(ws) -> None:
        thin = Side(style="thin")
        header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        header_font = Font(bold=True)
        header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, cell in enumerate(ws[1], start=1):
            if str(cell.value).strip() in DATE_COLS:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col_idx).number_format = DATE_FORMAT

        for col in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0 for c in col),
                default=10,
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)

    hojas = {
        "Match_Sistema": df_match_sistema,
        "Match_Arca": df_match_arca,
        "Falta_Sistema": df_falta_sistema,
        "Falta_Arca": df_falta_arca,
    }

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for nombre_hoja, df in hojas.items():
            df.to_excel(writer, sheet_name=nombre_hoja, index=False)

        wb = writer.book
        for nombre_hoja in hojas:
            _style_sheet(wb[nombre_hoja])

    return buf.getvalue()


# ─────────────────────────────────────────────
# PIPELINE COMPLETO
# ─────────────────────────────────────────────

def correr_cruce_sircreb(
    archivo_sircreb_txt, archivo_sircreb_excel, archivo_ret_excel, tolerancia_importe: float = 1.0
):
    df_arca = cargar_txt_sircreb(archivo_sircreb_txt)

    df_sircreb_excel = load_excel_sistema(archivo_sircreb_excel)
    df_ret_excel = load_excel_sistema(archivo_ret_excel)
    df_sistema_dep = depurar_sistema_sircreb(df_sircreb_excel, df_ret_excel)

    df_match_sistema, df_match_arca, df_falta_sistema, df_falta_arca = cruce_sircreb(
        df_arca, df_sistema_dep, tolerancia_importe=tolerancia_importe
    )

    stats = {
        "match": len(df_match_arca),
        "faltante_sistema": len(df_falta_sistema),
        "faltante_arca": len(df_falta_arca),
    }

    buf_reporte = generar_excel_sircreb(
        df_match_sistema, df_match_arca, df_falta_sistema, df_falta_arca
    )

    return buf_reporte, stats
