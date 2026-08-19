import re
from io import BytesIO

import pandas as pd
from rapidfuzz import fuzz
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from logica_percepciones import load_excel_sistema
from proveedores import PADRON_PROVEEDORES


# ─────────────────────────────────────────────
# CARGA
# ─────────────────────────────────────────────

_PATRON_TXT_SIRTAC = re.compile(
    r'^(?P<cuit>\d{2}-\d+-\d)'   # CUIT: nn-nnnnnnnn-n
    r'.{16}'                     # periodo(6) + fecha(10): se descartan
    r'(?P<digits>\d{37})'        # bloque de 37 dígitos: comprobante + importe entero
    r',(?P<dec>\d{2})$'          # decimales del importe
)


def cargar_txt_sirtac(file, padron: dict | None = None) -> pd.DataFrame:
    """
    Parsea el archivo de texto SIRTAC de retenciones IIBB PBA, quedándose
    únicamente con lo que se puede verificar con certeza: CUIT e Importe (la
    coma decimal está siempre en posición fija). No se extraen período,
    fecha, punto de venta ni comprobante.

    - CUIT: sin guiones, como texto.
    - Importe: últimos 11 dígitos antes de la coma (parte entera, con ceros
      a la izquierda) + los 2 decimales, convertido a float.
    - nombre: se busca el CUIT en el padrón de proveedores del repo
      (proveedores.py). Queda en None si el CUIT no está cargado.

    Retorna:
        pd.DataFrame con columnas ['cuit', 'importe', 'nombre'].
    """
    if padron is None:
        padron = PADRON_PROVEEDORES

    file.seek(0)
    contenido = file.read()
    if isinstance(contenido, bytes):
        contenido = contenido.decode("utf-8")

    rows = []
    for line in contenido.splitlines():
        line = line.rstrip()
        if not line.strip():
            continue
        m = _PATRON_TXT_SIRTAC.match(line)
        if m:
            rows.append(m.groupdict())

    if not rows:
        raise ValueError("No se pudo interpretar ninguna línea del archivo SIRTAC (formato inesperado).")

    df = pd.DataFrame(rows)

    df["importe"] = (df["digits"].str[-11:] + "." + df["dec"]).astype(float)
    df["cuit"] = df["cuit"].str.replace("-", "", regex=False)
    df["nombre"] = df["cuit"].map(padron)

    return df[["cuit", "importe", "nombre"]]


# ─────────────────────────────────────────────
# DEPURACIÓN SISTEMA
# ─────────────────────────────────────────────

def depurar_sistema_sirtac(df: pd.DataFrame, padron: dict | None = None) -> pd.DataFrame:
    """
    Depura el reporte del sistema:
    - Calcula 'Importe' = Debe - Haber.
    - Si el archivo ya trae una columna CUIT (sin importar mayúsculas o
      minúsculas), la depura como siempre (sin guiones ni espacios) y la usa
      directamente, sin pasar por el padrón. Si no viene, genera 'CUIT'
      buscando, para cada 'Tercero', qué nombre del padrón de proveedores
      del repo (proveedores.py) coincide (comparación exacta sin espacios ni
      mayúsculas), y le asigna el primer CUIT encontrado para ese nombre.
    """
    if padron is None:
        padron = PADRON_PROVEEDORES

    df = df.copy()

    def _normalizar(nombre):
        return re.sub(r"\s+", "", str(nombre)).replace(".", "").strip().upper()

    def _buscar_columna(df, objetivo):
        return next((c for c in df.columns if _normalizar(c) == _normalizar(objetivo)), None)

    col_debe = _buscar_columna(df, "Debe")
    col_haber = _buscar_columna(df, "Haber")
    col_tercero = _buscar_columna(df, "Tercero")

    columnas_encontradas = {"Debe": col_debe, "Haber": col_haber, "Tercero": col_tercero}
    faltantes = {nombre for nombre, col in columnas_encontradas.items() if col is None}
    if faltantes:
        raise ValueError(f"Faltan las siguientes columnas en el sistema: {faltantes}")

    df[col_debe] = pd.to_numeric(df[col_debe], errors="coerce").fillna(0)
    df[col_haber] = pd.to_numeric(df[col_haber], errors="coerce").fillna(0)
    df["Importe"] = (df[col_debe] - df[col_haber]).astype("float64").round(2)

    col_cuit = _buscar_columna(df, "CUIT")
    if col_cuit is not None:
        if pd.api.types.is_numeric_dtype(df[col_cuit]):
            df[col_cuit] = df[col_cuit].astype("Int64").astype(str)
        else:
            df[col_cuit] = df[col_cuit].astype(str)
        df[col_cuit] = (
            df[col_cuit]
            .str.replace("-", "", regex=False)
            .str.replace(" ", "", regex=False)
        )
        if col_cuit != "CUIT":
            df = df.rename(columns={col_cuit: "CUIT"})
    else:
        nombre_a_cuit = {}
        for cuit, nombre in padron.items():
            clave = str(nombre).strip().upper()
            nombre_a_cuit.setdefault(clave, cuit)

        df["CUIT"] = df[col_tercero].astype(str).str.strip().str.upper().map(nombre_a_cuit)

    if col_tercero != "Tercero":
        df = df.rename(columns={col_tercero: "Tercero"})

    return df


# ─────────────────────────────────────────────
# CRUCE
# ─────────────────────────────────────────────

def cruce_sirtac(
    df_sirtac: pd.DataFrame,
    df_sistema: pd.DataFrame,
    tolerancia_importe: float = 1.0,
    score_nombre_min: int = 80,
):
    """
    Cruza df_sirtac contra el sistema en tres pasadas:

    PASO 1 - CUIT + Importe, uno a uno (con tolerancia): para cada línea de
    sirtac busca en sistema una línea con el mismo CUIT y el mismo Importe.

    PASO 2 - Nombre + Importe, uno a uno, para lo que no matcheó en el paso
    1: compara 'nombre' (sirtac) vs 'Tercero' (sistema) con fuzzy matching.
    Solo se acepta si el score >= score_nombre_min y el Importe coincide.

    PASO 3 - Sumarizado por CUIT, para lo que sigue sin matchear: agrupa
    las líneas remanentes de sirtac por CUIT, sumando el Importe, y cruza
    ese importe sumarizado contra el sistema por CUIT + Importe. Si
    matchea, todas las líneas de sirtac de ese CUIT quedan ligadas a esa
    única línea de sistema.

    Retorna:
        tuple de 4 pd.DataFrame:
            - df_match_sistema, df_match_arca (ligados por 'id_match', con
              'match_tipo' = 'cuit' / 'nombre' / 'cuit_sumarizado')
            - df_falta_sistema (líneas de sirtac sin match)
            - df_falta_arca (líneas de sistema sin match)
    """
    df_sirtac = df_sirtac.reset_index(drop=True).copy()
    df_sistema = df_sistema.reset_index(drop=True).copy()

    def _normalizar(nombre):
        return re.sub(r"\s+", "", str(nombre)).replace(".", "").strip().upper()

    def _buscar_columna(df, *nombres):
        objetivo = {_normalizar(n) for n in nombres}
        return next((c for c in df.columns if _normalizar(c) in objetivo), None)

    col_cuit_sirtac = _buscar_columna(df_sirtac, "cuit")
    col_importe_sirtac = _buscar_columna(df_sirtac, "importe")
    col_nombre_sirtac = _buscar_columna(df_sirtac, "nombre")

    col_cuit_sistema = _buscar_columna(df_sistema, "CUIT")
    col_importe_sistema = _buscar_columna(df_sistema, "Importe")
    col_tercero_sistema = _buscar_columna(df_sistema, "Tercero")

    if col_cuit_sirtac is None or col_importe_sirtac is None:
        raise ValueError("df_sirtac necesita columnas de CUIT e Importe.")
    if col_cuit_sistema is None or col_importe_sistema is None:
        raise ValueError("df_sistema necesita columnas de CUIT e Importe.")

    usar_nombre = col_nombre_sirtac is not None and col_tercero_sistema is not None

    cuit_sirtac = df_sirtac[col_cuit_sirtac].astype(str).str.replace(r"[^0-9]", "", regex=True)
    cuit_sistema = df_sistema[col_cuit_sistema].astype(str).str.replace(r"[^0-9]", "", regex=True)

    matches = []  # (lista_sirtac_idx, sistema_idx, tipo, score)
    sirtac_disponibles = set(df_sirtac.index)
    sistema_disponibles = set(df_sistema.index)

    # PASO 1: CUIT + Importe, 1 a 1
    for s_idx in list(sirtac_disponibles):
        cuit_s = cuit_sirtac.loc[s_idx]
        monto_s = df_sirtac.loc[s_idx, col_importe_sirtac]

        candidatos = [
            i for i in sistema_disponibles
            if cuit_sistema.loc[i] == cuit_s
            and abs(df_sistema.loc[i, col_importe_sistema] - monto_s) <= tolerancia_importe
        ]
        if not candidatos:
            continue

        i_match = candidatos[0]
        matches.append(([s_idx], i_match, "cuit", 100))
        sirtac_disponibles.discard(s_idx)
        sistema_disponibles.discard(i_match)

    # PASO 2: Nombre + Importe, 1 a 1, para lo que no matcheó por CUIT
    if usar_nombre:
        for s_idx in list(sirtac_disponibles):
            monto_s = df_sirtac.loc[s_idx, col_importe_sirtac]
            nombre_s = str(df_sirtac.loc[s_idx, col_nombre_sirtac])

            candidatos = [
                i for i in sistema_disponibles
                if abs(df_sistema.loc[i, col_importe_sistema] - monto_s) <= tolerancia_importe
            ]
            if not candidatos:
                continue

            puntajes = [(i, fuzz.token_sort_ratio(nombre_s, str(df_sistema.loc[i, col_tercero_sistema])))
                        for i in candidatos]
            mejor_idx, mejor_score = max(puntajes, key=lambda x: x[1])

            if mejor_score < score_nombre_min:
                continue

            matches.append(([s_idx], mejor_idx, "nombre", mejor_score))
            sirtac_disponibles.discard(s_idx)
            sistema_disponibles.discard(mejor_idx)

    # PASO 3: Sumarizado por CUIT, contra sistema remanente
    if sirtac_disponibles:
        remanente = df_sirtac.loc[list(sirtac_disponibles)]
        sumarizado = remanente.groupby(cuit_sirtac.loc[remanente.index])[col_importe_sirtac].sum()

        for cuit_val, monto_sum in sumarizado.items():
            candidatos = [
                i for i in sistema_disponibles
                if cuit_sistema.loc[i] == cuit_val
                and abs(df_sistema.loc[i, col_importe_sistema] - monto_sum) <= tolerancia_importe
            ]
            if not candidatos:
                continue

            i_match = candidatos[0]
            idxs_cuit = [i for i in sirtac_disponibles if cuit_sirtac.loc[i] == cuit_val]

            matches.append((idxs_cuit, i_match, "cuit_sumarizado", 100))
            for i in idxs_cuit:
                sirtac_disponibles.discard(i)
            sistema_disponibles.discard(i_match)

    # Armado de resultados
    filas_sistema, filas_sirtac = [], []
    for id_match, (idxs_sirtac, s_idx_sistema, tipo, score) in enumerate(matches, start=1):
        fila_sis = df_sistema.loc[s_idx_sistema].copy()
        fila_sis["id_match"] = id_match
        fila_sis["match_tipo"] = tipo
        fila_sis["match_score"] = score
        filas_sistema.append(fila_sis)

        for idx in idxs_sirtac:
            fila_sir = df_sirtac.loc[idx].copy()
            fila_sir["id_match"] = id_match
            fila_sir["match_tipo"] = tipo
            fila_sir["match_score"] = score
            filas_sirtac.append(fila_sir)

    cols_extra = ["id_match", "match_tipo", "match_score"]

    if filas_sistema:
        df_match_sistema = pd.DataFrame(filas_sistema).reset_index(drop=True)
    else:
        df_match_sistema = pd.DataFrame(columns=list(df_sistema.columns) + cols_extra)

    if filas_sirtac:
        df_match_arca = pd.DataFrame(filas_sirtac).reset_index(drop=True)
    else:
        df_match_arca = pd.DataFrame(columns=list(df_sirtac.columns) + cols_extra)

    df_falta_sistema = df_sirtac.loc[list(sirtac_disponibles)].reset_index(drop=True)
    df_falta_arca = df_sistema.loc[list(sistema_disponibles)].reset_index(drop=True)

    return df_match_sistema, df_match_arca, df_falta_sistema, df_falta_arca


# ─────────────────────────────────────────────
# PROVEEDORES NUEVOS (matchearon por nombre, no están en el padrón)
# ─────────────────────────────────────────────

def detectar_proveedores_nuevos(
    df_match_arca: pd.DataFrame,
    padron: dict | None = None,
    umbral_nombre: int = 80,
) -> pd.DataFrame:
    """
    Devuelve los CUIT que matchearon por nombre (no por CUIT) con score >=
    umbral_nombre y que todavía no están en el padrón de proveedores del repo.
    """
    if padron is None:
        padron = PADRON_PROVEEDORES

    candidatos = df_match_arca[
        (df_match_arca["match_tipo"] == "nombre")
        & (df_match_arca["match_score"] >= umbral_nombre)
    ]

    nuevos = [
        {"CUIT": fila["cuit"], "Nombre": fila["nombre"], "score": fila["match_score"]}
        for _, fila in candidatos.iterrows()
        if fila["cuit"] not in padron
    ]

    return pd.DataFrame(nuevos, columns=["CUIT", "Nombre", "score"])


# ─────────────────────────────────────────────
# EXPORTAR A BUFFER EN MEMORIA (descargable único)
# ─────────────────────────────────────────────

def generar_excel_sirtac(
    df_match_sistema: pd.DataFrame,
    df_match_arca: pd.DataFrame,
    df_falta_sistema: pd.DataFrame,
    df_falta_arca: pd.DataFrame,
    df_proveedores_nuevos: pd.DataFrame | None = None,
) -> bytes:
    DATE_FORMAT = "DD/MM/YYYY"
    DATE_COLS = {"Fecha", "Fecha factura", "Fecha Percepcion", "Fecha Comprobante"}

    def _style_sheet(ws) -> None:
        thin = Side(style="thin")
        header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        header_font = Font(bold=True)
        header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, cell in enumerate(ws[1], start=1):
            if str(cell.value).strip() in DATE_COLS:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col_idx).number_format = DATE_FORMAT

        for col in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0 for c in col),
                default=10,
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)

    hojas = {
        "Match_Sistema": df_match_sistema,
        "Match_Arca": df_match_arca,
        "Falta_Sistema": df_falta_sistema,
        "Falta_Arca": df_falta_arca,
    }
    if df_proveedores_nuevos is not None and not df_proveedores_nuevos.empty:
        hojas["Proveedores_Nuevos"] = df_proveedores_nuevos

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for nombre_hoja, df in hojas.items():
            df.to_excel(writer, sheet_name=nombre_hoja, index=False)

        wb = writer.book
        for nombre_hoja in hojas:
            _style_sheet(wb[nombre_hoja])

    return buf.getvalue()


# ─────────────────────────────────────────────
# PIPELINE COMPLETO
# ─────────────────────────────────────────────

def correr_cruce_retenciones_pba(
    archivo_sirtac_txt, archivo_sistema, tolerancia_importe: float = 1.0, score_nombre_min: int = 80
):
    df_sirtac = cargar_txt_sirtac(archivo_sirtac_txt)

    df_sistema = load_excel_sistema(archivo_sistema)
    df_sistema_dep = depurar_sistema_sirtac(df_sistema)

    df_match_sistema, df_match_arca, df_falta_sistema, df_falta_arca = cruce_sirtac(
        df_sirtac, df_sistema_dep, tolerancia_importe=tolerancia_importe, score_nombre_min=score_nombre_min
    )

    df_proveedores_nuevos = detectar_proveedores_nuevos(df_match_arca, umbral_nombre=max(score_nombre_min, 80))

    stats = {
        "match": len(df_match_arca),
        "faltante_sistema": len(df_falta_sistema),
        "faltante_arca": len(df_falta_arca),
        "proveedores_nuevos": len(df_proveedores_nuevos),
    }

    buf_reporte = generar_excel_sirtac(
        df_match_sistema, df_match_arca, df_falta_sistema, df_falta_arca, df_proveedores_nuevos
    )

    return buf_reporte, stats
