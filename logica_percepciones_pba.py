import re
from io import BytesIO

import pandas as pd
from rapidfuzz import fuzz
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from logica_percepciones import load_excel_sistema
from proveedores import PADRON_PROVEEDORES


# ─────────────────────────────────────────────
# CARGA
# ─────────────────────────────────────────────

_PATRON_TXT_ARCA = re.compile(
    r'^(?P<cuit>\d{2}-\d+-\d)'          # CUIT: nn-nnnnnnnn-n
    r'(?P<fecha>\d{2}/\d{2}/\d{4})'      # Fecha: dd/mm/aaaa
    r'(?P<tipo>.{2})'                    # Tipo comprobante: FA, CA, F (espacio), etc.
    r'(?P<comprobante>\d{13})'           # Nro de comprobante (13 dígitos, con ceros)
    r'(?P<importe>-?\d+,\d{2})$'         # Importe: puede tener signo negativo
)


def cargar_txt_arca(file) -> pd.DataFrame:
    """
    Parsea el archivo de texto de percepciones IIBB PBA de ARCA, con formato:
    CUIT(13) + FECHA(10, dd/mm/aaaa) + TIPO(2) + COMPROBANTE(13) + IMPORTE(coma decimal).
    """
    file.seek(0)
    contenido = file.read()
    if isinstance(contenido, bytes):
        contenido = contenido.decode("utf-8")

    rows = []
    for line in contenido.splitlines():
        line = line.rstrip()
        if not line.strip():
            continue
        m = _PATRON_TXT_ARCA.match(line)
        if m:
            rows.append(m.groupdict())

    if not rows:
        raise ValueError("No se pudo interpretar ninguna línea del archivo de ARCA (formato inesperado).")

    df = pd.DataFrame(rows)

    df["fecha"] = pd.to_datetime(df["fecha"], format="%d/%m/%Y")
    df["tipo"] = df["tipo"].str.strip()
    df["importe"] = df["importe"].str.replace(",", ".", regex=False).astype(float)
    df["cuit"] = df["cuit"].str.replace("-", "", regex=False)
    df["n_comprobante"] = df["comprobante"].str[:8].str.lstrip("0").replace("", "0")
    df["pto_venta"] = df["comprobante"].str[8:].str.lstrip("0").replace("", "0")

    return df[["cuit", "fecha", "tipo", "pto_venta", "n_comprobante", "importe"]]


# ─────────────────────────────────────────────
# DEPURACIÓN
# ─────────────────────────────────────────────

def depurar_arca_pba(df_arca: pd.DataFrame, padron: dict | None = None) -> pd.DataFrame:
    """Agrega la columna 'nombre' buscando cada CUIT en el padrón de proveedores del repo."""
    if padron is None:
        padron = PADRON_PROVEEDORES

    df = df_arca.copy()
    df["nombre"] = df["cuit"].map(padron)
    return df


def _parse_nro(s: str) -> tuple[str, str]:
    s = str(s).strip()
    s = s.replace('−', '-').replace('–', '-')  # em dash / en dash
    s = re.sub(r'\s*-\s*', '-', s)
    count = s.count('-')
    if count == 0:
        left, right = s[:4], s[4:]
    elif count == 1:
        left, right = s.split('-', 1)
    else:
        idx = s.index('-')
        left = s[:idx]
        right = s[idx + 1:].replace('-', '')
    pto = left.lstrip('0')
    comp = right.lstrip('0')
    if pto == '' and len(right) >= 3 and right[0] != '0' and right[1:3] == '00':
        pto = right[0]
        comp = right[1:].lstrip('0')
    return pto, comp


def depurar_sistema_pba(df: pd.DataFrame, padron: dict | None = None) -> pd.DataFrame:
    """
    Depura el reporte del sistema:
    - Calcula 'Importe' = Debe - Haber.
    - A partir de 'Su Factura' genera 'Pto. Venta' y 'N°Comprobante'.
    - Si el archivo ya trae una columna CUIT (sin importar mayúsculas o
      minúsculas), la depura como siempre (sin guiones ni espacios) y la usa
      directamente, sin pasar por el padrón. Si no viene, genera 'CUIT'
      buscando, para cada 'Tercero', qué nombre del padrón de proveedores
      del repo coincide (comparación exacta sin espacios ni mayúsculas), y
      le asigna el primer CUIT encontrado para ese nombre.
    """
    if padron is None:
        padron = PADRON_PROVEEDORES

    df = df.copy()

    def _normalizar(nombre):
        return re.sub(r"\s+", "", str(nombre)).strip().upper()

    def _buscar_columna(df, objetivo):
        return next((c for c in df.columns if _normalizar(c) == _normalizar(objetivo)), None)

    col_debe = _buscar_columna(df, "Debe")
    col_haber = _buscar_columna(df, "Haber")
    col_su_factura = _buscar_columna(df, "Su Factura")
    col_tercero = _buscar_columna(df, "Tercero")

    columnas_encontradas = {
        "Debe": col_debe, "Haber": col_haber,
        "Su Factura": col_su_factura, "Tercero": col_tercero,
    }
    faltantes = {nombre for nombre, col in columnas_encontradas.items() if col is None}
    if faltantes:
        raise ValueError(f"Faltan las siguientes columnas en el sistema: {faltantes}")

    df[col_debe] = pd.to_numeric(df[col_debe], errors="coerce").fillna(0).round(2)
    df[col_haber] = pd.to_numeric(df[col_haber], errors="coerce").fillna(0).round(2)
    df["Importe"] = (df[col_debe] - df[col_haber]).round(2)

    df[["Pto. Venta", "N°Comprobante"]] = pd.DataFrame(
        df[col_su_factura].apply(_parse_nro).tolist(), index=df.index
    )

    col_cuit = _buscar_columna(df, "CUIT")
    if col_cuit is not None:
        if pd.api.types.is_numeric_dtype(df[col_cuit]):
            df[col_cuit] = df[col_cuit].astype("Int64").astype(str)
        else:
            df[col_cuit] = df[col_cuit].astype(str)
        df[col_cuit] = (
            df[col_cuit]
            .str.replace("-", "", regex=False)
            .str.replace(" ", "", regex=False)
        )
        if col_cuit != "CUIT":
            df = df.rename(columns={col_cuit: "CUIT"})
    else:
        nombre_a_cuit = {}
        for cuit, nombre in padron.items():
            clave = str(nombre).strip().upper()
            nombre_a_cuit.setdefault(clave, cuit)

        df["CUIT"] = df[col_tercero].astype(str).str.strip().str.upper().map(nombre_a_cuit)

    return df


# ─────────────────────────────────────────────
# CRUCE
# ─────────────────────────────────────────────

def cruce_percepciones_pba(
    df_arca: pd.DataFrame,
    df_sistema: pd.DataFrame,
    tolerancia_importe: float = 1.0,
    score_nombre_min: int = 70,
):
    """
    Cruza percepciones de ARCA (PBA) contra el sistema en dos pasadas:

    PASO 1 - CUIT + Importe: para cada línea de ARCA busca candidatas en
    sistema con el mismo CUIT y el mismo Importe (con tolerancia). Si hay
    más de un candidato, desempata por Pto. Venta / N°Comprobante.

    PASO 2 - Importe + Nombre (fuzzy), para lo que no matcheó en el paso 1:
    cubre el caso de proveedores cargados con un CUIT distinto en el sistema.

    Retorna:
        tuple de 4 pd.DataFrame:
            - df_match_sistema, df_match_arca (ligados por 'id_match', con
              'match_tipo' = 'cuit' o 'nombre', y 'match_score')
            - df_falta_sistema (líneas de ARCA sin match)
            - df_falta_arca (líneas de sistema sin match)
    """
    df_arca = df_arca.reset_index(drop=True).copy()
    df_sistema = df_sistema.reset_index(drop=True).copy()

    def _normalizar(nombre):
        return re.sub(r"\s+", "", str(nombre)).strip().upper()

    def _buscar_columna(df, *nombres):
        objetivo = {_normalizar(n) for n in nombres}
        return next((c for c in df.columns if _normalizar(c) in objetivo), None)

    def _norm_digitos(serie):
        def limpiar(x):
            if pd.isna(x):
                return ""
            s = re.sub(r"\D", "", str(x))
            return str(int(s)) if s else ""
        return serie.apply(limpiar)

    col_cuit_arca = _buscar_columna(df_arca, "CUIT")
    col_cuit_sistema = _buscar_columna(df_sistema, "CUIT")
    col_importe_arca = _buscar_columna(df_arca, "Importe", "Monto Percibido", "Monto")
    col_importe_sistema = _buscar_columna(df_sistema, "Importe", "Monto")
    col_nombre_arca = _buscar_columna(df_arca, "Nombre", "Razon Social")
    col_nombre_sistema = _buscar_columna(df_sistema, "Tercero", "Nombre", "Razon Social", "Proveedor")
    col_pv_arca = _buscar_columna(df_arca, "Pto Venta")
    col_nc_arca = _buscar_columna(df_arca, "N Comprobante")
    col_pv_sistema = _buscar_columna(df_sistema, "Pto Venta")
    col_nc_sistema = _buscar_columna(df_sistema, "N Comprobante")

    if col_cuit_arca is None or col_importe_arca is None:
        raise ValueError("df_arca necesita columnas de CUIT e Importe.")
    if col_importe_sistema is None:
        raise ValueError("df_sistema necesita una columna de Importe.")

    usar_cuit = col_cuit_sistema is not None
    usar_comprobante = all([col_pv_arca, col_nc_arca, col_pv_sistema, col_nc_sistema])
    usar_nombre = col_nombre_arca is not None and col_nombre_sistema is not None

    cuit_arca_norm = cuit_sistema_norm = None
    if usar_cuit:
        cuit_arca_norm = df_arca[col_cuit_arca].astype(str).str.replace(r"[^0-9]", "", regex=True)
        cuit_sistema_norm = df_sistema[col_cuit_sistema].astype(str).str.replace(r"[^0-9]", "", regex=True)

    pv_arca_norm = nc_arca_norm = pv_sistema_norm = nc_sistema_norm = None
    if usar_comprobante:
        pv_arca_norm = _norm_digitos(df_arca[col_pv_arca])
        nc_arca_norm = _norm_digitos(df_arca[col_nc_arca])
        pv_sistema_norm = _norm_digitos(df_sistema[col_pv_sistema])
        nc_sistema_norm = _norm_digitos(df_sistema[col_nc_sistema])

    def desempatar_por_comprobante(arca_idx, candidatos_idx):
        if len(candidatos_idx) <= 1 or not usar_comprobante:
            return candidatos_idx

        pv_a = pv_arca_norm.loc[arca_idx]
        nc_a = nc_arca_norm.loc[arca_idx]

        nivel_a = [i for i in candidatos_idx
                   if pv_sistema_norm.loc[i] == pv_a and nc_sistema_norm.loc[i] == nc_a and nc_a != ""]
        if nivel_a:
            return nivel_a

        nivel_b = [i for i in candidatos_idx if nc_sistema_norm.loc[i] == nc_a and nc_a != ""]
        if nivel_b:
            return nivel_b

        nivel_c = [i for i in candidatos_idx if pv_sistema_norm.loc[i] == pv_a and pv_a != ""]
        if nivel_c:
            return nivel_c

        concat_a = f"{pv_a}{nc_a}"
        mejor_score = -1
        ganadores = []
        for i in candidatos_idx:
            concat_s = f"{pv_sistema_norm.loc[i]}{nc_sistema_norm.loc[i]}"
            score = fuzz.ratio(concat_a, concat_s)
            if score > mejor_score:
                mejor_score, ganadores = score, [i]
            elif score == mejor_score:
                ganadores.append(i)
        return ganadores

    matches = []  # (a_idx, s_idx, tipo, score)
    sistema_disponibles = set(df_sistema.index)

    # PASO 1: CUIT + Importe
    if usar_cuit:
        for a_idx in df_arca.index:
            cuit_a = cuit_arca_norm.loc[a_idx]
            monto_a = df_arca.loc[a_idx, col_importe_arca]

            candidatos_idx = [
                i for i in sistema_disponibles
                if cuit_sistema_norm.loc[i] == cuit_a
                and abs(df_sistema.loc[i, col_importe_sistema] - monto_a) <= tolerancia_importe
            ]
            if not candidatos_idx:
                continue

            candidatos_idx = desempatar_por_comprobante(a_idx, candidatos_idx)
            s_idx = candidatos_idx[0]

            matches.append((a_idx, s_idx, "cuit", 100))
            sistema_disponibles.remove(s_idx)

    # PASO 2: Importe + Nombre (fuzzy), para lo que no matcheó por CUIT
    if usar_nombre:
        arca_matcheados_paso1 = {a_idx for a_idx, s_idx, tipo, score in matches}
        arca_pendientes = [i for i in df_arca.index if i not in arca_matcheados_paso1]

        for a_idx in arca_pendientes:
            monto_a = df_arca.loc[a_idx, col_importe_arca]
            nombre_a = str(df_arca.loc[a_idx, col_nombre_arca])

            candidatos_idx = [
                i for i in sistema_disponibles
                if abs(df_sistema.loc[i, col_importe_sistema] - monto_a) <= tolerancia_importe
            ]
            if not candidatos_idx:
                continue

            puntajes = [(i, fuzz.token_sort_ratio(nombre_a, str(df_sistema.loc[i, col_nombre_sistema])))
                        for i in candidatos_idx]
            mejor_score = max(p[1] for p in puntajes)

            if mejor_score < score_nombre_min:
                continue

            empatados = [i for i, sc in puntajes if sc == mejor_score]
            empatados = desempatar_por_comprobante(a_idx, empatados)
            s_idx = empatados[0]

            matches.append((a_idx, s_idx, "nombre", mejor_score))
            sistema_disponibles.remove(s_idx)

    arca_matcheados = {a_idx for a_idx, s_idx, tipo, score in matches}
    sistema_matcheados = {s_idx for a_idx, s_idx, tipo, score in matches}

    filas_sistema, filas_arca = [], []
    for id_match, (a_idx, s_idx, tipo, score) in enumerate(matches, start=1):
        fila_s = df_sistema.loc[s_idx].copy()
        fila_s["id_match"] = id_match
        fila_s["match_tipo"] = tipo
        fila_s["match_score"] = score
        filas_sistema.append(fila_s)

        fila_a = df_arca.loc[a_idx].copy()
        fila_a["id_match"] = id_match
        fila_a["match_tipo"] = tipo
        fila_a["match_score"] = score
        filas_arca.append(fila_a)

    cols_extra = ["id_match", "match_tipo", "match_score"]

    if filas_sistema:
        df_match_sistema = pd.DataFrame(filas_sistema).reset_index(drop=True)
    else:
        df_match_sistema = pd.DataFrame(columns=list(df_sistema.columns) + cols_extra)

    if filas_arca:
        df_match_arca = pd.DataFrame(filas_arca).reset_index(drop=True)
    else:
        df_match_arca = pd.DataFrame(columns=list(df_arca.columns) + cols_extra)

    df_falta_sistema = df_arca.loc[[i for i in df_arca.index if i not in arca_matcheados]].reset_index(drop=True)
    df_falta_arca = df_sistema.loc[[i for i in df_sistema.index if i not in sistema_matcheados]].reset_index(drop=True)

    return df_match_sistema, df_match_arca, df_falta_sistema, df_falta_arca


# ─────────────────────────────────────────────
# PROVEEDORES NUEVOS (matchearon por nombre, no están en el padrón)
# ─────────────────────────────────────────────

def detectar_proveedores_nuevos(
    df_match_arca: pd.DataFrame,
    padron: dict | None = None,
    umbral_nombre: int = 80,
) -> pd.DataFrame:
    """
    Devuelve los CUIT que matchearon por nombre (no por CUIT) con score >=
    umbral_nombre y que todavía no están en el padrón de proveedores del repo.
    """
    if padron is None:
        padron = PADRON_PROVEEDORES

    candidatos = df_match_arca[
        (df_match_arca["match_tipo"] == "nombre")
        & (df_match_arca["match_score"] >= umbral_nombre)
    ]

    nuevos = [
        {"CUIT": fila["cuit"], "Nombre": fila["nombre"], "score": fila["match_score"]}
        for _, fila in candidatos.iterrows()
        if fila["cuit"] not in padron
    ]

    return pd.DataFrame(nuevos, columns=["CUIT", "Nombre", "score"])


# ─────────────────────────────────────────────
# EXPORTAR A BUFFER EN MEMORIA (descargable único)
# ─────────────────────────────────────────────

def generar_excel_percepciones_pba(
    df_match_sistema: pd.DataFrame,
    df_match_arca: pd.DataFrame,
    df_falta_sistema: pd.DataFrame,
    df_falta_arca: pd.DataFrame,
    df_proveedores_nuevos: pd.DataFrame | None = None,
) -> bytes:
    DATE_FORMAT = "DD/MM/YYYY"
    DATE_COLS = {"fecha", "Fecha", "Fecha_Sistema", "Fecha_arca"}

    def _style_sheet(ws) -> None:
        thin = Side(style="thin")
        header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        header_font = Font(bold=True)
        header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, cell in enumerate(ws[1], start=1):
            if str(cell.value).strip() in DATE_COLS:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col_idx).number_format = DATE_FORMAT

        for col in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0 for c in col),
                default=10,
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)

    hojas = {
        "Match_Sistema": df_match_sistema,
        "Match_Arca": df_match_arca,
        "Falta_Sistema": df_falta_sistema,
        "Falta_Arca": df_falta_arca,
    }
    if df_proveedores_nuevos is not None and not df_proveedores_nuevos.empty:
        hojas["Proveedores_Nuevos"] = df_proveedores_nuevos

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for nombre_hoja, df in hojas.items():
            df.to_excel(writer, sheet_name=nombre_hoja, index=False)

        wb = writer.book
        for nombre_hoja in hojas:
            _style_sheet(wb[nombre_hoja])

    return buf.getvalue()


# ─────────────────────────────────────────────
# PIPELINE COMPLETO
# ─────────────────────────────────────────────

def correr_cruce_percepciones_pba(
    archivo_arca_txt,
    archivo_sistema,
    tolerancia_importe: float = 1.0,
    score_nombre_min: int = 70,
):
    df_arca = cargar_txt_arca(archivo_arca_txt)
    df_arca_dep = depurar_arca_pba(df_arca)

    df_sistema = load_excel_sistema(archivo_sistema)
    df_sistema_dep = depurar_sistema_pba(df_sistema)

    df_match_sistema, df_match_arca, df_falta_sistema, df_falta_arca = cruce_percepciones_pba(
        df_arca_dep, df_sistema_dep,
        tolerancia_importe=tolerancia_importe,
        score_nombre_min=score_nombre_min,
    )

    df_proveedores_nuevos = detectar_proveedores_nuevos(df_match_arca, umbral_nombre=max(score_nombre_min, 80))

    stats = {
        "match": len(df_match_arca),
        "faltante_sistema": len(df_falta_sistema),
        "faltante_arca": len(df_falta_arca),
        "proveedores_nuevos": len(df_proveedores_nuevos),
    }

    buf_reporte = generar_excel_percepciones_pba(
        df_match_sistema, df_match_arca, df_falta_sistema, df_falta_arca, df_proveedores_nuevos
    )

    return buf_reporte, stats
