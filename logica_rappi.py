import io
import re
import zipfile
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import pdfplumber


# ─────────────────────────────────────────────
# IMPORTAR LIQUIDACIONES (múltiples Excel)
# ─────────────────────────────────────────────

def _leer_nombre_tienda(wb):
    """Hoja 'Detalle': encabezados en la fila 2, solo se necesita la fila 3
    (una orden alcanza para identificar la tienda de esta liquidación)."""
    if 'Detalle' not in wb.sheetnames:
        return None
    ws_det = wb['Detalle']
    headers = [c.value for c in ws_det[2]]
    if 'Nombre de la tienda' not in headers:
        return None
    idx = headers.index('Nombre de la tienda')
    fila_datos = [c.value for c in ws_det[3]]
    return fila_datos[idx] if idx < len(fila_datos) else None


def _determinar_reporte(nombre_aliado, nombre_tienda):
    aliado = (nombre_aliado or '').strip().upper()
    tienda = (nombre_tienda or '').strip().upper()
    if 'ENTRETENIMIENTOS AVELLANEDA' in aliado:
        return 'Dean' if 'DEAN & DENNYS' in tienda else 'HIO'
    if 'PASEO RONDA' in aliado:
        return 'Atalaya' if 'ATALAYA' in tienda else 'HIO'
    return 'HIO'


def importar_liquidaciones(archivos):
    liquidaciones = []
    for archivo in archivos:
        wb = openpyxl.load_workbook(io.BytesIO(archivo.read()), data_only=True)
        ws = wb['Resumen']

        id_pago            = ws.cell(row=7, column=4).value
        inicio_periodo_liq = ws.cell(row=3, column=4).value
        fin_periodo_liq    = ws.cell(row=4, column=4).value
        nombre_aliado      = ws.cell(row=6, column=4).value
        nombre_tienda      = _leer_nombre_tienda(wb)

        merged_ranges = list(ws.merged_cells.ranges)
        for rango in merged_ranges:
            if rango.min_col == 2:
                valor = ws.cell(rango.min_row, 2).value
                ws.unmerge_cells(str(rango))
                for fila in range(rango.min_row, rango.max_row + 1):
                    ws.cell(fila, 2).value = valor

        registros = []
        for row in ws.iter_rows(min_row=10, values_only=True):
            if not any(c is not None for c in row):
                continue
            registros.append({
                'Grupo'  : row[1],
                'Valores': row[2],
                'Total'  : row[3],
            })

        df = pd.DataFrame(registros)
        df['Grupo']   = df['Grupo'].ffill()
        df['Valores'] = df['Valores'].ffill()
        df['Total']   = pd.to_numeric(df['Total'], errors='coerce')
        df = df.dropna(subset=['Valores']).reset_index(drop=True)

        fila_total = df[df['Grupo'] == 'Valor total a transferir']
        valor_total_transferir = round(fila_total['Total'].iloc[0], 2) if not fila_total.empty else None

        fila_venta_bruta = df[df['Valores'] == 'SUM of Venta Bruta']
        venta_bruta = round(fila_venta_bruta['Total'].iloc[0], 2) if not fila_venta_bruta.empty else None

        fila_desc_prod = df[df['Valores'] == 'SUM of Descuento de Producto']
        descuento_producto = round(fila_desc_prod['Total'].iloc[0], 2) if not fila_desc_prod.empty else None

        liquidaciones.append({
            'id_pago':                id_pago,
            'inicio_periodo_liq':     inicio_periodo_liq,
            'fin_periodo_liq':        fin_periodo_liq,
            'valor_total_transferir': valor_total_transferir,
            'periodo_str':            f'{inicio_periodo_liq} al {fin_periodo_liq}',
            'reporte':                _determinar_reporte(nombre_aliado, nombre_tienda),
            'venta_bruta':            venta_bruta,
            'descuento_producto':     descuento_producto,
            'df':                     df,
        })
    return liquidaciones


def depurar_liquidaciones(liquidaciones):
    excluir   = ['SUM of Venta Bruta', 'SUM of Descuento de Producto', 'SUM of Subtotal antes de impuestos']
    conservar = ['SUM of Uso y alquiler de plataforma Rappi', 'SUM of Tarifa transaccional']

    for liq in liquidaciones:
        df = liq['df']
        df = df[~df['Valores'].isin(excluir)]
        df = df[df['Grupo'] != 'Valor total a transferir']
        df = df[df['Total'].notna()]
        df = df[(df['Total'] != 0) | (df['Valores'].isin(conservar))]
        df['Total'] = df['Total'].round(2)
        liq['df'] = df.reset_index(drop=True)
    return liquidaciones


def construir_resumen_extracto(liquidaciones):
    filas = [
        {
            'Periodo':               liq.get('periodo_str'),
            'Reporte':               liq.get('reporte'),
            'Venta Bruta':           liq.get('venta_bruta'),
            'Descuento de Producto': liq.get('descuento_producto'),
            'Venta Neta':            liq.get('valor_total_transferir'),
        }
        for liq in liquidaciones
    ]
    return pd.DataFrame(filas, columns=['Periodo', 'Reporte', 'Venta Bruta', 'Descuento de Producto', 'Venta Neta'])


# ─────────────────────────────────────────────
# IMPORTAR FACTURAS (PDFs)
# ─────────────────────────────────────────────

def _nro_factura_desde_nombre_archivo(nombre_archivo):
    """Cuentas al día: la factura sale sin número (línea vacía / " " en el PDF).
    En esos casos se usa el identificador que trae el nombre del archivo
    (p.ej. 'VARACC21023015_27_al_31_07.pdf' -> 'VARACC21023015')."""
    if not nombre_archivo:
        return None
    stem = re.sub(r'\.pdf$', '', nombre_archivo, flags=re.IGNORECASE)
    stem = stem.split('_al_')[0]
    stem = re.sub(r'_\d+$', '', stem)
    return stem or None


def _parse_factura(file_obj, nombre_archivo=None):
    with pdfplumber.open(io.BytesIO(file_obj.read())) as pdf:
        words = pdf.pages[0].extract_words()

    COL = {'cod':(36,115), 'cant':(115,158), 'desc':(158,379),
           'imp':(379,440), 'punit':(400,500), 'ptotal':(500,600)}
    IMP = {'codigo':(200,257), 'alicuota':(316,363), 'impuesto':(363,520)}

    def x_in(w, rng): return rng[0] <= w['x0'] < rng[1]

    header_top = next(w['top'] for w in words if w['text']=='Cod.' and 50<w['x0']<70)
    obs_top    = next(w['top'] for w in words if w['text']=='Observaciones' and w['x0']<120)
    det_top    = next(w['top'] for w in words if w['text']=='Código' and w['x0']>200)
    total_top  = next(w['top'] for w in words if w['text']=='Total' and 460<w['x0']<510)

    # Número de factura
    nro_factura = None
    for i, w in enumerate(words):
        if w['text'] == 'N°:' and w['x0'] > 450:
            if i + 1 < len(words):
                candidato = words[i + 1]
                # El valor debe estar en la misma línea que "N°:"; si no hay
                # nada ahí, el siguiente word en orden de lectura es la
                # etiqueta "Fecha:" de la línea de abajo, no un número.
                if abs(candidato['top'] - w['top']) < 3:
                    nro_factura = candidato['text']
            break

    # Cuentas al día: la factura puede no traer número (línea vacía en el PDF).
    # En ese caso se usa el identificador del nombre del archivo.
    if not nro_factura or not nro_factura.strip():
        nro_factura = _nro_factura_desde_nombre_archivo(nombre_archivo)

    # PID
    pid = None
    for i, w in enumerate(words):
        if w['text'] == 'PID:':
            if i + 1 < len(words):
                pid = int(words[i + 1]['text'])
            break
        if w['text'] == 'Lot:':
            if i + 1 < len(words):
                pid = int(words[i + 1]['text'])
            break

    # Fecha Factura
    fecha_factura = None
    for i, w in enumerate(words):
        if w['text'] == 'Fecha:' and w['x0'] > 450:
            if i + 1 < len(words):
                fecha_factura = words[i + 1]['text']
            break

    # Inicio y Fin de Periodo (formato YYYY-MM-DD en zona descripción)
    date_pattern   = re.compile(r'^\d{4}-\d{2}-\d{2}$')
    dates_found    = [w['text'] for w in words if date_pattern.match(w['text']) and 150 < w['x0'] < 400]
    inicio_periodo = dates_found[0] if len(dates_found) >= 1 else None
    fin_periodo    = dates_found[1] if len(dates_found) >= 2 else None

    tbl_words = [w for w in words if header_top < w['top'] < obs_top]

    ptotal_anchors = sorted(
        [w for w in tbl_words if x_in(w, COL['ptotal']) and re.match(r'^-?\d', w['text'])],
        key=lambda w: w['top']
    )
    tops = [w['top'] for w in ptotal_anchors] + [obs_top]

    def collect_col(row_words, col_key):
        ws_ = sorted([w for w in row_words if x_in(w, COL[col_key])], key=lambda w:(w['top'],w['x0']))
        lines, cur_t, cur = [], None, []
        for w in ws_:
            if cur_t is None or abs(w['top']-cur_t) > 4:
                if cur: lines.append(' '.join(cur))
                cur, cur_t = [w['text']], w['top']
            else:
                cur.append(w['text'])
        if cur: lines.append(' '.join(cur))
        return ' '.join(lines).strip()

    rows = []
    for i, anchor in enumerate(ptotal_anchors):
        t_start = tops[i] - 3
        t_end   = tops[i+1] - 3
        rw = [w for w in tbl_words if t_start <= w['top'] < t_end]
        cant_vals = [w['text'] for w in rw if x_in(w, COL['cant']) and re.match(r'^\d', w['text'])]
        rows.append({
            'Cod':         collect_col(rw, 'cod'),
            'Cant':        cant_vals[0] if cant_vals else None,
            'Descripción': collect_col(rw, 'desc'),
            'Imp':         collect_col(rw, 'imp'),
            'P.Unit':      collect_col(rw, 'punit'),
            'P.Total':     anchor['text'],
        })
    df_main = pd.DataFrame(rows)

    imp_words = [w for w in words if det_top < w['top'] < total_top]
    ali_anchors = sorted(
        [w for w in imp_words if x_in(w, IMP['alicuota']) and '%' in w['text']],
        key=lambda w: w['top']
    )
    ali_tops = [w['top'] for w in ali_anchors] + [total_top]

    def collect_imp(row_words, col_key):
        ws_ = sorted([w for w in row_words if x_in(w, IMP[col_key])], key=lambda w:(w['top'],w['x0']))
        lines, cur_t, cur = [], None, []
        for w in ws_:
            if cur_t is None or abs(w['top']-cur_t) > 4:
                if cur: lines.append(' '.join(cur))
                cur, cur_t = [w['text']], w['top']
            else:
                cur.append(w['text'])
        if cur: lines.append(' '.join(cur))
        return ' '.join(lines).strip()

    imp_rows = []
    for i, anchor in enumerate(ali_anchors):
        t_start = ali_tops[i-1]+1 if i > 0 else det_top
        t_end   = ali_tops[i+1]-1
        rw = [w for w in imp_words if t_start <= w['top'] <= t_end]
        imp_rows.append({
            'Cod':         'Detalle de Impuesto',
            'Cant':        None,
            'Descripción': collect_imp(rw, 'codigo'),
            'Imp':         collect_imp(rw, 'alicuota'),
            'P.Unit':      collect_imp(rw, 'impuesto'),
            'P.Total':     collect_imp(rw, 'impuesto'),
        })

    df = pd.concat([df_main, pd.DataFrame(imp_rows)], ignore_index=True)

    primera_cod = df['Cod'].iloc[0].replace(' ', '').lower()
    tipo = 'publicidad' if 'serviciosdepublicidad' in primera_cod else 'servicios'

    # Total de la factura (misma fila que la etiqueta "Total", a la derecha)
    total_words = sorted(
        [w for w in words if abs(w['top'] - total_top) < 3 and w['x0'] > 510],
        key=lambda w: w['x0']
    )
    total_factura = ' '.join(w['text'] for w in total_words) or None

    return {
        'nro_factura':    nro_factura,
        'pid':            pid,
        'tipo':           tipo,
        'fecha_factura':  fecha_factura,
        'inicio_periodo': inicio_periodo,
        'fin_periodo':    fin_periodo,
        'total_factura':  total_factura,
        'df':             df,
    }


def importar_facturas(archivos_pdf):
    facturas = []
    for f in archivos_pdf:
        factura = _parse_factura(f, getattr(f, 'name', None))
        facturas.append(factura)
    return facturas


def depurar_facturas(facturas):
    for factura in facturas:
        df = factura['df']
        for col in ['P.Unit', 'P.Total']:
            df[col] = (
                df[col]
                .astype(str)
                .str.replace('$', '', regex=False)
                .str.replace('.', '', regex=False)
                .str.replace(',', '.', regex=False)
                .pipe(pd.to_numeric, errors='coerce')
                .round(2)
            )
        df = df[df['P.Total'].abs() > 0.5].reset_index(drop=True)
        factura['df'] = df

        total_str = factura.get('total_factura')
        valor = (
            str(total_str)
            .replace('$', '')
            .replace('.', '')
            .replace(',', '.')
            .strip()
        ) if total_str else None
        try:
            factura['total_factura'] = round(float(valor), 2) if valor else None
        except ValueError:
            factura['total_factura'] = None
    return facturas


def construir_resumen_facturas(facturas):
    """Igual que el resumen de asignaciones pero sin cruzar contra liquidaciones
    (por lo tanto sin columna 'id_pago'): solo lo que se puede leer directamente
    de las facturas."""
    filas = []
    for factura in facturas:
        fin_periodo   = factura.get('fin_periodo')
        fecha_factura = factura.get('fecha_factura')
        filas.append({
            'nro_factura':    factura['nro_factura'],
            'inicio_periodo': factura.get('inicio_periodo'),
            'fin_periodo':    datetime.strptime(fin_periodo, '%Y-%m-%d').date() if fin_periodo else None,
            'fecha_factura':  datetime.strptime(fecha_factura, '%d/%m/%Y').date() if fecha_factura else None,
            'total_factura':  factura.get('total_factura'),
        })
    columnas = ['nro_factura', 'inicio_periodo', 'fin_periodo', 'fecha_factura', 'total_factura']
    return pd.DataFrame(filas, columns=columnas)


# ─────────────────────────────────────────────
# ASIGNACIÓN: facturas → liquidaciones
# ─────────────────────────────────────────────

def _mismo_periodo(factura, liq):
    """True/False si se puede comparar; None si a alguno le falta el dato."""
    f_ini, f_fin = factura.get('inicio_periodo'), factura.get('fin_periodo')
    l_ini, l_fin = liq.get('inicio_periodo_liq'), liq.get('fin_periodo_liq')
    if not (f_ini and f_fin and l_ini and l_fin):
        return None
    return f_ini == l_ini and f_fin == l_fin


def _elegir_liquidacion(factura, candidatos):
    """Entre liquidaciones que ya matchean por monto, desambigua por período
    de venta. Si hay más de un candidato y no se puede confirmar cuál es por
    período, no elige ninguno (mejor no asignar que asignar mal)."""
    if len(candidatos) <= 1:
        return candidatos[0] if candidatos else None
    exactos = [liq for liq in candidatos if _mismo_periodo(factura, liq) is True]
    return exactos[0] if len(exactos) == 1 else None


def asignar_facturas(facturas, liquidaciones):
    TOLERANCIA = 1
    advertencias = []

    for liq in liquidaciones:
        liq['facturas_publicidad'] = []
        liq['facturas_servicios']  = []
        liq['nros_factura']        = []

    # Liquidaciones ya usadas para una factura de cada tipo: una vez asignada
    # una no vuelve a ofrecerse, así facturas repetidas en monto (misma
    # campaña, semanas distintas) no terminan todas pisando la primera.
    ocupadas_servicios  = set()
    ocupadas_publicidad = set()

    for factura in facturas:
        df_fac  = factura['df']
        tipo    = factura['tipo']
        nro_fac = factura['nro_factura']

        if tipo == 'servicios':
            mask_bank = df_fac['Cod'].str.replace(' ','').str.lower().str.contains('ar-bankfeerestaurantes', na=False)
            mask_com  = df_fac['Cod'].str.replace(' ','').str.lower().str.contains('ar-comisionesrestaurant', na=False)
            val_bank  = abs(df_fac.loc[mask_bank, 'P.Total'].values[0]) if mask_bank.any() else None
            val_com   = abs(df_fac.loc[mask_com,  'P.Total'].values[0]) if mask_com.any()  else None

            candidatos = []
            for liq in liquidaciones:
                if id(liq) in ocupadas_servicios:
                    continue
                df_liq   = liq['df']
                mask_tar = df_liq['Valores'].str.contains('SUM of Tarifa transaccional', na=False)
                mask_uso = df_liq['Valores'].str.contains('SUM of Uso y alquiler de plataforma Rappi', na=False)
                val_tar  = abs(df_liq.loc[mask_tar, 'Total'].values[0]) if mask_tar.any() else None
                val_uso  = abs(df_liq.loc[mask_uso, 'Total'].values[0]) if mask_uso.any() else None

                checks = []
                if val_bank is not None and val_tar is not None:
                    checks.append(abs(val_bank - val_tar) <= TOLERANCIA)
                if val_com is not None and val_uso is not None:
                    checks.append(abs(val_com - val_uso) <= TOLERANCIA)

                if checks and all(checks):
                    candidatos.append(liq)

            liq_elegida = _elegir_liquidacion(factura, candidatos)
            if liq_elegida is not None:
                liq_elegida['facturas_servicios'].append(factura)
                liq_elegida['nros_factura'].append(nro_fac)
                ocupadas_servicios.add(id(liq_elegida))
            elif len(candidatos) > 1:
                advertencias.append(
                    f"Factura {nro_fac} ({tipo}) coincide en monto con {len(candidatos)} liquidaciones "
                    "y no se pudo distinguir por período de venta"
                )
            else:
                advertencias.append(f"Factura {nro_fac} ({tipo}) no encontró liquidación")

        elif tipo == 'publicidad':
            mask_pub = df_fac['Cod'].str.replace(' ','').str.lower().str.contains('serviciosdepublicidadindexaccionrappi', na=False)
            suma_pub = abs(df_fac.loc[mask_pub, 'P.Total'].sum()) if mask_pub.any() else None

            candidatos = []
            for liq in liquidaciones:
                if id(liq) in ocupadas_publicidad:
                    continue
                df_liq   = liq['df']
                mask_ads = df_liq['Valores'].str.contains('SUM of Cuota de RappiAds', na=False)
                val_ads  = abs(df_liq.loc[mask_ads, 'Total'].values[0]) if mask_ads.any() else None

                if suma_pub is not None and val_ads is not None and abs(suma_pub - val_ads) <= TOLERANCIA:
                    candidatos.append(liq)

            liq_elegida = _elegir_liquidacion(factura, candidatos)
            if liq_elegida is not None:
                liq_elegida['facturas_publicidad'].append(factura)
                liq_elegida['nros_factura'].append(nro_fac)
                ocupadas_publicidad.add(id(liq_elegida))
            elif len(candidatos) > 1:
                advertencias.append(
                    f"Factura {nro_fac} ({tipo}) coincide en monto con {len(candidatos)} liquidaciones "
                    "y no se pudo distinguir por período de venta"
                )
            else:
                advertencias.append(f"Factura {nro_fac} ({tipo}) no encontró liquidación")

    return liquidaciones, advertencias


# ─────────────────────────────────────────────
# CRUCES INTERNOS
# ─────────────────────────────────────────────

def _cruce_publicidad(df_publicidad, df_liquidacion):
    TOLERANCIA = 1

    if df_publicidad.empty:
        return pd.DataFrame(), pd.DataFrame(), df_liquidacion.copy().reset_index(drop=True)

    disponibles_liq = list(df_liquidacion.index)
    idx_match_pub, idx_match_liq = [], []

    mask_pub = df_publicidad['Cod'].str.replace(' ','').str.lower().str.contains('serviciosdepublicidadindexaccionrappi', na=False)
    idx_pub  = df_publicidad[mask_pub].index.tolist()

    if idx_pub:
        suma_pub = df_publicidad.loc[idx_pub, 'P.Total'].sum()
        for j in disponibles_liq:
            if abs(abs(suma_pub) - abs(df_liquidacion.loc[j, 'Total'])) <= TOLERANCIA:
                idx_match_pub.extend(idx_pub)
                idx_match_liq.append(j)
                disponibles_liq.remove(j)
                break

    for i in df_publicidad[~mask_pub].index.tolist():
        val = abs(df_publicidad.loc[i, 'P.Total'])
        for j in disponibles_liq:
            if abs(val - abs(df_liquidacion.loc[j, 'Total'])) <= TOLERANCIA:
                idx_match_pub.append(i)
                idx_match_liq.append(j)
                disponibles_liq.remove(j)
                break

    match_liq = df_liquidacion.loc[idx_match_liq].reset_index(drop=True)
    falta_liq = df_liquidacion.drop(index=idx_match_liq).reset_index(drop=True)
    match_pub = df_publicidad.loc[idx_match_pub].reset_index(drop=True)
    return match_pub, match_liq, falta_liq


def _cruce_servicios(df_servicios, falta_pub):
    TOLERANCIA = 1

    if df_servicios.empty:
        return pd.DataFrame(), pd.DataFrame(), falta_pub.copy().reset_index(drop=True)

    disponibles_srv = list(df_servicios.index)
    disponibles_liq = list(falta_pub.index)
    idx_match_srv, idx_match_liq = [], []

    for i in list(disponibles_liq):
        val = abs(falta_pub.loc[i, 'Total'])
        for j in disponibles_srv:
            if abs(val - abs(df_servicios.loc[j, 'P.Total'])) <= TOLERANCIA:
                idx_match_liq.append(i)
                idx_match_srv.append(j)
                disponibles_liq.remove(i)
                disponibles_srv.remove(j)
                break

    # Pares base + "Descuento por inversión de Rappi DAR" que Rappi factura
    # como una única línea neta en el Detalle de Impuesto
    # (p.ej. IVA Uso y alquiler, o Percepción IIBB por provincia).
    PARES_DAR = [
        ('SUM of IVA Uso y alquiler de plataforma Rappi',
         'SUM of Descuento por inversión de Rappi a aplicar sobre el IVA Uso y alquiler de plataforma Rappi DAR'),
        ('SUM of Percepcion',
         'SUM of Descuento por inversión de Rappi a aplicar sobre la Percepción de BA DAR'),
        ('SUM of Percepcion Cordoba',
         'SUM of Descuento por inversión de Rappi a aplicar sobre la Percepción de Córdoba DAR'),
        ('SUM of Percepción Tucuman',
         'SUM of Descuento por inversión de Rappi a aplicar sobre la PERCEPCIÓN DE TUCUMAN DAR'),
        ('SUM of Percepción Corrientes',
         'SUM of Descuento por inversión de Rappi a aplicar sobre la PERCEPCIÓN DE CORRIENTES DAR'),
    ]

    for base_label, dsc_label in PARES_DAR:
        resto_liq = falta_pub.loc[disponibles_liq].copy()
        idx_par = resto_liq[resto_liq['Valores'].isin([base_label, dsc_label])].index.tolist()
        if not idx_par:
            continue

        suma_par = resto_liq.loc[idx_par, 'Total'].sum()
        for j in disponibles_srv:
            if abs(abs(suma_par) - abs(df_servicios.loc[j, 'P.Total'])) <= TOLERANCIA:
                idx_match_liq.extend(idx_par)
                idx_match_srv.append(j)
                disponibles_liq = [i for i in disponibles_liq if i not in idx_par]
                disponibles_srv.remove(j)
                break

    resto_liq  = falta_pub.loc[disponibles_liq].copy()
    grupos_sum = resto_liq.groupby('Grupo')['Total'].sum()
    idx_srv2, idx_grupos = [], []

    for grupo, suma in grupos_sum.items():
        for j in disponibles_srv:
            if abs(abs(suma) - abs(df_servicios.loc[j, 'P.Total'])) <= TOLERANCIA:
                idx_srv2.append(j)
                idx_grupos.append(grupo)
                disponibles_srv.remove(j)
                break

    idx_liq_grupo = resto_liq[resto_liq['Grupo'].isin(idx_grupos)].index.tolist()
    todos_liq     = idx_match_liq + idx_liq_grupo
    todos_srv     = idx_match_srv + idx_srv2

    match_liq = falta_pub.loc[todos_liq].reset_index(drop=True)
    falta_fac = falta_pub.drop(index=todos_liq).reset_index(drop=True)
    return df_servicios.loc[todos_srv].reset_index(drop=True), match_liq, falta_fac


# ─────────────────────────────────────────────
# PROCESAMIENTO POR LIQUIDACIÓN
# ─────────────────────────────────────────────

def procesar_liquidaciones(liquidaciones):
    COLUMNAS_FAC = ['Cod', 'Cant', 'Descripción', 'Imp', 'P.Unit', 'P.Total']
    resumen_rows = []

    for liq in liquidaciones:
        df_liq       = liq['df'].copy()
        facturas_pub = liq['facturas_publicidad']
        facturas_srv = liq['facturas_servicios']
        all_match_liq = []
        falta_actual  = df_liq.copy()

        if not facturas_pub:
            _, _, falta_actual = _cruce_publicidad(pd.DataFrame(columns=COLUMNAS_FAC), falta_actual)
        else:
            for factura in facturas_pub:
                df_pub = factura['df'].copy().reset_index(drop=True)
                _, match_liq, falta_actual = _cruce_publicidad(df_pub, falta_actual)
                if not match_liq.empty:
                    match_liq = match_liq.copy()
                    match_liq['nro_factura'] = factura['nro_factura']
                    all_match_liq.append(match_liq)

        if not facturas_srv:
            _, _, falta_actual = _cruce_servicios(pd.DataFrame(columns=COLUMNAS_FAC), falta_actual)
        else:
            for factura in facturas_srv:
                df_srv = factura['df'].copy().reset_index(drop=True)
                _, match_liq, falta_actual = _cruce_servicios(df_srv, falta_actual)
                if not match_liq.empty:
                    match_liq = match_liq.copy()
                    match_liq['nro_factura'] = factura['nro_factura']
                    all_match_liq.append(match_liq)

        liq['match_liquidacion'] = (
            pd.concat(all_match_liq, ignore_index=True)
            if all_match_liq
            else pd.DataFrame(columns=list(df_liq.columns) + ['nro_factura'])
        )
        liq['falta_factura'] = falta_actual.copy().reset_index(drop=True)

        # Se invierte el signo de 'Total': la liquidación lo expresa desde la
        # perspectiva de Rappi, acá lo pasamos a la perspectiva de la empresa.
        if not liq['match_liquidacion'].empty:
            liq['match_liquidacion']['Total'] = -liq['match_liquidacion']['Total']
        if not liq['falta_factura'].empty:
            liq['falta_factura']['Total'] = -liq['falta_factura']['Total']

        for factura in facturas_pub + facturas_srv:
            fin_periodo   = factura.get('fin_periodo')
            fecha_factura = factura.get('fecha_factura')
            resumen_rows.append({
                'nro_factura':    factura['nro_factura'],
                'id_pago':        liq['id_pago'],
                'inicio_periodo': factura.get('inicio_periodo'),
                'fin_periodo':    datetime.strptime(fin_periodo, '%Y-%m-%d').date() if fin_periodo else None,
                'fecha_factura':  datetime.strptime(fecha_factura, '%d/%m/%Y').date() if fecha_factura else None,
                'total_factura':  factura.get('total_factura'),
            })

    df_resumen = (
        pd.DataFrame(resumen_rows)
        if resumen_rows
        else pd.DataFrame(columns=['nro_factura', 'id_pago', 'inicio_periodo', 'fin_periodo', 'fecha_factura', 'total_factura'])
    )
    return liquidaciones, df_resumen


# ─────────────────────────────────────────────
# CUADRO DE CONCEPTOS (una hoja, agrupada por liquidación)
# ─────────────────────────────────────────────

def _formatear_fecha_liq(valor):
    if not valor:
        return ''
    try:
        valor = datetime.strptime(str(valor), '%Y-%m-%d')
    except ValueError:
        return str(valor)
    return valor.strftime('%d/%m/%Y')


def construir_cuadro_conceptos(liquidaciones):
    filas = []
    for liq in liquidaciones:
        inicio = _formatear_fecha_liq(liq.get('inicio_periodo_liq'))
        fin    = _formatear_fecha_liq(liq.get('fin_periodo_liq'))
        filas.append((liq.get('reporte'), None))
        filas.append((f'Liquidacion {inicio} a {fin}', None))
        filas.append(('Valor total a transferir', liq.get('valor_total_transferir')))

        for _, row in liq['falta_factura'].iterrows():
            concepto = re.sub(r'^SUM of ', '', str(row['Valores']), flags=re.IGNORECASE)
            filas.append((concepto, row['Total']))

        for factura in liq['facturas_publicidad'] + liq['facturas_servicios']:
            filas.append((factura['nro_factura'], factura.get('total_factura')))

        filas.append(None)
        filas.append(None)

    while filas and filas[-1] is None:
        filas.pop()

    return filas


# ─────────────────────────────────────────────
# FUNCIÓN PRINCIPAL → devuelve ZIP en memoria
# ─────────────────────────────────────────────

def correr_rappi(archivos_liq, archivos_pdf):
    liquidaciones = importar_liquidaciones(archivos_liq)
    liquidaciones = depurar_liquidaciones(liquidaciones)

    df_resumen_extracto = construir_resumen_extracto(liquidaciones)

    facturas = importar_facturas(archivos_pdf)
    facturas = depurar_facturas(facturas)

    liquidaciones, advertencias = asignar_facturas(facturas, liquidaciones)
    liquidaciones, df_resumen   = procesar_liquidaciones(liquidaciones)

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for liq in liquidaciones:
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                liq['match_liquidacion'].to_excel(writer, sheet_name='match_liquidacion', index=False)
                liq['falta_factura'].to_excel(writer, sheet_name='falta_factura', index=False)
            buf.seek(0)
            zf.writestr(f"liquidacion_{liq['id_pago']}.xlsx", buf.read())

        buf_res = io.BytesIO()
        with pd.ExcelWriter(buf_res, engine='openpyxl') as writer:
            df_resumen.to_excel(writer, index=False, sheet_name='Sheet1')
            ws = writer.sheets['Sheet1']
            formatos_col = {
                'fin_periodo':   'DD/MM/YYYY',
                'fecha_factura': 'DD/MM/YYYY',
                'total_factura': '$ #,##0.00',
            }
            for col_name, formato in formatos_col.items():
                if col_name in df_resumen.columns:
                    col_letter = get_column_letter(df_resumen.columns.get_loc(col_name) + 1)
                    for cell in ws[col_letter][1:]:
                        cell.number_format = formato
        buf_res.seek(0)
        zf.writestr('resumen_asignaciones.xlsx', buf_res.read())

        cuadro_filas = construir_cuadro_conceptos(liquidaciones)
        wb_cuadro = openpyxl.Workbook()
        ws_cuadro = wb_cuadro.active
        ws_cuadro.title = 'cuadro_conceptos'
        ws_cuadro.append(['Concepto', 'Monto'])
        for fila in cuadro_filas:
            ws_cuadro.append(list(fila) if fila is not None else [])
        for row in ws_cuadro.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '$ #,##0.00'
        buf_cuadro = io.BytesIO()
        wb_cuadro.save(buf_cuadro)
        buf_cuadro.seek(0)
        zf.writestr('cuadro_conceptos.xlsx', buf_cuadro.read())

        buf_extracto = io.BytesIO()
        with pd.ExcelWriter(buf_extracto, engine='openpyxl') as writer:
            df_resumen_extracto.to_excel(writer, index=False, sheet_name='resumen_extracto')
            ws_extracto = writer.sheets['resumen_extracto']
            for col_name in ['Venta Bruta', 'Descuento de Producto', 'Venta Neta']:
                col_letter = get_column_letter(df_resumen_extracto.columns.get_loc(col_name) + 1)
                for cell in ws_extracto[col_letter][1:]:
                    cell.number_format = '$ #,##0.00'
        buf_extracto.seek(0)
        zf.writestr('resumen_extracto.xlsx', buf_extracto.read())

    zip_buf.seek(0)

    stats = {
        'n_liquidaciones': len(liquidaciones),
        'n_facturas':      len(facturas),
        'advertencias':    advertencias,
        'detalle': [
            {
                'id_pago':  liq['id_pago'],
                'match':    len(liq['match_liquidacion']),
                'falta':    len(liq['falta_factura']),
                'facturas': liq['nros_factura'],
            }
            for liq in liquidaciones
        ]
    }
    return zip_buf, stats


def correr_rappi_resumen_facturas(archivos_pdf):
    """Modo simple: resume las facturas (mismo cálculo que en el cruce) sin
    necesitar liquidaciones ni cruzarlas contra nada."""
    facturas = importar_facturas(archivos_pdf)
    facturas = depurar_facturas(facturas)
    df_resumen_facturas = construir_resumen_facturas(facturas)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df_resumen_facturas.to_excel(writer, index=False, sheet_name='resumen_facturas')
        ws = writer.sheets['resumen_facturas']
        formatos_col = {
            'fin_periodo':   'DD/MM/YYYY',
            'fecha_factura': 'DD/MM/YYYY',
            'total_factura': '$ #,##0.00',
        }
        for col_name, formato in formatos_col.items():
            if col_name in df_resumen_facturas.columns:
                col_letter = get_column_letter(df_resumen_facturas.columns.get_loc(col_name) + 1)
                for cell in ws[col_letter][1:]:
                    cell.number_format = formato
    buf.seek(0)

    stats = {'n_facturas': len(facturas)}
    return buf, stats
